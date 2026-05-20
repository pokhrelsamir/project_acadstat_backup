from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from .forms import (ResultForm, GradeScaleForm, TeacherEvaluationForm, InvoiceQuickCreateForm,
                    SupportTicketForm, CertificateGenerateForm, LessonPlanForm, SyllabusCoverageForm,
                    GradingRubricForm, RubricCriterionForm, ExamSeatingPlanForm, QuestionBankForm,
                    QuestionPaperTemplateForm, ReminderForm, SubscriptionForm, InvoiceForm,
                    CertificateTemplateForm, TicketCommentForm, ParentUserForm,
                    UserRoleForm, UserProfileForm, ResultPublishSessionForm, ResultSessionRemarkForm)
from .models import (
    Result, Student, Subject, Teacher, CourseMaterial, Notification,
    GradeScale, TeacherEvaluation, ActivityLog,
    LessonPlan, SyllabusCoverage,
    GradingRubric, RubricCriterion, RubricScoreEntry, RubricTemplate,
    OnlineExam, Question, ExamAttempt, ExamAnswer,
    ExamSeatingPlan, SeatAllocation,
    TeacherLeave,
    AssignmentSubmission,
    Subscription, Invoice, Certificate, CertificateTemplate,
    APIKey, WebhookEndpoint, WebhookDeliveryLog,
    SupportTicket, TicketComment, SSOProvider, SystemConfig,
    UserProfile, LicenseKey, SystemBackup, UserRole, ParentUser,
    ResultPublishSession, ResultSessionEntry,
    Parent,
)
from core.models import (
    Assignment, Attendance, Exam, Fee, Announcement, StudentNote, MLPrediction,
    EducationLevel, Department, Semester, AcademicYear,
)
# missing new models
from core.models import (
    QRCodeAttendance, FaceAttendance, GPSAttendance, LateEntryLog, AttendanceReport,
    PlagiarismReport, MCQAutoGrade, RecheckingRequest,
    LiveClass, LiveClassAttendance, LiveClassRecording,
    DisciplineRecord, ParticipationScore, BehavioralAnalytics,
    SkillCourse, StudentSkill, Achievement, StudentAchievement, Leaderboard, LeaderboardEntry, DailyStreak,
    StudyGroup, StudyGroupMessage, SharedNote,
    StudentGoal, StudentIDCard, VideoLecture,
    CareerPortalListing, StudentApplication,
    AIQuiz, AILessonPlan,
    ParentNotification, ParentMeeting,
    BusRoute, BusStop, StudentTransportCard, TransportMonitor,
    HostelRoom, HostelAllocation,
    AIRecommendation, AIStudentReport,
    IoTDevice, SmartEnergyLog,
    Institution, InstitutionUser,
    Timetable, TimetableEntry, TimetableConflict, ExamRoom,
)
from django.db.models import Q, Sum, Avg, Count, F, FloatField, ExpressionWrapper, Case, When, Max, Min
from django.http import FileResponse
from django.conf import settings
from django.template import loader
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
try:
    import face_recognition
    HAS_FACE_RECOGNITION = True
except ImportError:
    HAS_FACE_RECOGNITION = False
import os
from core.models import UserProfile, LicenseKey, SystemBackup, SystemConfig
from core.models import UserRole, ParentUser, ResultPublishSession, ResultSessionEntry
from django.db.models import Q, Sum, Avg, Count, F, FloatField, ExpressionWrapper, Case, When
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User
from functools import wraps
import json
import qrcode
import csv
import io
from io import BytesIO
import base64
import openpyxl
from openpyxl.utils import get_column_letter
from django.contrib import messages
from datetime import date, datetime
from django.utils import timezone
from core import export_utils
from core.email_service import notify_student
from django.utils.crypto import get_random_string
import hmac
import hashlib
import uuid as _uuid


# Helper functions for teacher authentication
def get_teacher_profile(user):
    """Get teacher profile for a user, or None if not a teacher"""
    try:
        return Teacher.objects.get(user=user, is_active=True)
    except Teacher.DoesNotExist:
        return None

def is_teacher(user):
    """Check if user is a teacher"""
    return get_teacher_profile(user) is not None

def teacher_required(view_func):
    """Decorator to require teacher access"""
    @login_required
    def wrapper(request, *args, **kwargs):
        teacher = get_teacher_profile(request.user)
        if not teacher:
            messages.error(request, "Access denied. Teacher account required.")
            return redirect('/dashboard/')
        request.teacher = teacher
        return view_func(request, *args, **kwargs)
    return wrapper


DEFAULT_PASSWORD = "password123"

def home_view(request):
    """Landing page - accessible without login"""
    return render(request, 'core/dashboard/home.html')


# 🔐 LOGIN VIEW
def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('/dashboard/')  # ✅ go to dashboard
        else:
            return render(request, 'core/registration/login.html', {
                'error': 'Invalid username or password'
            })

    return render(request, 'core/registration/login.html')


# 🚪 LOGOUT VIEW
def logout_view(request):
    logout(request)
    return redirect('/')


# 📊 DASHBOARD (MAIN PAGE AFTER LOGIN)
@login_required
def dashboard(request):
    # Check if user is a teacher
    teacher = get_teacher_profile(request.user)
    if teacher:
        return teacher_dashboard(request)

    # Check if user is a student (has username matching a student)
    username = request.user.username.lower().strip()

    # Try to find student by exact name match (case-insensitive)
    students = Student.objects.all()
    for student in students:
        if student.name.lower().strip() == username:
            # Redirect students to their own dashboard
            return student_dashboard(request, student.id)

    # Also try partial match
    for student in students:
        if username in student.name.lower() or student.name.lower() in username:
            return student_dashboard(request, student.id)

    # Admin dashboard - Get all results for calculations
    results = Result.objects.all()

    # Calculate average marks
    avg_marks = 0
    if results.exists():
        total = sum(r.marks_obtained for r in results)
        avg_marks = round(total / results.count(), 1)

    context = {
        'total_students': Student.objects.count(),
        'total_subjects': Subject.objects.count(),
        'total_marks': Result.objects.count(),
        'average_marks': avg_marks,
        'user': request.user,
        'is_admin': True,
    }
    return render(request, 'core/dashboard/dashboard.html', context)


# 🎓 STUDENT DASHBOARD (Limited Access)
@login_required
def student_dashboard(request, student_id=None):
    """Student dashboard with limited access - shows only their own data"""
    import qrcode
    from io import BytesIO
    import base64
    from core.notification_service import NotificationService
    
    # Get terminal filter from request (default: '1st')
    selected_terminal = request.GET.get('terminal', '1st')
    # Validate terminal choice
    valid_terminals = ['1st', '2nd', '3rd', 'Final']
    if selected_terminal not in valid_terminals:
        selected_terminal = '1st'
    
    # If student_id is provided, use that, otherwise find by username
    if student_id:
        student = get_object_or_404(Student, id=student_id)
    else:
        # Try to find student by username
        try:
            student = Student.objects.get(name=request.user.username)
        except Student.DoesNotExist:
            return render(request, 'core/dashboard/student_dashboard.html', {
                'error': 'Student profile not found. Please contact admin.'
            })
    
    # Set student on request for notification access
    request.student = student
    
    # Get student's own results
    student_results = Result.objects.filter(student=student).select_related('subject')
    
    # Filter results for selected terminal (for the subject marks table)
    terminal_results = student_results.filter(terminal=selected_terminal)
    
    # Calculate OWN STATS (CUMULATIVE across ALL terminals - for the "1246/2000" display)
    total_marks_obtained = sum(r.marks_obtained for r in student_results)
    total_marks = sum(r.total_marks for r in student_results)
    overall_percentage = (total_marks_obtained / total_marks * 100) if total_marks > 0 else 0
    
    # Get marks by terminal for chart (still all terminals for chart overview)
    terminal_data = {}
    for terminal in ['1st', '2nd', '3rd', 'Final']:
        terminal_results_all = student_results.filter(terminal=terminal)
        if terminal_results_all.exists():
            t_marks = sum(r.marks_obtained for r in terminal_results_all)
            t_total = sum(r.total_marks for r in terminal_results_all)
            t_percentage = (t_marks / t_total * 100) if t_total > 0 else 0
            terminal_data[terminal] = {
                'obtained': t_marks,
                'total': t_total,
                'percentage': round(t_percentage, 1)
            }
    
    # Get subject-wise performance for the SELECTED terminal only (for table display)
    subject_data = {}
    for result in terminal_results:
        if result.subject.name not in subject_data:
            subject_data[result.subject.name] = []
        percentage = (result.marks_obtained / result.total_marks * 100) if result.total_marks > 0 else 0
        subject_data[result.subject.name].append({
            'terminal': result.terminal,
            'percentage': round(percentage, 1),
            'marks': result.marks_obtained,
            'total': result.total_marks
        })
    
    # Calculate terminal-specific totals (for the filtered view)
    terminal_marks_obtained = sum(r.marks_obtained for r in terminal_results)
    terminal_total_marks = sum(r.total_marks for r in terminal_results)
    
    # Get notifications for the student
    unread_notifications = NotificationService.get_student_notifications(student, unread_only=True, limit=10)
    recent_notifications = NotificationService.get_student_notifications(student, unread_only=False, limit=5)
    total_unread = NotificationService.get_unread_count(student)
    
    context = {
        'student': student,
        'student_results': terminal_results,  # Only results for selected terminal
        'total_subjects': student_results.values('subject').distinct().count(),
        'total_marks_obtained': total_marks_obtained,  # Cumulative across all terminals
        'total_marks': total_marks,  # Cumulative total
        'overall_percentage': round(overall_percentage, 1),
        'terminal_marks_obtained': terminal_marks_obtained,  # For selected terminal only
        'terminal_total_marks': terminal_total_marks,  # For selected terminal only
        'terminal_data': terminal_data,
        'subject_data': subject_data,
        'terminal_data_json': json.dumps(terminal_data),
        'subject_data_json': json.dumps(subject_data),
        'selected_terminal': selected_terminal,
        'terminal_choices': ['1st', '2nd', '3rd', 'Final'],
        'user': request.user,
        'is_student': True,
        # Notifications
        'notifications': recent_notifications,
        'unread_notifications': unread_notifications,
        'total_unread': total_unread,
    }
    
    return render(request, 'core/dashboard/student_dashboard.html', context)


# 👨‍🏫 TEACHER DASHBOARD
@teacher_required
def teacher_dashboard(request):
    """Teacher dashboard with restricted access to their subjects"""
    teacher = request.teacher

    # Get teacher's subjects
    teacher_subjects = teacher.subjects.all()

    # Get students assigned to this teacher
    teacher_students = teacher.students.all()

    # Get results for teacher's subjects and students
    teacher_results = Result.objects.filter(
        subject__in=teacher_subjects,
        student__in=teacher_students
    ).select_related('student', 'subject')

    # Calculate stats for teacher's data only
    avg_marks = 0
    if teacher_results.exists():
        total = sum(r.marks_obtained for r in teacher_results)
        avg_marks = round(total / teacher_results.count(), 1)

    context = {
        'teacher': teacher,
        'total_students': teacher_students.count(),
        'total_subjects': teacher_subjects.count(),
        'total_marks': teacher_results.count(),
        'average_marks': avg_marks,
        'user': request.user,
        'is_teacher': True,
        'teacher_subjects': teacher_subjects,
        'teacher_students': teacher_students[:10],  # Show recent students
    }
    return render(request, 'core/dashboard/teacher_dashboard.html', context)


# 👥 TEACHERS LIST - Admin view to see all teachers
@login_required
def teachers_list(request):
    """Admin view to list all teachers and their assignments"""
    if is_teacher(request.user):
        # Teachers can't access this view
        messages.error(request, "Access denied. Admin privileges required.")
        return redirect('/dashboard/')

    # Get all teachers with their assignments
    teachers = Teacher.objects.select_related('user').prefetch_related('subjects', 'students').filter(is_active=True)

    # Calculate statistics
    total_teachers = teachers.count()
    teachers_with_subjects = teachers.filter(subjects__isnull=False).distinct().count()
    teachers_with_students = teachers.filter(students__isnull=False).distinct().count()

    context = {
        'teachers': teachers,
        'total_teachers': total_teachers,
        'teachers_with_subjects': teachers_with_subjects,
        'teachers_with_students': teachers_with_students,
        'user': request.user,
        'is_admin': True,
    }

    return render(request, 'core/dashboard/teachers_list.html', context)


# ➕ ADD MARKS
@login_required
def add_marks(request):
    # Check if user is a teacher and restrict subjects accordingly
    teacher = get_teacher_profile(request.user)
    is_teacher_user = teacher is not None
    
    # Attach teacher to request for form access
    if teacher:
        request.teacher = teacher

    if is_teacher_user:
        # Teacher can only access their assigned subjects and students
        subjects = teacher.subjects.all().values('id', 'name', 'total_marks')
        subject_options = list(subjects)
        teacher_students = teacher.students.all()
        
        # Create subject total marks mapping for JS
        subject_total_map = {s['id']: s['total_marks'] for s in subject_options}
        
        # Order education levels: school (1-10) -> college (XI-XII) -> bachelor
        LEVEL_ORDER = {'school': 0, 'college': 1, 'bachelor': 2}
        teacher_levels = sorted(teacher.levels.all(), key=lambda lvl: LEVEL_ORDER.get(lvl.code, 99))
    else:
        # Admin can access all
        subjects = Subject.objects.all().values('id', 'name', 'total_marks')
        subject_options = list(subjects)
        teacher_students = None
        subject_total_map = {s['id']: s['total_marks'] for s in subject_options}

    # Prepare level/semester hierarchy for JS filtering based on teacher's assigned levels
    levels_hierarchy = {}
    available_levels = []

    if teacher:
        # Teacher's allowed education levels (already sorted above)
        # Build available_levels list from teacher.levels
        for level_obj in teacher_levels:
            available_levels.append({
                'value': level_obj.code,
                'label': level_obj.display_name  # Use display_name to show proper labels
            })

        # Get all students assigned to this teacher
        # (already defined as teacher_students above)
        # teacher_students = teacher.students.all() is already set in the if block above

        # Build hierarchy for each level the teacher is allowed to teach
        if teacher_students:
            for level_obj in teacher_levels:
                level_code = level_obj.code
                # Filter students by this level
                level_students = teacher_students.filter(level=level_code)

                if level_code == 'bachelor':
                    # Further filter by semesters the teacher teaches (if any)
                    if teacher.semesters.exists():
                        # Get semester numbers as strings for CharField comparison
                        teacher_semester_numbers = [str(s) for s in teacher.semesters.values_list('number', flat=True)]
                        level_students = level_students.filter(semester__in=teacher_semester_numbers)
                    # Get distinct semesters from the filtered students
                    semesters = level_students.values_list('semester', flat=True).distinct()
                    level_data = {'type': 'semester', 'options': []}
                    for sem in semesters:
                        if not sem:
                            continue
                        sem_students = level_students.filter(semester=sem).values('id', 'name', 'roll_number')
                        level_data['options'].append({
                            'value': str(sem),
                            'label': str(sem),  # Just the semester number
                            'students': list(sem_students)
                        })
                    levels_hierarchy[level_code] = level_data
                else:
                    # For school/college: group by class
                    classes = level_students.values_list('student_class', flat=True).distinct()
                    level_data = {'type': 'class', 'options': []}
                    for cls in classes:
                        cls_students = level_students.filter(student_class=cls).values('id', 'name', 'roll_number')
                        level_data['options'].append({
                            'value': str(cls),
                            'label': str(cls),  # Just the class identifier (e.g., "9", "XI")
                            'students': list(cls_students)
                        })
                    levels_hierarchy[level_code] = level_data
    else:
        # Admin user: no restrictions, leave empty (level dropdown not shown)
        pass
    
    if request.method == "POST":
        form = ResultForm(request.POST, request=request)
        if form.is_valid():
            student = form.cleaned_data['student']
            subject = form.cleaned_data['subject']
            terminal = request.POST.get('terminal', '1st')
            marks_obtained = form.cleaned_data['marks_obtained']
            total_marks = form.cleaned_data.get('total_marks', 100)

            # Additional validation for teachers
            if is_teacher_user:
                if not teacher.can_access_subject(subject):
                    messages.error(request, "You can only add marks for subjects assigned to you.")
                    return redirect('/add-marks/')
                if teacher_students and student not in teacher_students:
                    messages.error(request, "You can only add marks for students assigned to you.")
                    return redirect('/add-marks/')
                # Teachers cannot change total_marks - use subject's configured total
                if total_marks != subject.total_marks:
                    messages.error(request, f"Total marks must be {subject.total_marks} for {subject.name}.")
                    return redirect('/add-marks/')

            existing = Result.objects.filter(
                student_id=student.id,
                subject_id=subject.id,
                terminal=terminal
            ).first()

            if existing:
                existing.marks_obtained = marks_obtained
                existing.total_marks = total_marks
                existing.save()
                messages.success(request, f"Marks updated successfully for {student.name}!")
            else:
                Result.objects.create(
                    student=student,
                    subject=subject,
                    terminal=terminal,
                    marks_obtained=marks_obtained,
                    total_marks=total_marks
                )
                messages.success(request, "Marks added successfully!")

            return redirect('/marks-list/')
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    field_label = form.fields[field].label if form.fields[field].label else field
                    error_messages.append(f"{field_label}: {error}")

            if not error_messages:
                error_messages = ["Please correct the errors below."]

            for msg in error_messages:
                messages.error(request, msg)
    else:
        form = ResultForm(request=request)
        if is_teacher_user and teacher_students is not None:
            form.fields['student'].queryset = teacher_students

    return render(request, 'core/dashboard/add_marks.html', {
        'form': form,
        'subject_options': subject_options,
        'is_teacher': is_teacher_user,
        'teacher': teacher,
        'available_levels': available_levels,
        'levels_hierarchy_json': json.dumps(levels_hierarchy),
        'subject_total_map_json': json.dumps(subject_total_map),
    })



# 📤 API: Get Student Info
@login_required
def student_info(request, student_id):
    try:
        student = get_object_or_404(Student, id=student_id)
        image_url = None
        if student.image:
            image_url = student.image.url
        
        return JsonResponse({
            'success': True,
            'student': {
                'id': student.id,
                'name': student.name,
                'roll_number': student.roll_number,
                'class': student.student_class,
                'section': student.section,
                'image_url': image_url
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=404)


# 📋 MARKS LIST
@login_required
def marks_list(request):
    teacher = get_teacher_profile(request.user)

    if teacher:
        # Teacher can only see marks for their subjects and students
        all_marks = Result.objects.filter(
            subject__in=teacher.subjects.all(),
            student__in=teacher.students.all()
        ).select_related('student', 'subject')
    else:
        # Admin can see all marks
        all_marks = Result.objects.all().select_related('student', 'subject')

    return render(request, 'core/dashboard/marks_list.html', {
        'result': all_marks,
        'is_teacher': teacher is not None,
        'teacher': teacher,
    })

# ✏️ EDIT MARKS
@login_required
@require_http_methods(["POST"])
def edit_marks(request, mark_id):
    """API endpoint to update marks"""
    try:
        result = get_object_or_404(Result, id=mark_id)
        data = json.loads(request.body)
        
        teacher = get_teacher_profile(request.user)
        marks_obtained = float(data.get('marks_obtained', 0))
        total_marks = float(data.get('total_marks', 100))
        
        # Teacher permissions check
        if teacher:
            if not teacher.can_access_subject(result.subject):
                return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)
            if not teacher.students.filter(id=result.student.id).exists():
                return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)
            # Teachers cannot alter total_marks - enforce subject's configured total
            if total_marks != result.subject.total_marks:
                return JsonResponse({
                    'success': False, 
                    'message': f"Total marks must be {result.subject.total_marks} for {result.subject.name}"
                }, status=400)
        
        result.marks_obtained = marks_obtained
        result.total_marks = total_marks
        result.save()
        
        return JsonResponse({'success': True, 'message': 'Marks updated successfully'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


# 📚 COURSE MATERIALS - Teacher upload and management
@teacher_required
def course_materials(request):
    """Teacher's course materials management"""
    from core.notification_service import NotificationService
    
    teacher = request.teacher

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        subject_id = request.POST.get('subject')
        file_url = request.POST.get('file_url')

        if not title or not subject_id:
            messages.error(request, "Title and subject are required.")
            return redirect('/course-materials/')

        try:
            subject = Subject.objects.get(id=subject_id)

            # Check if teacher can upload for this subject
            if not teacher.can_access_subject(subject):
                messages.error(request, "You can only upload materials for subjects assigned to you.")
                return redirect('/course-materials/')

            material = CourseMaterial.objects.create(
                title=title,
                description=description or '',
                subject=subject,
                teacher=teacher,
                file=request.FILES.get('file'),
                file_url=file_url or None
            )

            # Send notification to all students of this teacher
            NotificationService.notify_course_material_upload(material, teacher)

            messages.success(request, f"Material '{title}' uploaded successfully! All your students have been notified.")

        except Subject.DoesNotExist:
            messages.error(request, "Invalid subject selected.")
        except Exception as e:
            messages.error(request, f"Error uploading material: {str(e)}")

        return redirect('/course-materials/')

    # Get materials uploaded by this teacher
    teacher_materials = CourseMaterial.objects.filter(
        teacher=teacher,
        is_active=True
    ).select_related('subject').order_by('-upload_date')

    context = {
        'teacher': teacher,
        'materials': teacher_materials,
        'teacher_subjects': teacher.subjects.all(),
        'is_teacher': True,
    }

    return render(request, 'core/dashboard/course_materials.html', context)


@login_required
def view_course_materials(request):
    """View course materials - students see materials for their subjects, teachers see their uploads"""
    teacher = get_teacher_profile(request.user)

    if teacher:
        # Teachers see all their uploaded materials
        materials = CourseMaterial.objects.filter(
            teacher=teacher,
            is_active=True
        ).select_related('subject').order_by('-upload_date')
    else:
        # Students see materials for their subjects (admin can see all)
        # For now, show all active materials - can be filtered by student subjects later
        materials = CourseMaterial.objects.filter(
            is_active=True
        ).select_related('subject', 'teacher').order_by('-upload_date')

    context = {
        'materials': materials,
        'is_teacher': teacher is not None,
        'teacher': teacher,
    }

    return render(request, 'core/dashboard/view_materials.html', context)


@teacher_required
def delete_course_material(request, material_id):
    """Delete course material - removes file, notifications, and DB record"""
    from django.db import transaction
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST method required'})

    teacher = request.teacher

    try:
        with transaction.atomic():
            material = CourseMaterial.objects.get(id=material_id, teacher=teacher)
            material_title = material.title
            
            # Delete physical file
            if material.file:
                try:
                    if material.file.storage.exists(material.file.name):
                        material.file.delete(save=False)
                except Exception as e:
                    print(f"File deletion error: {e}")
            
            # Clean up orphaned notifications (created before material FK existed)
            # These notifications have no material FK but belong to this teacher and mention the material title
            Notification.objects.filter(
                sender=teacher,
                title__icontains=material_title,
                material__isnull=True
            ).delete()
            
            # Delete the material (CASCADE will delete linked notifications automatically)
            material.delete()
            
        return JsonResponse({'success': True, 'message': 'Material deleted successfully'})
    except CourseMaterial.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Material not found or access denied'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def change_password(request):
    """Allow students to change their password"""
    try:
        data = json.loads(request.body)
        current_password = data.get('current_password', '')
        new_password = data.get('new_password', '')
        
        if not current_password or not new_password:
            return JsonResponse({'success': False, 'message': 'Both passwords are required'}, status=400)
         
        if len(new_password) < 4:
            return JsonResponse({'success': False, 'message': 'Password must be at least 4 characters'}, status=400)
         
        user = request.user
         
        # Verify current password
        if not user.check_password(current_password):
            return JsonResponse({'success': False, 'message': 'Current password is incorrect'}, status=400)
         
        # Set new password
        user.set_password(new_password)
        user.save()
         
        # Re-login the user with new password
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, user)
         
        return JsonResponse({'success': True, 'message': 'Password changed successfully!'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid request data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


# 🗑️ DELETE MARKS
@login_required
@require_http_methods(["POST"])
def delete_marks(request, mark_id):
    """API endpoint to delete marks"""
    try:
        result = get_object_or_404(Result, id=mark_id)
        
        # Check teacher permissions
        teacher = get_teacher_profile(request.user)
        if teacher:
            # Teacher can only delete marks for their subjects and students
            if not (teacher.can_access_subject(result.subject) and 
                   (teacher.students.filter(id=result.student.id).exists())):
                return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)
        
        student_name = result.student.name
        subject_name = result.subject.name
        result.delete()
        return JsonResponse({'success': True, 'message': f'Marks for {student_name} in {subject_name} deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)




# MARK SHEET VIEW
@login_required
def mark_sheet(request, student_id=None, terminal=None):
    """Generate official mark sheet for a student with terminal filter"""
    import qrcode
    from io import BytesIO
    import base64
    from datetime import datetime
    
    # Check teacher permissions
    teacher = get_teacher_profile(request.user)
    if teacher and not teacher.is_admin:
        # Teachers can only view mark sheets for their students
        if student_id:
            student = get_object_or_404(Student, id=student_id)
            if not teacher.students.filter(id=student.id).exists():
                return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)
        else:
            # If no student_id provided, show error or redirect
            return JsonResponse({'success': False, 'message': 'Student ID required'}, status=400)
    
    # Get all students for selection dropdown
    students = Student.objects.all().order_by('name')
    
    if student_id and terminal:
        # Get specific student's results for a specific terminal
        results = Result.objects.filter(
            student_id=student_id,
            terminal=terminal
        ).select_related('student', 'subject')
        selected_student = Student.objects.get(id=student_id)
    elif student_id:
        # Get specific student's results (all terminals)
        results = Result.objects.filter(student_id=student_id).select_related('student', 'subject')
        selected_student = Student.objects.get(id=student_id)
        terminal = 'All'
    else:
        results = []
        selected_student = None
        terminal = 'All'
    
    # Calculate totals
    total_subjects = len(results) if results else 0
    total_marks_obtained = sum(r.marks_obtained for r in results)
    total_marks = sum(r.total_marks for r in results)
    overall_percentage = (total_marks_obtained / total_marks * 100) if total_marks > 0 else 0
    
    # Determine overall pass/fail (all subjects must have >= 40%)
    all_passed = all((r.marks_obtained / r.total_marks * 100) >= 40 for r in results) if results else False
    
    # Build results list with grade and percentage attributes
    results_with_grades = []
    for r in results:
        pct = (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0
        results_with_grades.append({
            'result': r,
            'grade': 'P' if pct >= 40 else 'F',
            'percentage': round(pct, 1)
        })
    
    # Generate college QR code
    college_data = "SOCH_COLLEGE_OF_IT|RANIPAWA-12|POKHARA|ESTD:2020"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(college_data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    college_qr_code = base64.b64encode(buffer.getvalue()).decode()
    
    # Get current academic year
    academic_year = datetime.now().year
    
    # Generate document ID
    doc_id = f"SOCH-{datetime.now().strftime('%Y%m%d')}-{student_id or 'DEMO'}"
    
    context = {
        'students': students,
        'student_results': results_with_grades,
        'selected_student': selected_student,
        'selected_terminal': terminal,
        'total_subjects': total_subjects,
        'total_marks_obtained': total_marks_obtained,
        'total_marks': total_marks,
        'overall_percentage': round(overall_percentage, 2),
        'all_passed': all_passed,
        'academic_year': f"{academic_year}-{academic_year + 1}",
        'college_qr_code': college_qr_code,
        'doc_id': doc_id,
        'is_teacher': teacher is not None,
        'teacher': teacher,
    }
    
    return render(request, 'core/dashboard/mark_sheet.html', context)


# SELECT STUDENT FOR MARK SHEET
@login_required
def select_mark_sheet(request):
    """Student selection page for mark sheet generation"""
    # Check if user is teacher
    teacher = get_teacher_profile(request.user)
    
    # Get all students for selection
    if teacher and not teacher.is_admin:
        # Teachers can only see their students
        students = teacher.students.all().order_by('name')
    else:
        # Admin can see all students
        students = Student.objects.all().order_by('name')
    
    # Get unique terminals
    terminals = Result.objects.values_list('terminal', flat=True).distinct()
    
    context = {
        'students': students,
        'terminals': terminals,
        'is_teacher': teacher is not None,
        'teacher': teacher,
    }
    
    return render(request, 'core/dashboard/select_mark_sheet.html', context)


#  STUDENT ANALYSIS VIEW
@login_required
def student_analysis(request):
    """AI-powered analysis of student performance, focusing on students with low marks"""
    
    # Check if user is teacher
    teacher = get_teacher_profile(request.user)
    
    # Get filter parameters
    filter_type = request.GET.get('filter', 'attention')
    selected_class = request.GET.get('class', '')
    sort_by = request.GET.get('sort', 'lowest')
    
    # Get students with their results
    if teacher and not teacher.is_admin:
        # Teachers can only see their students
        students = teacher.students.all().prefetch_related('result_set')
    else:
        # Admin can see all students
        students = Student.objects.all().prefetch_related('result_set')
    
    student_analyses = []
    all_classes = []
    needs_attention_count = 0
    at_risk_count = 0
    total_percentage = 0
    student_count = 0
    
    for student in students:
        results = list(student.result_set.select_related('subject').all())
        
        if not results:
            continue
        
        # Get student's class for filtering and stats
        student_class = student.student_class
        all_classes.append(student_class)
        
        # Calculate percentage
        total_obtained = sum(r.marks_obtained for r in results)
        total_possible = sum(r.total_marks for r in results)
        percentage = (total_obtained / total_possible * 100) if total_possible > 0 else 0
        
        total_percentage += percentage
        student_count += 1
        
        # Find weak subjects (below 50%)
        weak_subjects = []
        for r in results:
            r_percentage = (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0
            if r_percentage < 50:
                weak_subjects.append(r)
        
        # Count students needing attention
        if percentage < 60:
            needs_attention_count += 1
        if percentage < 40:
            at_risk_count += 1
        
        # Add to analyses if matches filter
        should_add = False
        
        if filter_type == 'all':
            should_add = True
        elif filter_type == 'critical' and percentage < 40:
            should_add = True
        elif filter_type == 'warning' and 40 <= percentage < 60:
            should_add = True
        elif filter_type == 'attention' and percentage < 60:
            should_add = True
        
        # Filter by class
        if selected_class and student_class != selected_class:
            should_add = False
        
        if should_add:
            # Generate AI Summary
            ai_summary = generate_ai_summary(student, results, percentage, weak_subjects)
            
            # Generate Recommendations
            recommendations = generate_recommendations(student, results, percentage, weak_subjects)
            
            student_analyses.append({
                'student': student,
                'percentage': round(percentage, 2),
                'weak_subjects': weak_subjects,
                'total_attempts': len(results),
                'ai_summary': ai_summary,
                'recommendations': recommendations,
            })
    
    # Sort analyses
    if sort_by == 'lowest':
        student_analyses.sort(key=lambda x: x['percentage'])
    elif sort_by == 'highest':
        student_analyses.sort(key=lambda x: x['percentage'], reverse=True)
    elif sort_by == 'name':
        student_analyses.sort(key=lambda x: x['student'].name)
    
    # Calculate class average
    class_average = (total_percentage / student_count) if student_count > 0 else 0
    
    # Get unique classes
    all_classes = sorted(set(all_classes))
    
    context = {
        'student_analyses': student_analyses,
        'total_students': student_count,
        'needs_attention_count': needs_attention_count,
        'at_risk_count': at_risk_count,
        'class_average': class_average,
        'classes': all_classes,
        'selected_class': selected_class,
        'filter_type': filter_type,
        'sort_by': sort_by,
        'is_teacher': teacher is not None,
        'teacher': teacher,
    }
    
    return render(request, 'core/dashboard/student_analysis.html', context)


def generate_ai_summary(student, results, percentage, weak_subjects):
    """Generate AI-powered summary for a student"""
    
    total_obtained = sum(r.marks_obtained for r in results)
    total_possible = sum(r.total_marks for r in results)
    
    # Determine performance level
    if percentage < 40:
        level = "critical"
        intro = f"🚨 {student.name} is showing concerning performance with an overall average of {percentage:.1f}%. "
    elif percentage < 60:
        level = "warning"
        intro = f"⚠️ {student.name}'s performance needs improvement. Current average is {percentage:.1f}%. "
    elif percentage < 75:
        level = "good"
        intro = f"👍 {student.name} is performing at a satisfactory level with {percentage:.1f}% average. "
    else:
        level = "excellent"
        intro = f"🌟 {student.name} is performing excellently with {percentage:.1f}% average! "
    
    # Add subject-specific analysis
    subject_analysis = ""
    if weak_subjects:
        weak_names = [f"**{ws.subject.name}** ({ws.marks_obtained}/{ws.total_marks})" for ws in weak_subjects[:3]]
        if len(weak_names) > 1:
            subject_analysis = f" Areas of concern include {', '.join(weak_names[:-1])} and {weak_names[-1]}."
        else:
            subject_analysis = f" Primary area of concern is {weak_names[0]}."
    
    # Add best subject
    best_result = max(results, key=lambda r: (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0)
    best_percentage = (best_result.marks_obtained / best_result.total_marks * 100) if best_result.total_marks > 0 else 0
    if best_percentage >= 80:
        subject_analysis += f" Shows strong performance in **{best_result.subject.name}** ({best_percentage:.1f}%)."
    
    # Add summary
    if level == "critical":
        summary = "Immediate intervention is required. Student's current trajectory may lead to academic failure."
    elif level == "warning":
        summary = "With targeted improvement in weak areas, student can significantly improve performance."
    elif level == "good":
        summary = "Student is on track. Focus on weak subjects can help achieve distinction."
    else:
        summary = "Excellent performance! Student should be encouraged to maintain this standard."
    
    return intro + subject_analysis + " " + summary


def generate_recommendations(student, results, percentage, weak_subjects):
    """Generate personalized recommendations for a student"""
    
    recommendations = []
    
    # Priority recommendations based on performance
    if percentage < 40:
        recommendations.append({
            'priority': 'high',
            'title': '🔴 Immediate Parent Meeting Required',
            'description': 'Schedule urgent meeting with parents/guardians to discuss academic intervention strategies.'
        })
        recommendations.append({
            'priority': 'high',
            'title': '📚 Mandatory Tutoring Sessions',
            'description': 'Student requires additional support through remedial classes or private tutoring.'
        })
    elif percentage < 60:
        recommendations.append({
            'priority': 'medium',
            'title': '🟡 Schedule Academic Counseling',
            'description': 'Arrange meeting to identify specific learning challenges and set improvement goals.'
        })
    
    # Subject-specific recommendations
    for weak in weak_subjects[:2]:
        weak_percentage = (weak.marks_obtained / weak.total_marks * 100) if weak.total_marks > 0 else 0
        rec_title = f"Focus on {weak.subject.name}"
        
        if weak_percentage < 30:
            rec_desc = f"Critical weakness in {weak.subject.name} (scored {weak.marks_obtained}/{weak.total_marks}). Consider remedial classes."
        elif weak_percentage < 40:
            rec_desc = f"Weak performance in {weak.subject.name}. Recommend extra practice and teacher consultations."
        else:
            rec_desc = f"{weak.subject.name} needs improvement. Suggest focused study plan and regular assessments."
            
        recommendations.append({
            'priority': 'medium' if weak_percentage < 40 else 'low',
            'title': rec_title,
            'description': rec_desc
        })
    
    # General recommendations
    if percentage >= 40:
        recommendations.append({
            'priority': 'low',
            'title': '📖 Establish Study Schedule',
            'description': 'Create a structured daily study routine with dedicated time for each subject.'
        })
    
    recommendations.append({
        'priority': 'low',
        'title': '💪 Encourage Participation',
        'description': 'Promote active engagement in class discussions and practical activities.'
    })
    
    return recommendations[:5]  # Return top 5 recommendations


@login_required
def student_notifications(request):
    """Get all notifications for the logged-in student - excludes notifications for deleted materials"""
    from core.notification_service import NotificationService
    from django.db.models import Q
    
    # Get student
    student = None
    if hasattr(request, 'student'):
        student = request.student
    else:
        username = request.user.username.lower().strip()
        try:
            student = Student.objects.get(name__iexact=username)
        except Student.DoesNotExist:
            students = Student.objects.filter(name__icontains=username)
            if students.count() == 1:
                student = students.first()
            elif students.count() == 0:
                return JsonResponse({'success': True, 'notifications': [], 'unread_count': 0})
            else:
                student = students.first()
    
    if not student:
        return JsonResponse({'success': False, 'message': 'Student profile not found'}, status=404)
    
    # Filter logic:
    # Keep: General notifications (material=null) that are NOT material-related (title not starting with "New Material:")
    # Keep: Notifications linked to active materials (material__is_active=True)
    # Remove: Orphaned material notifications (material=null + title starts with "New Material:")
    # Remove: Notifications for soft-deleted materials (material__is_active=False)
    valid_notifications = Notification.objects.filter(recipient=student).exclude(
        Q(material__isnull=True, title__startswith='New Material:') |
        Q(material__is_active=False)
    )
    
    unread_count = valid_notifications.filter(is_read=False).count()
    recent = valid_notifications.order_by('-created_at')[:20]
    
    notifications = []
    for n in recent:
        notifications.append({
            'id': n.id,
            'title': n.title,
            'message': n.message,
            'type': n.notification_type,
            'priority': n.priority,
            'is_read': n.is_read,
            'created_at': n.created_at.strftime('%b %d, %Y %I:%M %p'),
            'link_url': n.link_url or '',
            'sender': n.sender.get_full_name() if n.sender else None,
        })
    
    return JsonResponse({
        'success': True,
        'notifications': notifications,
        'unread_count': unread_count
    })
    
    return JsonResponse({
        'success': True,
        'notifications': notifications,
        'unread_count': unread_count
    })


@login_required
@require_http_methods(["POST"])
def mark_notification_read(request, notification_id):
    """Mark a single notification as read"""
    from core.notification_service import NotificationService
    
    student = None
    if hasattr(request, 'student'):
        student = request.student
    else:
        username = request.user.username.lower().strip()
        try:
            student = Student.objects.get(name__iexact=username)
        except Student.DoesNotExist:
            students = Student.objects.filter(name__icontains=username)
            if students.count() == 1:
                student = students.first()
    
    if not student:
        return JsonResponse({'success': False, 'message': 'Student profile not found'}, status=404)
    
    success = NotificationService.mark_as_read(notification_id, student)
    
    if success:
        return JsonResponse({'success': True, 'message': 'Notification marked as read'})
    else:
        return JsonResponse({'success': False, 'message': 'Notification not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def mark_all_notifications_read(request):
    """Mark all notifications as read for the student"""
    from core.notification_service import NotificationService
    
    student = None
    if hasattr(request, 'student'):
        student = request.student
    else:
        username = request.user.username.lower().strip()
        try:
            student = Student.objects.get(name__iexact=username)
        except Student.DoesNotExist:
            students = Student.objects.filter(name__icontains=username)
            if students.count() == 1:
                student = students.first()
    
    if not student:
        return JsonResponse({'success': False, 'message': 'Student profile not found'}, status=404)
    
    NotificationService.mark_all_as_read(student)
    
    return JsonResponse({'success': True, 'message': 'All notifications marked as read'})


# 📊 CHART DATA API
@login_required
def chart_data(request):
    """Return dashboard chart data as JSON"""
    from django.db.models import Avg
    
    teacher = get_teacher_profile(request.user)
    
    if teacher:
        # Teacher sees only their students' data
        teacher_students = teacher.students.all()
        results = Result.objects.filter(student__in=teacher_students)
    else:
        # Admin sees all data
        results = Result.objects.all()
    
    # Subject-wise average marks
    subject_stats = []
    for subject in Subject.objects.all():
        subject_results = results.filter(subject=subject)
        if subject_results.exists():
            avg = subject_results.aggregate(avg=Avg('marks_obtained'))['avg']
            subject_stats.append({
                'subject': subject.name,
                'average': round(float(avg), 1) if avg else 0
            })
    
    # Terminal-wise performance
    terminal_stats = []
    for terminal in ['1st', '2nd', '3rd', 'Final']:
        terminal_results = results.filter(terminal=terminal)
        if terminal_results.exists():
            avg = terminal_results.aggregate(avg=Avg('marks_obtained'))['avg']
            terminal_stats.append({
                'terminal': terminal,
                'average': round(float(avg), 1) if avg else 0
            })
    
    # Top 5 students
    top_students = []
    for student in Student.objects.all():
        student_results = results.filter(student=student)
        if student_results.exists():
            total = sum(r.marks_obtained for r in student_results)
            total_possible = sum(r.total_marks for r in student_results)
            pct = (total / total_possible * 100) if total_possible > 0 else 0
            top_students.append({
                'name': student.name,
                'percentage': round(pct, 1)
            })
    
    top_students.sort(key=lambda x: x['percentage'], reverse=True)
    
    data = {
        'subject_stats': subject_stats,
        'terminal_stats': terminal_stats,
        'top_students': top_students[:5],
        'total_students': Student.objects.count(),
        'total_subjects': Subject.objects.count(),
        'total_results': results.count(),
    }
    return JsonResponse(data)


# 🤖 AI ANALYZE STUDENT (API) — removed
@login_required
@require_http_methods(["GET"])
def ai_analyze_student(request, student_id):
    """AI analysis not configured — OpenRouter integration removed."""
    return JsonResponse({
        'success': False,
        'message': 'AI analysis is not available. Configure OpenRouter to enable this feature.'
    }, status=503)


# 📧 SEND STUDENT NOTIFICATION (API) — removed
@login_required
@require_http_methods(["POST"])
def send_student_notification(request, student_id):
    """Notifications not configured — OpenRouter integration removed."""
    return JsonResponse({
        'success': False,
        'message': 'Notification service is not available. Configure OpenRouter to enable this feature.'
    }, status=503)


# 📢 NOTIFY ALL STUDENTS (API) — removed
@login_required
@require_http_methods(["POST"])
def notify_all_students(request):
    """Automated notifications not configured — OpenRouter integration removed."""
    return JsonResponse({
        'success': False,
        'message': 'Auto-notification service is not available. Configure OpenRouter to enable this feature.'
    }, status=503)


# 📋 DAILY REPORT (API) — removed
@login_required
def daily_report(request):
    """Daily report not configured — OpenRouter integration removed."""
    return JsonResponse({
        'success': False,
        'message': 'Daily report service is not available. Configure OpenRouter to enable this feature.'
    }, status=503)


# ═════════════════════════════════════════════════════════════════════════════
#  SMART ADMIN & ANALYTICS VIEWS
# ═════════════════════════════════════════════════════════════════════════════

def _is_admin_user(user):
    """Return True if user has Teacher.is_admin=True OR has a RBAC super_admin/admin role."""
    try:
        if Teacher.objects.filter(user=user, is_active=True, is_admin=True).exists():
            return True
    except Exception:
        pass
    try:
        profile = UserProfile.objects.select_related('role').get(user=user)
        return profile.role.code in ('super_admin', 'admin')
    except (UserProfile.DoesNotExist, Exception):
        return False


def _admin_required(view_func):
    """Decorator: only users with Teacher.is_admin=True may pass."""
    from functools import wraps
    @login_required
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not _is_admin_user(request.user):
            messages.error(request, "Admin privileges required.")
            return redirect('/dashboard/')
        return view_func(request, *args, **kwargs)
    return wrapper


@_admin_required
def smart_dashboard(request):
    """Smart Analytics Dashboard: institution-wide KPIs, graphs, toppers, weak students"""
    from django.db.models import Avg, Sum, Count, Q, F, FloatField, ExpressionWrapper, Case, When
    from core.models import GradeScale
    from datetime import date

    period = request.GET.get('period', 'all')  # all / month / semester
    selected_class = request.GET.get('student_class', '')

    # Base result queryset
    all_results = Result.objects.select_related('student', 'subject')
    if selected_class:
        all_results = all_results.filter(student__student_class=selected_class)

    # ── KPI Stats ──────────────────────────────────────────────────────────────
    total_students = Student.objects.count()
    total_subjects = Subject.objects.count()
    total_results_count = all_results.count()
    total_teachers = Teacher.objects.filter(is_active=True).count()

    if all_results.exists():
        avg_all = all_results.aggregate(avg=Avg('marks_obtained'))['avg'] or 0
        total_possible = all_results.aggregate(s=Sum('total_marks'))['s'] or 0
        total_obt = all_results.aggregate(s=Sum('marks_obtained'))['s'] or 0
        overall_pct = (total_obt / total_possible * 100) if total_possible > 0 else 0
    else:
        avg_all = overall_pct = 0

    # Pass / Fail
    gs = GradeScale.objects.filter(is_active=True).first()
    pass_mark_pct = gs.pass_mark_percent if gs else 40.0
    total_pass = total_fail = 0
    if all_results.exists():
        for r in all_results:
            p = (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0
            if p >= pass_mark_pct:
                total_pass += 1
            else:
                total_fail += 1

    pass_fail_ratio = f"{total_pass}:{total_fail}" if (total_pass + total_fail) > 0 else "0:0"

    # Pass rate %
    total_attempts = total_pass + total_fail
    pass_rate = round((total_pass / total_attempts * 100), 1) if total_attempts > 0 else 0

    # ── Top Students ────────────────────────────────────────────────────────────
    student_scores = []
    for stu in Student.objects.prefetch_related('result_set'):
        res = stu.result_set.all()
        if not res.exists():
            continue
        obt = sum(r.marks_obtained for r in res)
        pos = sum(r.total_marks for r in res)
        p = (obt / pos * 100) if pos > 0 else 0
        student_scores.append({'student': stu, 'percentage': round(p, 1), 'obtained': obt, 'total': pos})
    student_scores.sort(key=lambda x: x['percentage'], reverse=True)
    top_students = student_scores[:10]

    # ── Weak Students ───────────────────────────────────────────────────────────
    weak_students = [s for s in student_scores if s['percentage'] < pass_mark_pct][:10]

    # ── Subject-wise Performance ───────────────────────────────────────────────
    subj_stats = []
    for subj in Subject.objects.all():
        s_res = all_results.filter(subject=subj)
        if s_res.exists():
            avg = s_res.aggregate(avg=Avg('marks_obtained'))['avg'] or 0
            pos_sum = s_res.aggregate(s=Sum('total_marks'))['s'] or 0
            obt_sum = s_res.aggregate(s=Sum('marks_obtained'))['s'] or 0
            pct = (obt_sum / pos_sum * 100) if pos_sum > 0 else 0
            subj_stats.append({
                'subject': subj.name,
                'avg': round(avg, 1),
                'percentage': round(pct, 1),
            })
    subj_stats.sort(key=lambda x: x['percentage'])

    # ── Class-wise Performance ─────────────────────────────────────────────────
    classes = Student.objects.values_list('student_class', flat=True).distinct().order_by('student_class')
    class_stats = []
    for cls in classes:
        cls_students = Student.objects.filter(student_class=cls)
        cls_results = all_results.filter(student__student_class=cls)
        if cls_results.exists():
            pos = cls_results.aggregate(s=Sum('total_marks'))['s'] or 0
            obt = cls_results.aggregate(s=Sum('marks_obtained'))['s'] or 0
            pct = (obt / pos * 100) if pos > 0 else 0
            class_stats.append({'class': cls, 'students': cls_students.count(), 'percentage': round(pct, 1)})

    # ── Terminal-wise ─────────────────────────────────────────────────────────
    term_stats = []
    for term in ['1st', '2nd', '3rd', 'Final']:
        t_res = all_results.filter(terminal=term)
        if t_res.exists():
            pos = t_res.aggregate(s=Sum('total_marks'))['s'] or 0
            obt = t_res.aggregate(s=Sum('marks_obtained'))['s'] or 0
            pct = (obt / pos * 100) if pos > 0 else 0
            term_stats.append({'terminal': term, 'percentage': round(pct, 1)})

    # ── Attendance Summary ───────────────────────────────────────────────────
    att_present = Attendance.objects.filter(status='present').count()
    att_total = Attendance.objects.count()
    att_pct = round((att_present / att_total * 100), 1) if att_total > 0 else 0

    # ── Levels ──
    all_classes_list = list(classes)

    context = {
        'total_students': total_students,
        'total_subjects': total_subjects,
        'total_teachers': total_teachers,
        'total_results': total_results_count,
        'overall_average': round(avg_all, 1),
        'overall_percentage': round(overall_pct, 1),
        'pass_count': total_pass,
        'fail_count': total_fail,
        'pass_fail_ratio': pass_fail_ratio,
        'pass_rate': pass_rate,
        'top_students': top_students,
        'weak_students': weak_students,
        'subj_stats': subj_stats,
        'class_stats': class_stats,
        'term_stats': term_stats,
        'att_present': att_present,
        'att_total': att_total,
        'att_pct': att_pct,
        'all_classes': all_classes_list,
        'selected_class': selected_class,
        'pass_mark_pct': pass_mark_pct,
    }
    return render(request, 'core/dashboard/smart_dashboard.html', context)


# ═════════════════════════════════════════════════════════════════════════════
#  STUDENT MANAGEMENT
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def bulk_import_students(request):
    """Bulk import students from CSV/Excel file"""
    from core.views_excel import _resolve_student, _resolve_subject
    import csv, io

    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        if not csv_file or not csv_file.name.lower().endswith(('.csv', '.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid CSV or Excel file.')
            return redirect('core:bulk_import_students')

        try:
            data = csv_file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(data))
            imported = skipped = errors = 0
            error_log = []
            for row_num, row in enumerate(reader, start=2):
                try:
                    name = row.get('name', '').strip()
                    roll = row.get('roll_number', '').strip()
                    level = row.get('level', 'school').strip()
                    cls = row.get('student_class', '').strip()
                    section = row.get('section', 'A').strip()
                    email = row.get('email', '').strip()
                    phone = row.get('phone', '').strip()
                    dob_str = row.get('date_of_birth', '').strip()
                    gender = row.get('gender', '').strip()
                    blood = row.get('blood_group', '').strip()

                    if not name or not cls:
                        errors += 1
                        error_log.append(f'Row {row_num}: name and student_class are required')
                        continue

                    if roll and Student.objects.filter(roll_number=roll).exists():
                        skipped += 1
                        continue

                    dob = None
                    if dob_str:
                        try:
                            dob = date.fromisoformat(dob_str)
                        except ValueError:
                            pass

                    # Auto-create Django user
                    username = name.replace(' ', '').lower()
                    if not User.objects.filter(username=username).exists():
                        user = User.objects.create_user(username=username, password=DEFAULT_PASSWORD,
                                                         first_name=name.split()[0], email=email or '')
                    else:
                        user = User.objects.get(username=username)

                    existing = Student.objects.filter(name__iexact=name).first()
                    if existing:
                        skipped += 1
                        continue

                    Student.objects.create(
                        name=name, roll_number=roll or None, level=level, student_class=cls,
                        section=section, email=email or None, phone=phone or None,
                        date_of_birth=dob, gender=gender, blood_group=blood,
                    )
                    imported += 1
                except Exception as e:
                    errors += 1
                    error_log.append(f'Row {row_num}: {str(e)}')

            msg_parts = []
            if imported:
                msg_parts.append(f'{imported} students imported')
            if skipped:
                msg_parts.append(f'{skipped} skipped (already exist)')
            if errors:
                msg_parts.append(f'{errors} errors')
            messages.success(request, 'Bulk import: ' + ', '.join(msg_parts) + '.')

            NotificationService.log_activity(request.user, 'bulk_upload',
                                              f'Bulk imported {imported} students ({errors} errors)')
        except Exception as e:
            messages.error(request, f'CSV parsing error: {str(e)}')

        return redirect('core:students_list')

    return render(request, 'core/dashboard/bulk_import_students.html', {})


@_admin_required
def promote_students(request):
    """Bulk promote students to next class/year and archive academic history"""
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        promoted_class = request.POST.get('promoted_to_class', '').strip()
        academic_year_id = request.POST.get('academic_year')

        if not student_ids or not promoted_class:
            messages.error(request, 'Please select students and specify the target class.')
            return redirect('core:promote_students')

        try:
            acad_year = AcademicYear.objects.get(id=academic_year_id) if academic_year_id else None
        except AcademicYear.DoesNotExist:
            acad_year = None

        promoted = 0
        for sid in student_ids:
            try:
                stu = Student.objects.get(id=sid)
                prev = stu.student_class
                prev_sec = stu.section
                total_obt = sum((r.marks_obtained for r in stu.result_set.all()), 0.0)
                total_pos = sum((r.total_marks for r in stu.result_set.all()), 0.0)
                pct = (total_obt / total_pos * 100) if total_pos > 0 else 0

                gs = GradeScale.objects.filter(is_active=True).first()
                grade = gs.get_grade(pct) if gs else ''

                StudentAcademicHistory.objects.create(
                    student=stu,
                    academic_year=acad_year,
                    previous_class=prev,
                    previous_section=prev_sec,
                    promoted_to_class=promoted_class,
                    total_marks_obtained=total_obt,
                    total_marks=total_pos,
                    percentage=round(pct, 2),
                    grade=grade,
                )
                stu.student_class = promoted_class
                stu.is_promoted = True
                stu.save(update_fields=['student_class', 'is_promoted'])
                promoted += 1
            except Student.DoesNotExist:
                continue

        messages.success(request, f'{promoted} students promoted to {promoted_class}.')
        NotificationService.log_activity(request.user, 'promote_student',
                                          f'Promoted {promoted} students to {promoted_class}')
        return redirect('core:students_list')

    students = Student.objects.all().order_by('student_class', 'name')
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    return render(request, 'core/dashboard/promote_students.html', {
        'students': students, 'academic_years': academic_years,
    })


@_admin_required
def student_history(request, student_id):
    """View academic history for a specific student"""
    student = get_object_or_404(Student, id=student_id)
    history = student.history.select_related('academic_year').all()
    return render(request, 'core/dashboard/student_history.html', {
        'student': student, 'history': history,
    })


@login_required
def manage_parent(request, student_id=None):
    """Create / edit parent info for a student"""
    is_admin = _is_admin_user(request.user)
    if student_id:
        student = get_object_or_404(Student, id=student_id)
        if not is_admin:
            messages.error(request, 'Access denied.')
            return redirect('core:students_list')
    else:
        student = None

    parent_obj = None
    if student and hasattr(student, 'parent_info'):
        parent_obj = student.parent_info

    if request.method == 'POST':
        if not parent_obj:
            parent_obj = Parent(student=student)
        parent_obj.father_name = request.POST.get('father_name', '')
        parent_obj.mother_name = request.POST.get('mother_name', '')
        parent_obj.guardian_name = request.POST.get('guardian_name', '')
        parent_obj.phone = request.POST.get('phone', '')
        parent_obj.email = request.POST.get('email', '')
        parent_obj.address = request.POST.get('address', '')
        parent_obj.relation = request.POST.get('relation', 'Father')
        parent_obj.is_primary_contact = request.POST.get('is_primary_contact') == 'on'
        parent_obj.save()
        # A11 — if "create_parent_user" action, create a linked ParentUser portal account
        create_pu = request.POST.get('create_parent_user')
        if create_pu and parent_obj and not hasattr(parent_obj, 'user_account'):
            import secrets, uuid
            token = uuid.uuid4().hex
            username = f"{student.name.replace(' ', '').lower()}_parent"
            if not User.objects.filter(username=username).exists():
                user = User.objects.create_user(username=username, password=DEFAULT_PASSWORD)
            else:
                user = User.objects.get(username=username)
            ParentUser.objects.create(
                user=user, parent=parent_obj,
                phone=parent_obj.phone,
                verification_token=token,
            )
            from core.notification_service import NotificationService
            NotificationService.log_activity(
                request.user, 'add_student',
                f'Created ParentUser portal account for {parent_obj.guardian_name or parent_obj.father_name or "parent"} of {student.name}',
            )
            messages.success(request,
                f'Parent info saved. Portal account "{username}" created. Verification token: {token}')
            return redirect('core:manage_parent', student_id=student.id)
        messages.success(request, f'Parent info saved for {student.name}.')
        return redirect('core:students_list')

    return render(request, 'core/dashboard/manage_parent.html', {
        'student': student,
        'parent': parent_obj,
        'has_parent_user_account': hasattr(parent_obj, 'user_account') if parent_obj else False,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  TEACHER MANAGEMENT (ADMIN)
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def teacher_performance(request):
    """Teacher performance evaluation summary for admin"""
    teachers = Teacher.objects.filter(is_active=True).prefetch_related('evaluations').order_by('first_name')
    data = []
    for t in teachers:
        evals = t.evaluations.all()
        if evals.exists():
            avg = evals.aggregate(avg=Avg('overall_score'))['avg'] or 0
            ev_count = evals.count()
        else:
            avg = 0
            ev_count = 0
        data.append({'teacher': t, 'avg_score': round(avg, 1), 'eval_count': ev_count,
                      'students': t.students.count(), 'subjects': t.subjects.count()})
    return render(request, 'core/dashboard/teacher_performance.html', {'teacher_data': data})


@_admin_required
def workload_analytics(request):
    """Teacher workload distribution view"""
    teachers = Teacher.objects.filter(is_active=True).prefetch_related('subjects', 'students').order_by('first_name')
    data = []
    max_s = max_t = 1
    for t in teachers:
        sub_count = t.subjects.count()
        stu_count = t.students.count()
        data.append({'teacher': t, 'subjects': sub_count, 'students': stu_count})
        max_s = max(max_s, sub_count)
        max_t = max(max_t, stu_count)
    return render(request, 'core/dashboard/workload_analytics.html', {
        'teacher_data': data, 'max_subjects': max_s, 'max_students': max_t,
    })


@_admin_required
def activity_logs(request):
    """System-wide activity log viewer for admin"""
    logs = ActivityLog.objects.select_related('user').all()[:500]
    action_filter = request.GET.get('action', '')
    user_filter = request.GET.get('user', '')
    if action_filter:
        logs = logs.filter(action=action_filter)
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    actions = ActivityLog.ACTION_CHOICES
    return render(request, 'core/dashboard/activity_logs.html', {
        'logs': logs, 'actions': actions, 'action_filter': action_filter, 'user_filter': user_filter,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  GRADE SCALE & ACADEMIC YEAR CONFIG (ADMIN)
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def grade_scale_config(request):
    """Manage grade scale: thresholds and labels"""
    from core.forms import GradeScaleForm
    gs = GradeScale.objects.filter(is_active=True).first()
    if not gs:
        gs = GradeScale.objects.create(name='Standard', pass_mark_percent=40.0)

    if request.method == 'POST':
        form = GradeScaleForm(request.POST, instance=gs)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grade scale configuration saved.')
            return redirect('core:grade_scale_config')
    else:
        form = GradeScaleForm(instance=gs)

    return render(request, 'core/dashboard/grade_scale_config.html', {'form': form, 'grade_scale': gs})


@_admin_required
def academic_year_config(request):
    """Manage academic years — create, set current, close"""
    years = AcademicYear.objects.all().order_by('-start_date')
    current_year = AcademicYear.objects.filter(is_current=True).first()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name', '').strip()
            sd = request.POST.get('start_date', '').strip()
            ed = request.POST.get('end_date', '').strip()
            if name and sd and ed:
                from datetime import date
                try:
                    AcademicYear.objects.create(name=name, start_date=date.fromisoformat(sd),
                                                 end_date=date.fromisoformat(ed))
                    messages.success(request, f'Academic year "{name}" created.')
                except Exception as e:
                    messages.error(request, str(e))
        elif action == 'set_current':
            yid = request.POST.get('year_id')
            try:
                yr = AcademicYear.objects.get(id=yid)
                yr.is_current = True
                yr.save()
                messages.success(request, f'"{yr.name}" set as current academic year.')
            except AcademicYear.DoesNotExist:
                messages.error(request, 'Year not found.')
        return redirect('core:academic_year_config')

    return render(request, 'core/dashboard/academic_year_config.html', {
        'years': years, 'current_year': current_year,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  ATTENDANCE
# ═════════════════════════════════════════════════════════════════════════════

def _attendance_date_range(request):
    """Shared helper to resolve date range from GET params"""
    from datetime import timedelta
    today = date.today()
    start_str = request.GET.get('start_date', '')
    end_str = request.GET.get('end_date', '')
    try:
        start = date.fromisoformat(start_str) if start_str else today - timedelta(days=6)
        end = date.fromisoformat(end_str) if end_str else today
    except ValueError:
        start = today - timedelta(days=6)
        end = today
    return start, end


@teacher_required
def attendance_view(request):
    """Record daily attendance for assigned students — teacher only"""
    teacher = request.teacher
    today = date.today()
    start, end = _attendance_date_range(request)
    selected_class = request.GET.get('student_class', '')
    selected_subject = request.GET.get('subject', '')

    students = teacher.students.all()
    if selected_class:
        students = students.filter(student_class=selected_class)

    if request.method == 'POST':
        raw_date = request.POST.get('att_date', today.isoformat())
        try:
            att_date = date.fromisoformat(raw_date)
        except ValueError:
            att_date = today

        for key, val in request.POST.items():
            if key.startswith('att_') and val in ('present', 'absent', 'late', 'excused'):
                s_id = key.replace('att_', '')
                try:
                    stu = Student.objects.get(id=s_id)
                    subj_id = request.POST.get('subject_id', '')
                    subj = Subject.objects.get(id=subj_id) if subj_id else None
                    att, created = Attendance.objects.update_or_create(
                        student=stu, subject=subj, date=att_date,
                        defaults={'status': val, 'recorded_by': teacher}
                    )
                    NotificationService.check_and_alert_low_attendance(stu)
                except (Student.DoesNotExist, Subject.DoesNotExist):
                    continue

        NotificationService.log_activity(request.user, 'add_attendance',
                                          f'Attendance recorded for {att_date}')
        messages.success(request, f'Attendance saved for {att_date}.')
        return redirect(request.path)

    teacher_subjects = teacher.subjects.all()
    teacher_classes = students.values_list('student_class', flat=True).distinct().order_by('student_class')

    # Prefetch today's attendance
    att_map = {}
    period_att = Attendance.objects.filter(student__in=students, date=today)
    for a in period_att:
        att_map[f"{a.student_id}_{a.subject_id or 'none'}"] = a.status

    return render(request, 'core/dashboard/attendance.html', {
        'teacher': teacher, 'students': students,
        'today': today, 'start': start, 'end': end,
        'teacher_subjects': teacher_subjects,
        'available_classes': teacher_classes,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
        'att_map': att_map,
    })


@teacher_required
def attendance_report(request):
    """Attendance report: summary + date-wise breakdown — teacher only"""
    teacher = request.teacher
    students = teacher.students.all()
    selected_class = request.GET.get('student_class', '')
    if selected_class:
        students = students.filter(student_class=selected_class)

    start, end = _attendance_date_range(request)
    records = Attendance.objects.filter(student__in=students, date__range=(start, end))

    per_student = {}
    for stu in students:
        stu_recs = records.filter(student=stu)
        total = stu_recs.count()
        present = stu_recs.filter(status='present').count()
        absent = stu_recs.filter(status='absent').count()
        late = stu_recs.filter(status='late').count()
        pct = round((present / total * 100), 1) if total > 0 else 0
        per_student[stu] = {'total': total, 'present': present, 'absent': absent,
                             'late': late, 'percentage': pct, 'records': list(stu_recs)}

    teacher_classes = students.values_list('student_class', flat=True).distinct().order_by('student_class')
    return render(request, 'core/dashboard/attendance_report.html', {
        'teacher': teacher, 'per_student': per_student,
        'start': start, 'end': end, 'available_classes': teacher_classes,
        'selected_class': selected_class,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  ASSIGNMENTS
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def assignments_view(request):
    """Teacher: list + create + edit assignments"""
    teacher = request.teacher
    assignments = Assignment.objects.filter(teacher=teacher).select_related('subject').order_by('-created_at')

    if request.method == 'POST':
        action = request.POST.get('action', 'create')
        if action == 'create':
            title = request.POST.get('title', '').strip()
            subj_id = request.POST.get('subject', '').strip()
            description = request.POST.get('description', '').strip()
            target_class = request.POST.get('target_class', '').strip()
            target_section = request.POST.get('target_section', '').strip()
            due_date_str = request.POST.get('due_date', '').strip()
            total_marks = int(request.POST.get('total_marks', 10))
            priority = request.POST.get('priority', 'medium')
            # Parse datetime
            from datetime import datetime
            try:
                due_date = datetime.fromisoformat(due_date_str) if due_date_str else None
            except ValueError:
                due_date = None

            if title and subj_id:
                try:
                    subj = Subject.objects.get(id=subj_id)
                    assign = Assignment.objects.create(
                        title=title, description=description, subject=subj, teacher=teacher,
                        target_class=target_class, target_section=target_section,
                        due_date=due_date, total_marks=total_marks, priority=priority,
                        is_published=True,
                    )
                    for sid in request.POST.getlist('target_students'):
                        try:
                            assign.target_students.add(Student.objects.get(id=sid))
                        except Student.DoesNotExist:
                            pass
                    # Notify students
                    targets = assign.target_students.all() or Student.objects.filter(
                        student_class=target_class, section=target_section)
                    NotificationService.create_bulk_notifications(
                        targets,
                        title=f"📚 New Assignment: {title}",
                        message=f"A new assignment has been posted for {subj.name}. Due: {due_date.strftime('%b %d, %Y') if due_date else 'TBD'}",
                        notification_type='info', sender=teacher,
                        link_url='/student-assignments/'
                    )
                    # A1 — email target students
                    try:
                        from core.email_service import send_template_mail
                        email_list = [s.email for s in targets if s.email]
                        if email_list:
                            send_template_mail(
                                'assignment_due.html',
                                {
                                    'title': title,
                                    'subject_name': subj.name,
                                    'target_class': target_class,
                                    'due_date': due_date.strftime('%b %d, %Y') if due_date else 'TBD',
                                    'total_marks': total_marks,
                                    'description': description,
                                    'link_url': '/student-assignments/',
                                },
                                email_list,
                                f"📚 New Assignment: {title}",
                            )
                    except Exception as exc:
                        logger = __import__('logging').getLogger(__name__)
                        logger.warning("assignments_view email failed: %s", exc)
                    messages.success(request, f'Assignment "{title}" created.')
                except Subject.DoesNotExist:
                    messages.error(request, 'Invalid subject.')
            else:
                messages.error(request, 'Title and subject are required.')

        return redirect('core:assignments')

    teacher_subjects = teacher.subjects.all()
    teacher_students = teacher.students.all()
    student_classes = teacher_students.values_list('student_class', flat=True).distinct().order_by('student_class')

    return render(request, 'core/dashboard/assignments.html', {
        'teacher': teacher, 'assignments': assignments,
        'teacher_subjects': teacher_subjects,
        'available_classes': student_classes,
    })


@teacher_required
def assignment_detail_view(request, assignment_id):
    """Teacher: view + grade one assignment"""
    teacher = request.teacher
    assign = get_object_or_404(Assignment, id=assignment_id, teacher=teacher)
    submissions = assign.submissions.select_related('student').all()

    if request.method == 'POST':
        for sub in submissions:
            marks_key = f'marks_{sub.id}'
            feedback_key = f'feedback_{sub.id}'
            if marks_key in request.POST:
                try:
                    marks = float(request.POST[marks_key])
                    sub.marks_obtained = marks
                    sub.teacher_feedback = request.POST.get(feedback_key, '')
                    sub.graded_by = teacher
                    from django.utils import timezone
                    sub.graded_at = timezone.now()
                    sub.status = 'graded'
                    sub.save()
                except (ValueError, TypeError):
                    continue
        messages.success(request, 'Grades and feedback saved.')
        return redirect('core:assignment_detail', assignment_id=assignment_id)

    return render(request, 'core/dashboard/assignment_detail.html', {
        'teacher': teacher, 'assignment': assign, 'submissions': submissions,
    })


@login_required
def student_assignments_view(request):
    """Student: view their assignments with due dates"""
    teacher = get_teacher_profile(request.user)
    is_student = not teacher and not _is_admin_user(request.user)

    student = None
    if is_student:
        student = _get_current_student(request)
    elif teacher:
        student = _get_current_student(request)

    if not student:
        return render(request, 'core/dashboard/student_assignments.html', {
            'error': 'Student profile not found.', 'assignments': [], 'is_student': is_student,
        })

    all_assignments = Assignment.objects.filter(
        is_published=True,
        target_class=student.student_class,
    ).select_related('subject', 'teacher').order_by('-due_date')

    # Filter by section if assignment targets section
    sec_assignments = [a for a in all_assignments
                       if not a.target_section or a.target_section.strip() == ''
                       or a.target_section == student.section]

    student_submitted = {sub.assignment_id for sub in AssignmentSubmission.objects.filter(student=student)}

    return render(request, 'core/dashboard/student_assignments.html', {
        'assignments': sec_assignments, 'student': student, 'is_student': True,
        'submitted_ids': student_submitted,
    })


@login_required
def submit_assignment(request, assignment_id):
    """Student: submit assignment file / text answer"""
    teacher = get_teacher_profile(request.user)
    student = _get_current_student(request)
    if not student:
        messages.error(request, 'Student profile not found.')
        return redirect('core:dashboard')

    assign = get_object_or_404(Assignment, id=assignment_id)

    sub, _ = AssignmentSubmission.objects.get_or_create(
        assignment=assign, student=student,
        defaults={'status': 'missing'}
    )
    if request.method == 'POST':
        text_ans = request.POST.get('text_answer', '').strip()
        up_file = request.FILES.get('file')
        sub.text_answer = text_ans
        if up_file:
            sub.file = up_file
        sub.status = 'submitted'
        sub.save()
        messages.success(request, 'Assignment submitted successfully!')
        return redirect('core:student_assignments')

    return render(request, 'core/dashboard/submit_assignment.html', {
        'assignment': assign, 'submission': sub, 'student': student,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  EXAM MANAGEMENT
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def exam_schedule_view(request):
    """Teacher: create and view exam schedule"""
    teacher = request.teacher
    class_list = teacher.students.values_list('student_class', flat=True).distinct().order_by('student_class')
    exams = Exam.objects.filter(
        subject__in=teacher.subjects.all()
    ).select_related('subject', 'published_by').order_by('exam_date', 'start_time')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        subj_id = request.POST.get('subject', '').strip()
        exam_type = request.POST.get('exam_type', 'terminal')
        target_class = request.POST.get('target_class', '').strip()
        target_section = request.POST.get('target_section', '').strip()
        exam_date = request.POST.get('exam_date', '').strip()
        start = request.POST.get('start_time', '').strip()
        end = request.POST.get('end_time', '').strip()
        total = request.POST.get('total_marks', 100)
        pass_m = request.POST.get('passing_marks', 40)
        venue = request.POST.get('venue', '').strip()
        instructions = request.POST.get('instructions', '').strip()
        is_pub = request.POST.get('is_published') == 'on'

        if title and subj_id and target_class and exam_date:
            try:
                from datetime import datetime
                subj = Subject.objects.get(id=subj_id)
                ed = datetime.strptime(exam_date, '%Y-%m-%d').date()
                st = datetime.strptime(start, '%H:%M').time()
                et = datetime.strptime(end, '%H:%M').time() if end else None
                exam = Exam.objects.create(
                    title=title, exam_type=exam_type, subject=subj,
                    target_class=target_class, target_section=target_section,
                    exam_date=ed, start_time=st, end_time=et,
                    total_marks=int(total), passing_marks=int(pass_m),
                    venue=venue, instructions=instructions,
                    is_published=is_pub, published_by=teacher if is_pub else None,
                )
                messages.success(request, f'Exam "{title}" created.')
            except (Subject.DoesNotExist, ValueError) as e:
                messages.error(request, str(e))
        else:
            messages.error(request, 'Title, subject, class and date are required.')
        return redirect('core:exam_schedule')

    return render(request, 'core/dashboard/exam_schedule.html', {
        'teacher': teacher, 'exams': exams,
        'available_classes': class_list,
        'teacher_subjects': teacher.subjects.all(),
    })


@login_required
def exam_schedule_student(request):
    """Student: view their exam schedule"""
    teacher = get_teacher_profile(request.user)
    student = _get_current_student(request)
    if not student:
        return render(request, 'core/dashboard/exam_schedule_student.html',
                       {'exams': [], 'is_student': True})

    exams = Exam.objects.filter(
        is_published=True,
        target_class=student.student_class,
    ).filter(
        Q(target_section='') | Q(target_section=student.section)
    ).select_related('subject').order_by('exam_date', 'start_time')

    today = date.today()
    upcoming = [e for e in exams if e.exam_date >= today]
    past = [e for e in exams if e.exam_date < today]

    return render(request, 'core/dashboard/exam_schedule_student.html', {
        'upcoming': upcoming, 'past': past, 'student': student, 'is_student': True,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  FEE MANAGEMENT
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def fee_management(request):
    """Teacher: view and record fee payments"""
    teacher = request.teacher
    teacher_students = teacher.students.all().order_by('name')
    selected_class = request.GET.get('student_class', '')
    status_filter = request.GET.get('status', '')

    if selected_class:
        teacher_students = teacher_students.filter(student_class=selected_class)

    qs = Fee.objects.filter(student__in=teacher_students)
    if status_filter:
        qs = qs.filter(status=status_filter)
    fees = qs.select_related('student').order_by('-due_date')

    total_amount = fees.aggregate(t=Sum('amount'))['t'] or 0
    total_paid = fees.aggregate(t=Sum('amount_paid'))['t'] or 0
    total_pending = float(total_amount) - float(total_paid)

    student_classes = teacher.students.values_list('student_class', flat=True).distinct().order_by('student_class')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'record_payment':
            fid = request.POST.get('fee_id')
            paid_amt = request.POST.get('amount_paid', '')
            try:
                fee = Fee.objects.get(id=fid, student__in=teacher.students.all())
                fee.amount_paid = float(paid_amt)
                fee.paid_date = date.today()
                fee.save()
                # A1 — send receipt email to student
                try:
                    student = fee.student
                    if student and student.email:
                        from core.email_service import send_template_mail
                        send_template_mail(
                            'fee_reminder.html',
                            {
                                'student_name': student.name,
                                'guardian_name': 'Guardian',
                                'fee_type': fee.get_fee_type_display(),
                                'amount': f"Rs. {fee.amount}",
                                'due_date': fee.due_date.strftime('%b %d, %Y'),
                                'receipt_number': fee.receipt_number or '',
                                'receipt_url': '/my-fees/',
                            },
                            [student.email],
                            f'Payment Receipt — {fee.get_fee_type_display()}',
                        )
                except Exception as exc:
                    logger = __import__('logging').getLogger(__name__)
                    logger.warning("fee_management email failed: %s", exc)
                messages.success(request, f'Payment of {paid_amt} recorded for {fee.student.name}.')
            except (Fee.DoesNotExist, ValueError):
                messages.error(request, 'Invalid fee record or amount.')
        return redirect('core:fee_management')

    return render(request, 'core/dashboard/fee_management.html', {
        'teacher': teacher, 'fees': fees,
        'student_classes': student_classes,
        'selected_class': selected_class, 'status_filter': status_filter,
        'total_amount': total_amount, 'total_paid': total_paid, 'total_pending': total_pending,
    })


@login_required
def student_fee_view(request):
    """Student: view their own fee records"""
    student = _get_current_student(request)
    if not student:
        return render(request, 'core/dashboard/student_fee.html', {'error': 'Not found.', 'fees': []})

    fees = Fee.objects.filter(student=student).select_related('student').order_by('-due_date')
    total_owed = sum(f.balance for f in fees)
    return render(request, 'core/dashboard/student_fee.html', {
        'fees': fees, 'student': student, 'total_owed': total_owed,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  ANNOUNCEMENTS
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def announcements_view(request):
    """Teacher: post and manage announcements"""
    teacher = request.teacher
    announcements = Announcement.objects.filter(
        Q(published_by=teacher) | Q(published_by__is_admin=True)
    ).order_by('-published_at', '-created_at')[:50]

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        audience = request.POST.get('audience', 'all')
        target_class = request.POST.get('target_class', '')
        priority = request.POST.get('priority', 'medium')
        exp_str = request.POST.get('expires_at', '')

        if title and content:
            ann = Announcement.objects.create(
                title=title, content=content, audience=audience, target_class=target_class,
                priority=priority, published_by=teacher, is_published=True,
                published_at=timezone.now(),
            )
            messages.success(request, 'Announcement published.')
            # A1 — send email to relevant recipients
            try:
                from core.models import Student, Teacher as TeacherModel
                from core.email_service import send_template_mail, notify_teachers_bulk
                now_str = timezone.now().strftime('%b %d, %Y %I:%M %p')
                context = {
                    'title': title,
                    'content': content,
                    'audience': dict(Announcement.AUDIENCE_CHOICES).get(audience, ''),
                    'priority': priority,
                    'attachment_link': ann.attachment.url if ann.attachment else '',
                    'published_by': teacher.get_full_name(),
                    'published_at': now_str,
                }
                subject = f"📢 Announcement: {title}"
                if audience == 'all':
                    student_emails = [s.email for s in Student.objects.all() if s.email]
                    if student_emails:
                        send_template_mail('announcement_alert.html', context, student_emails, subject)
                    all_teachers = TeacherModel.objects.filter(is_active=True).exclude(user=teacher.user)
                    notify_teachers_bulk(all_teachers, subject,
                                         f"New announcement: {title}\n\n{content[:300]}")
                elif audience == 'students':
                    student_emails = [s.email for s in Student.objects.all() if s.email]
                    if student_emails:
                        send_template_mail('announcement_alert.html', context, student_emails, subject)
                elif audience == 'class' and target_class:
                    student_emails = [s.email for s in Student.objects.filter(student_class=target_class) if s.email]
                    if student_emails:
                        send_template_mail('announcement_alert.html', context, student_emails, subject)
                elif audience == 'parents':
                    parents = Parent.objects.filter(student__student_class=target_class
                                                    ) if target_class else Parent.objects.all()
                    parent_emails = [p.email for p in parents if p.email]
                    if parent_emails:
                        send_template_mail('announcement_alert.html', context, parent_emails,
                                           f"📢 Announcement: {title} (for Parents)")
            except Exception as exc:
                logger = __import__('logging').getLogger(__name__)
                logger.warning("announcements_view email failed: %s", exc)
        else:
            messages.error(request, 'Title and content are required.')
        return redirect('core:announcements')

    teacher_classes = teacher.students.values_list('student_class', flat=True).distinct().order_by('student_class')
    return render(request, 'core/dashboard/announcements.html', {
        'teacher': teacher, 'announcements': announcements,
        'available_classes': teacher_classes,
    })


@login_required
def student_announcements_view(request):
    """Student: view all active announcements"""
    student = _get_current_student(request)
    now = timezone.now()
    announcements = Announcement.objects.filter(is_published=True).filter(
        Q(expires_at__gte=now) | Q(expires_at__isnull=True)
    ).filter(
        Q(audience='all') |
        Q(audience='students') |
        Q(audience='class', target_class=student.student_class if student else '')
    ).order_by('-published_at')[:50] if student else Announcement.objects.none()

    return render(request, 'core/dashboard/student_announcements.html', {
        'announcements': announcements, 'student': student, 'is_student': True,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  STUDENT PROFILE & SETTINGS
# ═════════════════════════════════════════════════════════════════════════════

def _get_current_student(request):
    """Return Student instance for the current logged-in user, or None"""
    username = request.user.username.lower().strip()
    try:
        return Student.objects.get(name__iexact=username)
    except Student.DoesNotExist:
        students = Student.objects.filter(name__icontains=username)
        if students.count() == 1:
            return students.first()
    return None


@login_required
def profile_view(request):
    """Student: view and edit their own profile"""
    student = _get_current_student(request)
    if not student:
        messages.error(request, 'Student profile not found.')
        return redirect('core:dashboard')
    return render(request, 'core/dashboard/profile.html', {'student': student, 'is_student': True})


@require_http_methods(["POST"])
@login_required
def update_profile(request):
    """Update student profile fields"""
    student = _get_current_student(request)
    if not student:
        return JsonResponse({'success': False, 'message': 'Profile not found'}, status=404)
    student.email = request.POST.get('email', '') or student.email
    student.phone = request.POST.get('phone', '') or student.phone
    try:
        student.save(update_fields=['email', 'phone'])
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)
    return JsonResponse({'success': True, 'message': 'Profile updated.'})


@login_required
def teachers_list_view(request):
    """Admin: enhanced teachers list with search/filter"""
    teachers = Teacher.objects.select_related('user', 'department').prefetch_related(
        'subjects', 'students', 'levels').filter(is_active=True).order_by('first_name')

    search = request.GET.get('search', '')
    dept_filter = request.GET.get('department', '')
    level_filter = request.GET.get('level', '')

    if search:
        teachers = teachers.filter(
            Q(first_name__icontains=search) | Q(last_name__icontains=search)
            | Q(email__icontains=search) | Q(subjects__name__icontains=search)
        ).distinct()
    if dept_filter:
        teachers = teachers.filter(department_id=dept_filter)
    if level_filter:
        teachers = teachers.filter(levels__code=level_filter).distinct()

    departments = Department.objects.filter(is_active=True).order_by('name')
    return render(request, 'core/dashboard/teachers_list.html', {
        'teachers': teachers, 'departments': departments,
        'search': search, 'dept_filter': dept_filter, 'level_filter': level_filter,
        'is_admin': True,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  STUDENT LIST (ADMIN)
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def students_list_view(request):
    """Admin: enhanced student list with search/filter and manage/delete"""
    students = Student.objects.select_related('academic_year').prefetch_related(
        'teachers', 'history').order_by('student_class', 'name')

    search = request.GET.get('search', '')
    level_filter = request.GET.get('level', '')
    class_filter = request.GET.get('student_class', '')
    section_filter = request.GET.get('section', '')

    if search:
        students = students.filter(
            Q(name__icontains=search) | Q(roll_number__icontains=search)
            | Q(phone__icontains=search)
        )
    if level_filter:
        students = students.filter(level=level_filter)
    if class_filter:
        students = students.filter(student_class=class_filter)
    if section_filter:
        students = students.filter(section=section_filter)

    all_classes = Student.objects.values_list('student_class', flat=True).distinct().order_by('student_class')
    all_sections = Student.objects.values_list('section', flat=True).distinct().order_by('section')

    if request.method == 'POST':
        action = request.POST.get('action', '')
        sid = request.POST.get('student_id', '')
        try:
            stu = Student.objects.get(id=sid)
            if action == 'delete':
                stu_name = stu.name
                stu.delete()
                NotificationService.log_activity(request.user, 'delete_student',
                                                  f'Deleted student: {stu_name}')
                messages.success(request, f'Student "{stu_name}" deleted.')
            elif action == 'manage':
                return redirect('core:manage_parent', student_id=sid)
        except Student.DoesNotExist:
            messages.error(request, 'Student not found.')
        return redirect('core:students_list')

    return render(request, 'core/dashboard/students_list.html', {
        'students': students, 'all_classes': all_classes, 'all_sections': all_sections,
        'search': search, 'level_filter': level_filter, 'class_filter': class_filter,
        'section_filter': section_filter,
    })


@login_required
def add_student_view(request):
    """Admin: manual student creation form"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        roll = request.POST.get('roll_number', '').strip()
        level = request.POST.get('level', 'school')
        cls = request.POST.get('student_class', '').strip()
        section = request.POST.get('section', 'A').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        dob_str = request.POST.get('date_of_birth', '').strip()

        if not name or not cls:
            messages.error(request, 'Name and Class are required.')
        else:
            dob = None
            if dob_str:
                try:
                    dob = date.fromisoformat(dob_str)
                except ValueError:
                    pass
            if roll and Student.objects.filter(roll_number=roll).exists():
                messages.error(request, f'Roll number {roll} already exists.')
            else:
                # Auto-create Django user
                username = name.replace(' ', '').lower()
                if not User.objects.filter(username=username).exists():
                    User.objects.create_user(username=username, password=DEFAULT_PASSWORD,
                                              first_name=name.split()[0], email=email or '')
                Student.objects.create(name=name, roll_number=roll or None, level=level,
                                        student_class=cls, section=section, email=email or None,
                                        phone=phone or None, date_of_birth=dob)
                messages.success(request, f'Student "{name}" added successfully.')
                return redirect('core:students_list')

    return render(request, 'core/dashboard/add_students.html', {})


@login_required
def edit_student_view(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.method == 'POST':
        student.name = request.POST.get('name', student.name)
        student.roll_number = request.POST.get('roll_number', student.roll_number)
        student.level = request.POST.get('level', student.level)
        student.student_class = request.POST.get('student_class', student.student_class)
        student.section = request.POST.get('section', student.section)
        student.email = request.POST.get('email', student.email or '') or None
        student.phone = request.POST.get('phone', student.phone or '') or None
        dob_str = request.POST.get('date_of_birth', '')
        if dob_str:
            try:
                student.date_of_birth = date.fromisoformat(dob_str)
            except ValueError:
                pass
        student.save()
        messages.success(request, f'Student "{student.name}" updated.')
        return redirect('core:students_list')
    return render(request, 'core/dashboard/edit_students.html', {'student': student})


# ═════════════════════════════════════════════════════════════════════════════
#  TEACHER EVALUATION
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def teacher_evaluation_view(request, teacher_id=None):
    """Admin: evaluate a teacher"""
    teacher = get_object_or_404(Teacher, id=teacher_id) if teacher_id else None
    if not teacher:
        messages.error(request, 'Teacher not found.')
        return redirect('core:teacher_performance')

    if request.method == 'POST':
        from core.forms import TeacherEvaluationForm
        form = TeacherEvaluationForm(request.POST)
        if form.is_valid():
            eval_obj = form.save(commit=False)
            eval_obj.teacher = teacher
            eval_obj.evaluator = request.user
            from core.models import AcademicYear
            yr = AcademicYear.objects.filter(is_current=True).first()
            eval_obj.academic_year = yr
            eval_obj.save()
            messages.success(request, f'Evaluation for {teacher.get_full_name()} saved.')
            return redirect('core:teacher_performance')
    else:
        from core.forms import TeacherEvaluationForm
        form = TeacherEvaluationForm()

    return render(request, 'core/dashboard/teacher_evaluation.html', {
        'teacher': teacher, 'form': form,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  STUDENT CHART DATA API (enhanced)
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def student_chart_data(request, student_id):
    """Return chart data for a single student — for student-profile analytics"""
    from django.db.models import Avg
    student = get_object_or_404(Student, id=student_id)
    results = student.result_set.select_related('subject').all()

    subject_labels = []
    subject_scores = []
    for subj in Subject.objects.all():
        s_res = results.filter(subject=subj)
        if s_res.exists():
            obt = sum(r.marks_obtained for r in s_res)
            pos = sum(r.total_marks for r in s_res)
            p = round((obt / pos * 100), 1) if pos > 0 else 0
            subject_labels.append(subj.name)
            subject_scores.append(p)

    term_stats = []
    for term in ['1st', '2nd', '3rd', 'Final']:
        tr = results.filter(terminal=term)
        if tr.exists():
            obt = sum(r.marks_obtained for r in tr)
            pos = sum(r.total_marks for r in tr)
            term_stats.append({
                'terminal': term, 'percentage': round((obt / pos * 100), 1) if pos > 0 else 0,
            })

    return JsonResponse({
        'subject_labels': subject_labels,
        'subject_scores': subject_scores,
        'term_stats': term_stats,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  EXPORT API
# ═════════════════════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def export_students_excel(request):
    """Export filtered student list as Excel"""
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    hdr_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    headers = ['Name', 'Roll No.', 'Class', 'Section', 'Level', 'Phone', 'Email']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = hdr_border
    qs = Student.objects.all().order_by('student_class', 'name')
    search = request.GET.get('search', '')
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(roll_number__icontains=search))
    for r_idx, stu in enumerate(qs, 2):
        ws.cell(r_idx, 1, stu.name)
        ws.cell(r_idx, 2, stu.roll_number or '')
        ws.cell(r_idx, 3, stu.student_class)
        ws.cell(r_idx, 4, stu.section)
        ws.cell(r_idx, 5, stu.get_level_display())
        ws.cell(r_idx, 6, stu.phone or '')
        ws.cell(r_idx, 7, stu.email or '')
    for col in ws.columns:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
                         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
    return resp


# ═════════════════════════════════════════════════════════════════════════════
#  RESULT LOCKING (LOCK GRADE SHEET AFTER PUBLICATION)
# ═════════════════════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["POST"])
def lock_result(request, result_id):
    result = get_object_or_404(Result, id=result_id)
    if get_teacher_profile(request.user):
        teacher = get_teacher_profile(request.user)
        if not teacher.can_access_subject(result.subject):
            return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)
        lock, created = ResultLock.objects.get_or_create(
            result=result,
            defaults={'locked_by': teacher, 'reason': 'Locked by teacher'}
        )
        NotificationService.log_activity(request.user, 'publish_result',
                                          f'Locked result for {result.student.name} / {result.subject.name}')
        # A1 — email student that result was locked
        try:
            student = result.student
            if student and student.email:
                from core.email_service import send_template_mail
                send_template_mail(
                    'result_published.html',
                    {
                        'student_name': student.name,
                        'terminal': result.get_terminal_display(),
                        'results': [{
                            'subject': result.subject.name,
                            'marks_obtained': result.marks_obtained,
                            'total_marks': result.total_marks,
                            'percentage': result.percentage,
                            'grade': result.grade,
                        }],
                        'dashboard_url': '/student-dashboard/',
                    },
                    [student.email],
                    f'Result Locked — {result.subject.name}',
                )
        except Exception as exc:
            logger = __import__('logging').getLogger(__name__)
            logger.warning("lock_result email failed: %s", exc)
        return JsonResponse({'success': True, 'message': 'Result locked', 'created': created})
    return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)


@login_required
@require_http_methods(["POST"])
def lock_all_results(request):
    """Lock all results for a student — admin or teacher who owns them"""
    student_id = request.POST.get('student_id')
    reason = request.POST.get('reason', 'Published by admin')
    if not student_id:
        return JsonResponse({'success': False, 'message': 'student_id required'}, status=400)
    student = get_object_or_404(Student, id=student_id)
    teacher = get_teacher_profile(request.user)
    if teacher:
        for r in student.result_set.filter(subject__in=teacher.subjects.all()):
            ResultLock.objects.get_or_create(
                result=r,
                defaults={'locked_by': teacher, 'reason': reason}
            )
    else:
        for r in student.result_set.all():
            ResultLock.objects.get_or_create(
                result=r,
                defaults={'reason': reason}
            )
    NotificationService.log_activity(request.user, 'publish_result',
                                      f'Locked all results for {student.name}')
    return JsonResponse({'success': True, 'message': f'All results locked for {student.name}'})


@login_required
@require_http_methods(["POST"])
def unlock_result(request, result_id):
    result = get_object_or_404(Result, id=result_id)
    if not _is_admin_user(request.user):
        return JsonResponse({'success': False, 'message': 'Admin rights required'}, status=403)
    ResultLock.objects.filter(result=result).delete()
    NotificationService.log_activity(request.user, 'edit_marks', f'Unlocked result {result.id}')
    return JsonResponse({'success': True, 'message': 'Result unlocked'})


# ═════════════════════════════════════════════════════════════════════════════
#  MARK SHEET PDF DOWNLOAD
# ═════════════════════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["GET"])
def export_mark_sheet_pdf(request, student_id, terminal):
    """Render a styled mark-sheet HTML page and return it as PDF view."""
    from datetime import datetime
    import qrcode
    from io import BytesIO
    import base64

    student = get_object_or_404(Student, id=student_id)
    results = Result.objects.filter(
        student=student, terminal=terminal
    ).select_related('subject').order_by('subject__name')

    teacher = get_teacher_profile(request.user)
    if teacher and not teacher.is_admin:
        assigned = teacher.students.filter(id=student.id).exists()
        if not assigned:
            from django.http import JsonResponse
            return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)

    gs = GradeScale.objects.filter(is_active=True).first()
    rows = []
    total_obt = 0
    total_pos = 0
    for r in results:
        pct = (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0
        rows.append({
            'subject': r.subject.name,
            'code': r.subject.code,
            'marks': r.marks_obtained,
            'total': r.total_marks,
            'percentage': round(pct, 1),
            'grade': gs.get_grade(pct) if gs else '-',
        })
        total_obt += r.marks_obtained
        total_pos += r.total_marks
    overall_pct = (total_obt / total_pos * 100) if total_pos > 0 else 0
    overall_grade = gs.get_grade(overall_pct) if gs else '-'

    # QR code
    college_data = f"SOCH_COLLEGE|{student.name}|{student.roll_number}|{terminal}"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=4, border=2)
    qr.add_data(college_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    context = {
        'student': student,
        'terminal': terminal,
        'rows': rows,
        'total_obt': total_obt,
        'total_pos': total_pos,
        'overall_pct': round(overall_pct, 2),
        'overall_grade': overall_grade,
        'academic_year': f"{datetime.now().year}-{datetime.now().year + 1}",
        'qr_code': qr_b64,
        'doc_id': f"MSH-{datetime.now().strftime('%Y%m%d%H%M')}-{student.id}",
        'generated_at': datetime.now().strftime('%b %d, %Y %I:%M %p'),
    }
    return render(request, 'core/dashboard/mark_sheet_pdf.html', context)


# ═════════════════════════════════════════════════════════════════════════════
#  AI RECOMMENDATION GENERATOR (TEACHER)
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def ai_recommendations(request):
    teacher = request.teacher
    teacher_subjects = teacher.subjects.all()
    teacher_students = teacher.students.all()

    selected_class = request.GET.get('student_class', '').strip()
    selected_subject_id = request.GET.get('subject', '').strip()
    context_key = request.GET.get('context', '').strip()   # 'performance' | 'homework' | 'attendance'

    base_qs = teacher_students
    if selected_class:
        base_qs = base_qs.filter(student_class=selected_class)

    results = Result.objects.filter(
        student__in=base_qs, subject__in=teacher_subjects
    ).select_related('student', 'subject')

    if selected_subject_id:
        results = results.filter(subject_id=selected_subject_id)

    # Build student performance map
    student_map = {}
    for r in results:
        key = r.student.id
        if key not in student_map:
            student_map[key] = {'student': r.student, 'results': [], 'pcts': []}
        pct = (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0
        student_map[key]['results'].append(r)
        student_map[key]['pcts'].append(pct)

    recommendations = []

    for sid, data in student_map.items():
        stu = data['student']
        pcts = data['pcts']
        avg_pct = sum(pcts) / len(pcts) if pcts else 0

        # Determine risk level
        if avg_pct < 40:
            level = 'critical'
        elif avg_pct < 60:
            level = 'warning'
        elif avg_pct < 75:
            level = 'fair'
        else:
            level = 'good'

        # AI-style summary
        if level == 'critical':
            level_intro = f"🚨 {stu.name} is scoring very low — avg {avg_pct:.1f}%. Requires urgent intervention."
            action_title = "🔴 Urgent Action Required"
            action_desc = "Schedule parent meeting, assign remedial homework, and consider peer tutoring."
        elif level == 'warning':
            level_intro = f"⚠️ {stu.name} is underperforming — avg {avg_pct:.1f}%. Needs targeted support."
            action_title = "🟡 Support Needed"
            action_desc = "Schedule additional practice sessions and identify root causes of struggles."
        elif level == 'fair':
            level_intro = f"👍 {stu.name} is doing reasonably well — avg {avg_pct:.1f}%. Improvement possible."
            action_title = "🟢 Encourage Consistency"
            action_desc = "Set small improvement goals and praise progress to build momentum."
        else:
            level_intro = f"🌟 {stu.name} is excelling — avg {avg_pct:.1f}%. Keep challenging."
            action_title = "⭐ Maintain Excellence"
            action_desc = "Provide enrichment material and consider leadership/mentorship opportunities."

        # Weak subjects
        weak_subs = [r for r in data['results']
                     if (r.marks_obtained / r.total_marks * 100) < 50 and r.total_marks > 0]
        weak_names = ', '.join(r.subject.name for r in weak_subs[:3]) if weak_subs else 'None'

        # Study tip based on subject
        study_tips = {
            'Mathematics': 'Daily problem-solving practice (20 min/day). Focus on concept clarity.',
            'English': 'Encourage reading novels and practicing essay writing weekly.',
            'Science': 'Use visual aids, conduct simple experiments at home.',
            'Physics': 'Focus on numerical problems and conceptual understanding.',
            'Chemistry': 'Practice balancing equations, memorize key reactions.',
            'Biology': 'Use diagrams and flashcards for terminology.',
            'History': 'Create timelines, use mnemonics for dates.',
            'Geography': 'Draw maps from memory, use atlas daily.',
            'Computer': 'Encourage hands-on coding practice on online platforms.',
        }
        tip = study_tips.get(weak_subs[0].subject.name if weak_subs else '', 'Encourage regular study and practice.')

        recommendations.append({
            'student': stu,
            'level': level,
            'avg_pct': round(avg_pct, 1),
            'color': {'critical': '#ef4444', 'warning': '#f59e0b', 'fair': '#3b82f6', 'good': '#10b981'}[level],
            'level_intro': level_intro,
            'action_title': action_title,
            'action_desc': action_desc,
            'weak_subjects': weak_subs[:3],
            'weak_names': weak_names,
            'study_tip': tip,
            'total_subjects': len(data['results']),
        })

    available_classes = sorted(teacher_students.values_list('student_class', flat=True).distinct())
    all_subjects_values = list(teacher_subjects.values_list('id', 'name'))

    context = {
        'teacher': teacher,
        'recommendations': recommendations,
        'available_classes': available_classes,
        'selected_class': selected_class,
        'all_subjects_values': all_subjects_values,
        'selected_subject_id': selected_subject_id,
        'context_key': context_key,
        'is_teacher': True,
    }
    return render(request, 'core/dashboard/ai_recommendations.html', context)


# ═════════════════════════════════════════════════════════════════════════════
#  CLASS PERFORMANCE HEATMAP (TEACHER)
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def performance_heatmap(request):
    teacher = request.teacher
    teacher_students = teacher.students.all()
    teacher_subjects = teacher.subjects.all()
    selected_class = request.GET.get('student_class', '').strip()

    if selected_class:
        teacher_students = teacher_students.filter(student_class=selected_class)

    results = Result.objects.filter(
        student__in=teacher_students, subject__in=teacher_subjects
    ).select_related('student', 'subject')

    students = teacher_students.order_by('name')
    subjects = teacher_subjects.order_by('name')

    # Build percentage matrix: {student_id: {subject_id: pct}}
    matrix = {}
    for r in results:
        sid = r.student.id
        subjid = r.subject.id
        pct = (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0
        if sid not in matrix:
            matrix[sid] = {}
        matrix[sid][subjid] = round(pct, 1)

    # Subject averages (per-col)
    subj_avg = {}
    for subj in subjects:
        subj_res = results.filter(subject=subj)
        if subj_res:
            total_pct = sum(
                (r.marks_obtained / r.total_marks * 100) for r in subj_res if r.total_marks > 0
            )
            count = sum(1 for r in subj_res if r.total_marks > 0)
            subj_avg[subj.id] = round(total_pct / count, 1) if count else 0

    # Student row averages (per-row)
    stu_avg = {}
    for stu in students:
        sresults = results.filter(student=stu)
        if sresults:
            total_pct = sum(
                (r.marks_obtained / r.total_marks * 100) for r in sresults if r.total_marks > 0
            )
            count = sum(1 for r in sresults if r.total_marks > 0)
            stu_avg[stu.id] = round(total_pct / count, 1) if count else 0

    available_classes = sorted(teacher_students.values_list('student_class', flat=True).distinct())

    def pct_color(pct):
        if pct >= 80:   return '#10b981'
        if pct >= 60:   return '#3b82f6'
        if pct >= 40:   return '#f59e0b'
        if pct > 0:     return '#ef4444'
        return '#334155'

    row_data = []
    for stu in students:
        row = {
            'student': stu,
            'cells': [],
            'avg': stu_avg.get(stu.id, '-'),
            'avg_color': pct_color(stu_avg.get(stu.id, 0)),
        }
        for subj in subjects:
            pct = matrix.get(stu.id, {}).get(subj.id)
            row['cells'].append({'pct': pct, 'color': pct_color(pct or 0)})
        row_data.append(row)

    context = {
        'teacher': teacher,
        'students': students,
        'subjects': subjects,
        'row_data': row_data,
        'subj_avg': subj_avg,
        'available_classes': available_classes,
        'selected_class': selected_class,
        'pct_color': pct_color,
        'is_teacher': True,
    }
    return render(request, 'core/dashboard/performance_heatmap.html', context)


# ═════════════════════════════════════════════════════════════════════════════
#  IMPROVEMENT TRACKING (TEACHER)
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def improvement_tracking(request):
    teacher = request.teacher
    selected_student_id = request.GET.get('student_id', '').strip()

    teacher_students = teacher.students.all()
    results = Result.objects.filter(
        student__in=teacher_students
    ).select_related('student', 'subject').order_by('created_at')

    improvement_map = {}  # {student_id: {subject_id: [results_sorted]}}

    for r in results:
        sid = r.student.id
        subjid = r.subject.id
        key = (sid, subjid)
        if key not in improvement_map:
            improvement_map[key] = {
                'student': r.student,
                'subject': r.subject,
                'records': [],
            }
        pct = (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0
        improvement_map[key]['records'].append({
            'terminal': r.terminal,
            'pct': round(pct, 1),
            'marks': r.marks_obtained,
            'total': r.total_marks,
            'date': r.created_at.strftime('%b %d, %Y'),
        })

    improvements = []
    for key, data in improvement_map.items():
        pcts = [r['pct'] for r in data['records']]
        first_pct = pcts[0] if pcts else 0
        last_pct = pcts[-1] if pcts else 0
        delta = round(last_pct - first_pct, 1)
        direction = 'up' if delta >= 0 else 'down'
        improvements.append({
            'student': data['student'],
            'subject': data['subject'],
            'records': data['records'],
            'first_pct': first_pct,
            'last_pct': last_pct,
            'delta': delta,
            'direction': direction,
            'total_records': len(data['records']),
        })

    if selected_student_id:
        improvements = [i for i in improvements if i['student'].id == int(selected_student_id)]

    improvements.sort(key=lambda x: (-x['delta']))

    available_students = teacher_students.order_by('name')

    context = {
        'teacher': teacher,
        'improvements': improvements,
        'available_students': available_students,
        'selected_student_id': selected_student_id,
        'is_teacher': True,
    }
    return render(request, 'core/dashboard/improvement_tracking.html', context)


# ═════════════════════════════════════════════════════════════════════════════
#  TEACHER → STUDENT MESSAGING
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def teacher_messages(request):
    teacher = request.teacher
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'send':
            student_id = request.POST.get('recipient_student', '').strip()
            subject = request.POST.get('subject', '').strip()
            user = request.POST.get('user', '').strip()
            if student_id and user:
                try:
                    student = Student.objects.get(id=student_id)
                    # Verify teacher can message this student
                    if teacher.students.filter(id=student.id).exists():
                        Message.objects.create(
                            sender=teacher,
                            sender_student=None,
                            recipient_student=student,
                            subject=subject,
                            user=user,
                        )
                    # Also send notification
                    Notification.objects.create(
                        recipient=student,
                        title=f"📬 New Message: {subject or 'No Subject'}",
                        message=user[:250],
                        notification_type='info',
                        priority='medium',
                        sender=teacher,
                        link_url='/student-messages/',
                    )
                    messages.success(request, f'Message sent to {student.name}.')
                except Student.DoesNotExist:
                    messages.error(request, 'Student not found.')
            else:
                messages.error(request, 'All fields are required.')
        return redirect('core:teacher_messages')

    students = teacher.students.all().order_by('name')
    sent_messages = Message.objects.filter(sender=teacher).select_related('recipient_student').order_by('-created_at')[:50]
    unread_count = Message.objects.filter(recipient_teacher=teacher, status='unread').count()

    return render(request, 'core/dashboard/teacher_messages.html', {
        'teacher': teacher, 'students': students,
        'sent_messages': sent_messages, 'unread_count': unread_count,
        'is_teacher': True,
    })


@teacher_required
def teacher_messages_inbox(request):
    teacher = request.teacher
    msgs = Message.objects.filter(recipient_teacher=teacher).select_related(
        'sender_student', 'recipient_teacher'
    ).order_by('-created_at')
    unread_ids = request.POST.getlist('mark_read') if request.method == 'POST' else []
    if unread_ids:
        Message.objects.filter(id__in=unread_ids, recipient_teacher=teacher).update(status='read')
    return render(request, 'core/dashboard/teacher_messages_inbox.html', {
        'teacher': teacher, 'messages': msgs,
        'is_teacher': True,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  STUDENT MESSAGES INBOX
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def student_messages(request):
    student = _get_current_student(request)
    if not student:
        return render(request, 'core/dashboard/student_messages.html', {
            'error': 'Student profile not found.', 'messages': [], 'is_student': True,
        })

    if request.method == 'POST':
        msg_ids = request.POST.getlist('mark_read')
        if msg_ids:
            Message.objects.filter(id__in=msg_ids, recipient_student=student).update(status='read')
        return redirect('core:student_messages')

    sent = Message.objects.filter(sender_student=student).order_by('-created_at')[:20]
    received = Message.objects.filter(recipient_student=student).order_by('-created_at')
    unread_count = received.filter(status='unread').count()

    return render(request, 'core/dashboard/student_messages.html', {
        'messages': received, 'sent_messages': sent,
        'unread_count': unread_count, 'student': student,
        'is_student': True,
    })


@login_required
@require_http_methods(["POST"])
def send_student_message(request):
    student = _get_current_student(request)
    if not student:
        return JsonResponse({'success': False, 'message': 'Not found'}, status=404)

    teacher_id = request.POST.get('teacher_id', '').strip()
    subject = request.POST.get('subject', '').strip()
    user = request.POST.get('user', '').strip()
    if not teacher_id or not user:
        return JsonResponse({'success': False, 'message': 'Teacher and message required'}, status=400)

    try:
        teacher = Teacher.objects.get(id=teacher_id)
    except Teacher.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Teacher not found'}, status=404)

    if not teacher.students.filter(id=student.id).exists():
        return JsonResponse({'success': False, 'message': 'Not your teacher'}, status=403)

    Message.objects.create(
        sender_student=student,
        sender=None,
        recipient_teacher=teacher,
        subject=subject,
        user=user,
    )
    return JsonResponse({'success': True, 'message': 'Message sent'})


@login_required
def student_parent_message(request):
    student = _get_current_student(request)
    if not student:
        return render(request, 'core/dashboard/student_messages.html', {'error': 'Not found', 'is_student': True})
    parent = getattr(student, 'parent_info', None)
    return render(request, 'core/dashboard/parent_message.html', {
        'student': student, 'parent': parent, 'is_student': True,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  GLOBAL SEARCH
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def global_search(request):
    q = request.GET.get('q', '').strip()
    results = {'students': [], 'teachers': [], 'subjects': []}

    if q and len(q) >= 2:
        results['students'] = (
            Student.objects.filter(
                Q(name__icontains=q) | Q(roll_number__icontains=q)
            )[:10]
        )
        results['teachers'] = (
            Teacher.objects.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(email__icontains=q)
            )[:10]
        )
        results['subjects'] = (
            Subject.objects.filter(
                Q(name__icontains=q) | Q(code__icontains=q)
            )[:10]
        )

    return render(request, 'core/dashboard/global_search.html', {
        'query': q, 'results': results, 'is_admin': True, 'is_teacher': get_teacher_profile(request.user) is not None,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  STUDENT NOTES & SHARING
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def student_notes_view(request):
    """Student: view & manage their own notes"""
    student = _get_current_student(request)
    if not student:
        return render(request, 'core/dashboard/student_notes.html', {
            'error': 'Student profile not found.', 'notes': [], 'is_student': True,
        })
    notes = StudentNote.objects.filter(student=student).order_by('-created_at')
    shared_notes = StudentNote.objects.filter(teacher__in=student.teachers.all()).filter(
        ~Q(student=student)
    ).distinct().order_by('-created_at')[:20]
    return render(request, 'core/dashboard/student_notes.html', {
        'notes': notes, 'shared_notes': shared_notes,
        'student': student, 'is_student': True,
    })


@login_required
def add_student_note(request):
    """Create a personal note (student)"""
    student = _get_current_student(request)
    if not student:
        return JsonResponse({'success': False, 'message': 'Not found'}, status=404)

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        if not title or not content:
            return JsonResponse({'success': False, 'message': 'Title and content required'}, status=400)
        note = StudentNote.objects.create(student=student, title=title, content=content)
        return JsonResponse({'success': True, 'note_id': note.id, 'message': 'Note saved!'})
    return JsonResponse({'success': False, 'message': 'POST required'}, status=405)


@login_required
def share_note(request):
    """Teacher shares a note with one of their students"""
    teacher = get_teacher_profile(request.user)
    if not teacher:
        return JsonResponse({'success': False, 'message': 'Teacher only'}, status=403)

    if request.method == 'POST':
        student_id = request.POST.get('student_id', '').strip()
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        priority = request.POST.get('priority', 'medium').strip()
        if not student_id or not title:
            return JsonResponse({'success': False, 'message': 'Student and title required'}, status=400)
        try:
            student = Student.objects.get(id=student_id)
            if not teacher.students.filter(id=student.id).exists():
                return JsonResponse({'success': False, 'message': 'Not your student'}, status=403)
            note = StudentNote.objects.create(
                student=student, title=title, content=content,
                priority=priority, created_by=teacher,
            )
            return JsonResponse({'success': True, 'note_id': note.id})
        except Student.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Student not found'}, status=404)
    return JsonResponse({'success': False, 'message': 'POST required'}, status=405)


@login_required
@require_http_methods(["POST"])
def delete_student_note(request, note_id):
    student = _get_current_student(request)
    if not student:
        return JsonResponse({'success': False, 'message': 'Not found'}, status=404)
    try:
        note = StudentNote.objects.get(id=note_id, student=student)
        note.delete()
        return JsonResponse({'success': True, 'message': 'Note deleted'})
    except StudentNote.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def teacher_share_note_with_class(request):
    """Teacher creates a note and broadcasts to all students of a class/section"""
    teacher = get_teacher_profile(request.user)
    if not teacher:
        return JsonResponse({'success': False, 'message': 'Teacher only'}, status=403)

    title = request.POST.get('title', '').strip()
    content = request.POST.get('content', '').strip()
    target_class = request.POST.get('target_class', '').strip()
    priority = request.POST.get('priority', 'medium').strip()
    if not title or not target_class:
        return JsonResponse({'success': False, 'message': 'Title and target_class required'}, status=400)

    students = teacher.students.filter(student_class=target_class)
    count = 0
    for stu in students:
        StudentNote.objects.create(
            student=stu, title=title, content=content,
            priority=priority, created_by=teacher,
        )
        count += 1
    return JsonResponse({'success': True, 'shared_with': count, 'message': f'Shared with {count} students'})


@login_required
def api_search(request):
    q = request.GET.get('q', '').strip()
    out = {'students': [], 'teachers': [], 'subjects': []}
    if q and len(q) >= 1:
        for s in Student.objects.filter(name__icontains=q)[:8]:
            out['students'].append({'id': s.id, 'name': s.name, 'class': s.student_class, 'roll': s.roll_number})
        for t in Teacher.objects.filter(first_name__icontains=q)[:5]:
            out['teachers'].append({'id': t.id, 'name': t.get_full_name(), 'dept': t.department.name if t.department else ''})
        for sub in Subject.objects.filter(name__icontains=q)[:5]:
            out['subjects'].append({'id': sub.id, 'name': sub.name, 'code': sub.code})
    return JsonResponse(out)


# ═════════════════════════════════════════════════════════════════════════════
#  REST API — RESULTS
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def api_results(request):
    teacher = get_teacher_profile(request.user)
    qs = Result.objects.select_related('student', 'subject')
    if teacher:
        qs = qs.filter(subject__in=teacher.subjects.all(), student__in=teacher.students.all())
    student_id = request.GET.get('student_id')
    if student_id:
        qs = qs.filter(student_id=student_id)
    results = [{
        'id': r.id, 'student': r.student.name, 'student_id': r.student.id,
        'subject': r.subject.name, 'terminal': r.terminal,
        'marks_obtained': r.marks_obtained, 'total_marks': r.total_marks,
        'percentage': r.percentage, 'grade': r.grade,
    } for r in qs.order_by('-created_at')[:200]]
    return JsonResponse({'results': results})


@login_required
def api_result_detail(request, result_id):
    r = get_object_or_404(Result, id=result_id)
    teacher = get_teacher_profile(request.user)
    if teacher and not (teacher.can_access_subject(r.subject) and teacher.students.filter(id=r.student.id).exists()):
        return JsonResponse({'error': 'Forbidden'}, status=403)
    return JsonResponse({
        'id': r.id, 'student': r.student.name, 'subject': r.subject.name,
        'terminal': r.terminal, 'marks_obtained': r.marks_obtained,
        'total_marks': r.total_marks, 'percentage': r.percentage, 'grade': r.grade,
    })


@login_required
def api_students(request):
    qs = Student.objects.all()
    teacher = get_teacher_profile(request.user)
    if teacher:
        qs = qs.filter(teacher=teacher)
    level = request.GET.get('level', '')
    cls = request.GET.get('student_class', '')
    search_q = request.GET.get('search', '')
    if level:
        qs = qs.filter(level=level)
    if cls:
        qs = qs.filter(student_class=cls)
    if search_q:
        qs = qs.filter(Q(name__icontains=search_q) | Q(roll_number__icontains=search_q))
    data = list(qs.order_by('name')[:100].values('id', 'name', 'roll_number', 'student_class', 'section', 'level', 'email'))
    return JsonResponse({'students': data})


# ═════════════════════════════════════════════════════════════════════════════
#  ADMIN EXPORT — ALL ENTITIES EXCEL
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def admin_export_all(request):
    if not _is_admin_user(request.user):
        return HttpResponse('Forbidden', status=403)

    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr = Font(bold=True, color='ffffff', size=11)
    fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    align = Alignment(horizontal='center')

    def style_header(ws, headers):
        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr; cell.fill = fill; cell.alignment = align; cell.border = thin
        ws.row_dimensions[1].height = 22

    # Sheet 1: Students
    ws1 = wb.create_sheet('Students')
    style_header(ws1, ['ID', 'Name', 'Roll No.', 'Class', 'Section', 'Level', 'Phone', 'Email', 'DOB', 'Admission'])
    for i, stu in enumerate(Student.objects.order_by('student_class', 'name'), 2):
        ws1.cell(i, 1, stu.id)
        ws1.cell(i, 2, stu.name)
        ws1.cell(i, 3, stu.roll_number or '')
        ws1.cell(i, 4, stu.student_class)
        ws1.cell(i, 5, stu.section)
        ws1.cell(i, 6, stu.get_level_display())
        ws1.cell(i, 7, stu.phone or '')
        ws1.cell(i, 8, stu.email or '')
        ws1.cell(i, 9, str(stu.date_of_birth) if stu.date_of_birth else '')
        ws1.cell(i, 10, str(stu.admission_date) if stu.admission_date else '')
    for col in ws1.columns:
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 18

    # Sheet 2: Teachers
    ws2 = wb.create_sheet('Teachers')
    style_header(ws2, ['ID', 'Name', 'Email', 'Phone', 'Department', 'Subjects', 'Active'])
    for i, t in enumerate(Teacher.objects.order_by('first_name'), 2):
        ws2.cell(i, 1, t.id)
        ws2.cell(i, 2, t.get_full_name())
        ws2.cell(i, 3, t.email or '')
        ws2.cell(i, 4, t.phone or '')
        ws2.cell(i, 5, t.department.name if t.department else '')
        ws2.cell(i, 6, ', '.join(s.name for s in t.subjects.all()))
        ws2.cell(i, 7, 'Yes' if t.is_active else 'No')
    for col in ws2.columns:
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 18

    # Sheet 3: Subjects
    ws3 = wb.create_sheet('Subjects')
    style_header(ws3, ['ID', 'Name', 'Code', 'Total Marks', 'Pass Marks', 'Practical'])
    for i, subj in enumerate(Subject.objects.order_by('name'), 2):
        ws3.cell(i, 1, subj.id)
        ws3.cell(i, 2, subj.name)
        ws3.cell(i, 3, subj.code)
        ws3.cell(i, 4, subj.total_marks)
        ws3.cell(i, 5, subj.pass_marks)
        ws3.cell(i, 6, 'Yes' if subj.is_practical else 'No')
    for col in ws3.columns:
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 18

    # Sheet 4: Results
    ws4 = wb.create_sheet('Results')
    style_header(ws4, ['ID', 'Student', 'Subject', 'Terminal', 'Marks', 'Total', '%', 'Grade', 'Created'])
    for i, r in enumerate(Result.objects.select_related('student', 'subject').order_by('-created_at'), 2):
        ws4.cell(i, 1, r.id)
        ws4.cell(i, 2, r.student.name)
        ws4.cell(i, 3, r.subject.name)
        ws4.cell(i, 4, r.terminal)
        ws4.cell(i, 5, r.marks_obtained)
        ws4.cell(i, 6, r.total_marks)
        ws4.cell(i, 7, r.percentage)
        ws4.cell(i, 8, r.grade)
        ws4.cell(i, 9, r.created_at.strftime('%Y-%m-%d'))
    for col in ws4.columns:
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 18

    # Sheet 5: ML Predictions
    ws5 = wb.create_sheet('ML Predictions')
    style_header(ws5, ['ID', 'Student', 'Subject', 'Predicted Grade', 'Confidence %', 'Actual Grade', 'Created'])
    for i, p in enumerate(MLPrediction.objects.select_related('student', 'subject').order_by('-created_at'), 2):
        ws5.cell(i, 1, p.id)
        ws5.cell(i, 2, p.student.name)
        ws5.cell(i, 3, p.subject.name)
        ws5.cell(i, 4, p.predicted_grade)
        ws5.cell(i, 5, p.confidence_score)
        ws5.cell(i, 6, p.actual_grade)
        ws5.cell(i, 7, p.created_at.strftime('%Y-%m-%d'))
    for col in ws5.columns:
        ws5.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="academic_data_export.xlsx"'
    return resp


# ═════════════════════════════════════════════════════════════════════════════
#  ML PREDICTION API
# ═════════════════════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["POST"])
def create_ml_prediction(request):
    if not _is_admin_user(request.user):
        return JsonResponse({'success': False, 'message': 'Admin only'}, status=403)
    student_id = request.POST.get('student_id', '').strip()
    subject_id = request.POST.get('subject_id', '').strip()
    predicted_grade = request.POST.get('predicted_grade', '').strip()
    confidence_str = request.POST.get('confidence_score', '0').strip()

    if not student_id or not subject_id:
        return JsonResponse({'success': False, 'message': 'Student and subject required'}, status=400)

    try:
        student = Student.objects.get(id=student_id)
        subject = Subject.objects.get(id=subject_id)
        confidence = float(confidence_str)
    except (Student.DoesNotExist, Subject.DoesNotExist, ValueError):
        return JsonResponse({'success': False, 'message': 'Invalid input'}, status=400)

    prediction, created = MLPrediction.objects.update_or_create(
        student=student, subject=subject, model_version='v1',
        defaults={'predicted_grade': predicted_grade, 'confidence_score': confidence}
    )
    return JsonResponse({'success': True, 'created': created, 'predicted_grade': predicted_grade})


@login_required
@require_http_methods(["GET"])
def get_ml_predictions(request):
    student_id = request.GET.get('student_id', '').strip()
    subject_id = request.GET.get('subject_id', '').strip()

    qs = MLPrediction.objects.select_related('student', 'subject')
    if student_id:
        qs = qs.filter(student_id=student_id)
    if subject_id:
        qs = qs.filter(subject_id=subject_id)

    data = []
    for p in qs.order_by('-created_at')[:100]:
        data.append({
            'id': p.id,
            'student': p.student.name,
            'subject': p.subject.name,
            'predicted_grade': p.predicted_grade,
            'confidence': p.confidence_score,
            'actual_grade': p.actual_grade,
            'created_at': p.created_at.strftime('%Y-%m-%d'),
        })
    return JsonResponse({'predictions': data})


# ═════════════════════════════════════════════════════════════════════════════
#  AI CONTEXTUAL HELP  — public API
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def ai_contextual_help(request):
    teacher = get_teacher_profile(request.user)
    action_type = request.GET.get('action', '')

    if action_type == 'add_marks':
        tip = (
            'Pick subject and terminal first. Scroll to find the student, '
            'enter marks and total. The system auto-calculates grade points. '
            'Click Save All Marks to commit.'
        )
    elif action_type == 'attendance':
        tip = (
            'Today\'s date is pre-filled. Select subject to filter by period. '
            'Mark each student Present / Absent / Late / Excused, then hit Save.'
        )
    elif action_type == 'assignment':
        tip = (
            'Fill title, description, due date, and priority. '
            'Save and point students to the Assignments page to submit.'
        )
    elif action_type == 'exam':
        tip = (
            'Create exam with date/time/venue. Check "Published" to release instantly. '
            'Students see their exam schedule on the Exams page.'
        )
    else:
        tip = (
            'Use the dashboard for quick overview. Navigate sidebar for a specific task. '
            'AI Recommendations and Performance Heatmap highlight weak spots quickly.'
        )
    return JsonResponse({'tip': tip, 'action': action_type})


# ═════════════════════════════════════════════════════════════════════════════
#  A7 — BACKUP & RESTORE (Admin)
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def backup_restore(request):
    """Admin: list existing backups, create new backup, restore from file."""
    import os as _os
    backups = SystemBackup.objects.all().order_by('-created_at')

    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'restore':
            restore_file = request.FILES.get('restore_file')
            if restore_file and restore_file.name.endswith('.json'):
                output_dir = _os.path.join(settings.BASE_DIR, 'backups')
                _os.makedirs(output_dir, exist_ok=True)
                saved_path = _os.path.join(output_dir, restore_file.name)
                with open(saved_path, 'wb+') as dest:
                    for chunk in restore_file.chunks():
                        dest.write(chunk)
                try:
                    import json, io
                    from django.apps import apps
                    with open(saved_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    created_count = 0
                    for entry in data:
                        model_label = entry.get('model', '')
                        if '.' not in model_label:
                            continue
                        app_label, model_name = model_label.split('.', 1)
                        try:
                            model_class = apps.get_model(app_label, model_name)
                        except LookupError:
                            continue
                        try:
                            obj = model_class(**entry['fields'])
                            obj.pk = entry.get('pk')
                            obj.save()
                            created_count += 1
                        except Exception:
                            pass
                    backup_kb = _os.path.getsize(saved_path) if _os.path.exists(saved_path) else 0
                    SystemBackup.objects.create(
                        filename=restore_file.name,
                        size_bytes=backup_kb,
                        backup_type='django_export',
                        triggered_by=request.user if request.user.is_authenticated else None,
                        status='success',
                    )
                    messages.success(request, f'Restore complete. {created_count} records restored.')
                except Exception as exc:
                    messages.error(request, f'Restore failed: {exc}')
            else:
                messages.error(request, 'Please upload a valid JSON backup file.')
        return redirect('core:backup_restore')

    return render(request, 'core/dashboard/backup_restore.html', {
        'backups': backups,
    })


@_admin_required
def download_backup(request, backup_id):
    """Download a specific backup file."""
    import os as _os
    backup = get_object_or_404(SystemBackup, id=backup_id)
    backup_path = _os.path.join(settings.BASE_DIR, 'backups', backup.filename)
    if not _os.path.exists(backup_path):
        return HttpResponse('Backup file not found on disk.', status=404)
    try:
        with open(backup_path, 'rb') as f:
            resp = HttpResponse(f.read(), content_type='application/octet-stream')
            resp['Content-Disposition'] = f'attachment; filename="{backup.filename}"'
            return resp
    except OSError as exc:
        return HttpResponse(f'Error reading file: {exc}', status=500)


@_admin_required
def api_create_backup(request):
    """API endpoint: trigger a new database backup."""
    import os, json, datetime
    from django.core import serializers
    from django.apps import apps
    from django.core.management import call_command
    output_dir = os.path.join(settings.BASE_DIR, 'backups')
    try:
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"backup_{timestamp}.json"
        filepath = os.path.join(output_dir, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            models_list = [
                'core.Subject', 'core.EducationLevel', 'core.Semester',
                'core.AcademicYear', 'core.Department', 'core.GradeScale',
                'core.Teacher', 'core.Parent', 'core.Student',
                'core.StudentAcademicHistory', 'core.Attendance',
                'core.Assignment', 'core.AssignmentSubmission',
                'core.Exam', 'core.Fee', 'core.Announcement',
                'core.ActivityLog', 'core.TeacherEvaluation',
                'core.Notification', 'core.CourseMaterial',
                'core.Result', 'core.ResultLock', 'core.Message',
                'core.StudentNote', 'core.MLPrediction',
            ]
            counts = {}
            for label in models_list:
                try:
                    model_class = apps.get_model(label)
                    objs = model_class.objects.all()
                    counts[label] = objs.count()
                    if objs.exists():
                        serializers.serialize('json', objs, stream=f, indent=2)
                        f.write('\n')
                except LookupError:
                    pass
        size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
        SystemBackup.objects.create(
            filename=filename,
            size_bytes=size,
            backup_type='django_export',
            triggered_by=request.user if request.user.is_authenticated else None,
            status='success',
        )
        call_command('backup_db', output=output_dir)
        return JsonResponse({'success': True, 'filename': filename, 'size': size, 'counts': counts})
    except Exception as exc:
        SystemBackup.objects.create(
            filename=filename if 'filename' in dir() else 'unknown',
            size_bytes=0,
            backup_type='django_export',
            triggered_by=request.user if request.user.is_authenticated else None,
            status='failed',
        )
        return JsonResponse({'success': False, 'error': str(exc)}, status=500)


# ═════════════════════════════════════════════════════════════════════════════
#  A11 — PARENT / GUARDIAN PORTAL
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def parent_portal(request):
    """Parent portal: view their ward's academic summary."""
    if not hasattr(request.user, 'parent_profile'):
        messages.error(request, 'Parent portal access only. No linked parent account found.')
        return redirect('core:dashboard')
    parent_user = request.user.parent_profile
    parent = parent_user.parent
    student = parent.student

    results = student.result_set.select_related('subject').all()
    total_obt = sum(r.marks_obtained for r in results)
    total_pos = sum(r.total_marks for r in results)
    overall_pct = round((total_obt / total_pos * 100), 1) if total_pos > 0 else 0

    fees = Fee.objects.filter(student=student).order_by('-due_date')
    outstanding = [f for f in fees if int(f.balance) > 0]
    att_pct = student.attendance_percentage

    context = {
        'student': student,
        'parent': parent,
        'results': results,
        'fees': fees,
        'outstanding': outstanding,
        'att_pct': att_pct,
        'overall_pct': overall_pct,
    }
    return render(request, 'core/dashboard/parent_portal.html', context)


@login_required
def parent_verification(request):
    """Simple token-based verification for a parent user account."""
    if request.method == 'POST':
        token = request.POST.get('token', '').strip()
        try:
            pu = ParentUser.objects.get(user=request.user, verification_token=token)
            pu.is_verified = True
            pu.save(update_fields=['is_verified'])
            messages.success(request, 'Account verified successfully.')
            return redirect('core:parent_portal')
        except ParentUser.DoesNotExist:
            messages.error(request, 'Invalid verification token.')
    return render(request, 'core/dashboard/parent_verification.html', {
        'parent_user': getattr(request.user, 'parent_profile', None),
    })


@login_required
def parent_api_student_info(request):
    """JSON: return the ward's basic info scoped to the authenticated parent."""
    if not hasattr(request.user, 'parent_profile'):
        return JsonResponse({'success': False, 'message': 'Forbidden'}, status=403)
    student = request.user.parent_profile.parent.student
    return JsonResponse({
        'success': True,
        'student': {
            'id': student.id,
            'name': student.name,
            'class': student.student_class,
            'section': student.section,
            'roll_number': student.roll_number,
            'level': student.get_level_display(),
        }
    })


@login_required
def parent_api_fees(request):
    """JSON: fee records for the parent's ward."""
    if not hasattr(request.user, 'parent_profile'):
        return JsonResponse({'success': False, 'message': 'Forbidden'}, status=403)
    student = request.user.parent_profile.parent.student
    fees = Fee.objects.filter(student=student).order_by('-due_date')
    data = [{'id': f.id, 'type': f.get_fee_type_display(),
             'amount': str(f.amount), 'paid': str(f.amount_paid),
             'balance': str(f.balance), 'due_date': str(f.due_date),
             'status': f.get_status_display()} for f in fees]
    return JsonResponse({'success': True, 'fees': data})


@login_required
def parent_api_attendance(request):
    """JSON: attendance summary for the parent's ward."""
    if not hasattr(request.user, 'parent_profile'):
        return JsonResponse({'success': False, 'message': 'Forbidden'}, status=403)
    student = request.user.parent_profile.parent.student
    records = Attendance.objects.filter(student=student).order_by('-date')[:50]
    present = sum(1 for r in records if r.status == 'present')
    data = [{'date': str(r.date), 'status': r.get_status_display(),
             'subject': r.subject.name if r.subject else ''} for r in records]
    pct = round((present / records.count() * 100), 1) if records.count() else 0
    return JsonResponse({'success': True, 'total': records.count(),
                         'present': present, 'percentage': pct, 'records': data})


@login_required
def parent_api_results(request):
    """JSON: result summary for the parent's ward."""
    if not hasattr(request.user, 'parent_profile'):
        return JsonResponse({'success': False, 'message': 'Forbidden'}, status=403)
    student = request.user.parent_profile.parent.student
    results = Result.objects.filter(student=student).select_related('subject').order_by('subject__name')
    data = [{'subject': r.subject.name, 'terminal': r.get_terminal_display(),
             'marks': r.marks_obtained, 'total': r.total_marks,
             'percentage': r.percentage, 'grade': r.grade} for r in results]
    return JsonResponse({'success': True, 'results': data})


# ═════════════════════════════════════════════════════════════════════════════
#  A13 — RESULT PUBLICATION WORKFLOW
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def publish_session_list(request):
    """Admin: list all result publish sessions."""
    sessions = ResultPublishSession.objects.all().order_by('-created_at')
    subject_id = request.GET.get('subject', '')
    status_filter = request.GET.get('status', '')
    target_class = request.GET.get('target_class', '')
    if subject_id:
        sessions = sessions.filter(subject_id=subject_id)
    if status_filter:
        sessions = sessions.filter(status=status_filter)
    if target_class:
        sessions = sessions.filter(target_class=target_class)
    subjects = Subject.objects.all().order_by('name')
    classes = Student.objects.values_list('student_class', flat=True).distinct().order_by('student_class')
    return render(request, 'core/dashboard/publish_session_list.html', {
        'sessions': sessions,
        'subjects': subjects,
        'classes': classes,
        'subject_id': subject_id,
        'status_filter': status_filter,
        'target_class': target_class,
        'status_choices': ResultPublishSession.PUBLISH_STATUS_CHOICES,
    })


@_admin_required
def publish_session_create(request):
    """Admin: create a new publish session — auto-collects Result objects."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        subject_id = request.POST.get('subject', '').strip()
        target_class = request.POST.get('target_class', '').strip()
        academic_year_id = request.POST.get('academic_year', '').strip()
        terminal = request.POST.get('terminal', '1st')
        if not name or not subject_id or not target_class:
            messages.error(request, 'Name, subject and target class are required.')
            return redirect('core:publish_session_create')
        try:
            subject = Subject.objects.get(id=subject_id)
            acad_year = AcademicYear.objects.filter(id=academic_year_id).first() if academic_year_id else None
        except Subject.DoesNotExist:
            messages.error(request, 'Invalid subject.')
            return redirect('core:publish_session_create')
        session = ResultPublishSession.objects.create(
            name=name,
            subject=subject,
            target_class=target_class,
            academic_year=acad_year,
            created_by=request.user,
        )
        results = Result.objects.filter(
            subject=subject, terminal=terminal,
            student__student_class=target_class,
        ).select_related('student')
        entry_count = 0
        for result in results:
            ResultSessionEntry.objects.create(session=session, result=result)
            entry_count += 1
        messages.success(request, f'Session "{name}" created with {entry_count} result entries.')
        return redirect('core:publish_session_detail', session_id=session.id)
    subjects = Subject.objects.all().order_by('name')
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    classes = Student.objects.values_list('student_class', flat=True).distinct().order_by('student_class')
    terminals = [('1st', '1st Terminal'), ('2nd', '2nd Terminal'), ('3rd', '3rd Terminal'), ('Final', 'Final Terminal')]
    return render(request, 'core/dashboard/publish_session_create.html', {
        'subjects': subjects, 'academic_years': academic_years,
        'classes': classes, 'terminals': terminals,
    })


@_admin_required
def publish_session_detail(request, session_id):
    """Admin: view/edit individual publish session — lock/unlock results."""
    session = get_object_or_404(ResultPublishSession, id=session_id)
    entries = session.entries.select_related('result__student', 'result__subject').all()
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'lock_entry':
            entry_id = request.POST.get('entry_id')
            remarks = request.POST.get('remarks', '')
            try:
                entry = ResultSessionEntry.objects.get(id=entry_id, session=session)
                entry.is_locked = True
                if remarks:
                    entry.result.remarks = remarks
                    entry.result.save()
                entry.save()
                messages.success(request, 'Entry locked.')
            except ResultSessionEntry.DoesNotExist:
                messages.error(request, 'Entry not found.')
        elif action == 'unlock_entry':
            entry_id = request.POST.get('entry_id')
            try:
                entry = ResultSessionEntry.objects.get(id=entry_id, session=session)
                entry.is_locked = False
                entry.save()
                messages.success(request, 'Entry unlocked.')
            except ResultSessionEntry.DoesNotExist:
                messages.error(request, 'Entry not found.')
        elif action == 'lock_all':
            entries.update(is_locked=True)
            messages.success(request, 'All entries locked.')
        elif action == 'unlock_all':
            entries.update(is_locked=False)
            messages.success(request, 'All entries unlocked.')
        elif action == 'save_remarks':
            entry_id = request.POST.get('entry_id')
            remarks = request.POST.get('remarks', '')
            try:
                entry = ResultSessionEntry.objects.get(id=entry_id, session=session)
                entry.result.remarks = remarks
                entry.result.save()
                messages.success(request, 'Remarks saved.')
            except ResultSessionEntry.DoesNotExist:
                messages.error(request, 'Entry not found.')
        elif action == 'add_remark':
            entry_id = request.POST.get('entry_id')
            remarks = request.POST.get('remarks', '').strip()
            try:
                entry = ResultSessionEntry.objects.get(id=entry_id, session=session)
                result = entry.result
                result.remarks = remarks
                result.save(update_fields=['remarks'])
                messages.success(request, f'Remark added for {result.student.name}.')
            except ResultSessionEntry.DoesNotExist:
                messages.error(request, 'Entry not found.')
        return redirect('core:publish_session_detail', session_id=session.id)
    return render(request, 'core/dashboard/publish_session_detail.html', {
        'session': session,
        'entries': entries,
    })


def _set_session_status(session, new_status, request, by_field, reason_field=''):
    session.status = new_status
    if by_field:
        setattr(session, by_field, request.user)
    if reason_field:
        reason_key = reason_field
        setattr(session, reason_key, request.POST.get('rejection_reason', ''))
    session.save()


@_admin_required
def publish_session_submit_for_review(request, session_id):
    """DRAFT → UNDER_REVIEW"""
    session = get_object_or_404(ResultPublishSession, id=session_id)
    if session.status != 'draft':
        messages.error(request, 'Only Draft sessions can be submitted for review.')
        return redirect('core:publish_session_detail', session_id=session.id)
    session.status = 'under_review'
    session.reviewed_by = request.user
    session.save(update_fields=['status', 'reviewed_by'])
    messages.success(request, 'Session submitted for review.')
    return redirect('core:publish_session_detail', session_id=session.id)


@_admin_required
def publish_session_approve(request, session_id):
    """UNDER_REVIEW → APPROVED"""
    session = get_object_or_404(ResultPublishSession, id=session_id)
    if session.status != 'under_review':
        messages.error(request, 'Only Under Review sessions can be approved.')
        return redirect('core:publish_session_detail', session_id=session.id)
    session.status = 'approved'
    session.approved_by = request.user
    session.save(update_fields=['status', 'approved_by'])
    messages.success(request, 'Session approved.')
    return redirect('core:publish_session_detail', session_id=session.id)


@_admin_required
def publish_session_publish(request, session_id):
    """APPROVED → PUBLISHED — triggers notification burst to students"""
    session = get_object_or_404(ResultPublishSession, id=session_id)
    if session.status != 'approved':
        messages.error(request, 'Only Approved sessions can be published.')
        return redirect('core:publish_session_detail', session_id=session.id)
    session.status = 'published'
    session.published_at = timezone.now()
    session.save(update_fields=['status', 'published_at'])
    NotificationService.log_activity(
        request.user, 'publish_result',
        f'Published session: {session.name}',
    )
    # Notification burst to all students in this session
    try:
        from core.email_service import send_template_mail, notify_student
        affected_students = {e.result.student for e in session.entries.select_related('result__student')}
        for student in affected_students:
            student_results = [
                {
                    'subject': e.result.subject.name,
                    'marks_obtained': e.result.marks_obtained,
                    'total_marks': e.result.total_marks,
                    'percentage': e.result.percentage,
                    'grade': e.result.grade,
                }
                for e in session.entries.select_related('result__subject').filter(result__student=student)
            ]
            notify_student(
                student,
                f"✅ Results Published — {session.name}",
                f"Your results for {session.subject.name} ({session.name}) have been published. Check your dashboard.",
                link_url='/student-dashboard/',
                notification_type='success',
            )
            if student.email:
                send_template_mail(
                    'result_published.html',
                    {
                        'student_name': student.name,
                        'terminal': session.name,
                        'results': student_results,
                        'dashboard_url': '/student-dashboard/',
                    },
                    [student.email],
                    f'✅ Results Published — {session.name}',
                )
    except Exception as exc:
        logger = __import__('logging').getLogger(__name__)
        logger.warning("publish notification burst failed: %s", exc)
    messages.success(request, f'Session "{session.name}" published. {len(affected_students)} students notified.')
    return redirect('core:publish_session_list')


@_admin_required
def publish_session_reject(request, session_id):
    """Any → REJECTED with reason."""
    session = get_object_or_404(ResultPublishSession, id=session_id)
    reason = request.POST.get('rejection_reason', '').strip()
    if not reason:
        messages.error(request, 'Rejection reason is required.')
        return redirect('core:publish_session_detail', session_id=session.id)
    session.status = 'rejected'
    session.rejection_reason = reason
    session.save(update_fields=['status', 'rejection_reason'])
    messages.success(request, 'Session rejected.')
    return redirect('core:publish_session_detail', session_id=session.id)


@_admin_required
def publish_session_archive(request, session_id):
    """PUBLISHED → ARCHIVED"""
    session = get_object_or_404(ResultPublishSession, id=session_id)
    if session.status != 'published':
        messages.error(request, 'Only Published sessions can be archived.')
        return redirect('core:publish_session_detail', session_id=session.id)
    session.status = 'archived'
    session.save(update_fields=['status'])
    messages.success(request, 'Session archived.')
    return redirect('core:publish_session_list')


# ═════════════════════════════════════════════════════════════════════════════
#  A2 — LICENSE KEY SYSTEM (Admin)
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def license_config(request):
    """Admin: view current license, create new key, deactivate existing."""
    active_license = LicenseKey.objects.filter(is_active=True).first()
    all_licenses = LicenseKey.objects.all().order_by('-issued_at')[:20]
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'create':
            name = request.POST.get('institution_name', '').strip()
            max_teachers = int(request.POST.get('max_teachers', 5) or 5)
            max_students = int(request.POST.get('max_students', 500) or 500)
            max_branches = int(request.POST.get('max_branches', 1) or 1)
            expires_str = request.POST.get('expires_at', '')
            notes = request.POST.get('notes', '').strip()
            expires_at = None
            if expires_str:
                try:
                    from datetime import date as _date
                    expires_at = _date.fromisoformat(expires_str)
                except ValueError:
                    messages.error(request, 'Invalid expiry date.')
                    return redirect('core:license_config')
            # Deactivate existing active license
            LicenseKey.objects.filter(is_active=True).update(is_active=False)
            lk = LicenseKey.objects.create(
                institution_name=name or 'Default Institution',
                max_teachers=max_teachers, max_students=max_students,
                max_branches=max_branches, expires_at=expires_at,
                created_by=request.user, notes=notes,
            )
            lk.refresh_counts()
            messages.success(request, f'New license key created: {lk.key[:20]}...')
            return redirect('core:license_config')
        elif action == 'deactivate':
            lk_id = request.POST.get('license_id')
            try:
                lk = LicenseKey.objects.get(id=lk_id)
                lk.is_active = False
                lk.save(update_fields=['is_active'])
                messages.success(request, f'License {lk.key[:20]}... deactivated.')
            except LicenseKey.DoesNotExist:
                messages.error(request, 'License not found.')
        elif action == 'set_system_key':
            key = request.POST.get('system_key', '').strip()
            SystemConfig.set('license_key', key)
            retrieved = SystemConfig.get('license_key', '')
            messages.success(request, f'System license key saved. Current stored value: {retrieved}')
        return redirect('core:license_config')
    stored_key = SystemConfig.get('license_key', '')
    return render(request, 'core/dashboard/license_config.html', {
        'active_license': active_license,
        'all_licenses': all_licenses,
        'stored_key': stored_key,
    })


@login_required
def license_status(request):
    """Lightweight JSON endpoint indicating the licensing status of the system."""
    active = LicenseKey.objects.filter(is_active=True).first()
    today = __import__('datetime').date.today()
    expired = active and active.expires_at and active.expires_at < today if active else True
    if active and not expired:
        ok, detail = active.check_limits()
    else:
        ok, detail = False, "No active / valid license"
    return JsonResponse({
        'licensed': bool(active and not expired and ok),
        'license_key': active.key[:20] + '...' if active else '',
        'institution': active.institution_name if active else '',
        'expires_at': str(active.expires_at) if active else '',
        'status': detail,
        'teachers_used': active.current_teachers if active else 0,
        'students_used': active.current_students if active else 0,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  A6 — RBAC ROLE & PERMISSION MANAGER (Admin UI)
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def role_management(request):
    """Admin: manage roles, permissions, and user-role assignments."""
    roles = UserRole.objects.all().order_by('code')
    users = User.objects.all().order_by('username')
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'assign_role':
            user_id = request.POST.get('user_id')
            role_id = request.POST.get('role_id')
            try:
                user = User.objects.get(id=user_id)
                role = UserRole.objects.get(id=role_id)
                profile, created = UserProfile.objects.update_or_create(
                    user=user, defaults={'role': role},
                )
                messages.success(request, f'Assigned {role.name} to {user.username}.')
            except (User.DoesNotExist, UserRole.DoesNotExist):
                messages.error(request, 'Invalid user or role.')
        return redirect('core:role_management')
    user_roles = {}
    for up in UserProfile.objects.select_related('role', 'user').all():
        rid = up.role_id
        if rid not in user_roles:
            user_roles[rid] = []
        user_roles[rid].append(up)
    return render(request, 'core/dashboard/role_management.html', {
        'roles': roles, 'users': users, 'user_roles': user_roles,
    })


@_admin_required
def user_permissions(request):
    """Admin: per-user permission overview."""
    profiles = UserProfile.objects.select_related('user', 'role').all().order_by('user__username')
    perm_fields = [
        'can_manage_students', 'can_manage_teachers', 'can_manage_subjects',
        'can_manage_fees', 'can_manage_exams', 'can_manage_attendance',
        'can_manage_materials', 'can_view_reports', 'can_manage_system',
    ]
    return render(request, 'core/dashboard/user_permissions.html', {
        'profiles': profiles, 'perm_fields': perm_fields,
    })



# ═════════════════════════════════════════════════════════════════════════════
#  MILESTONE ACHIEVEMENT CHECK
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def milestone_check(request):
    """Check if student has reached a milestone (100% attendance, 90%+ avg, etc.)"""
    from core.notification_service import NotificationService
    student = _get_current_student(request)

    milestones_achieved = []

    if student:
        # Check attendance milestone
        att_pct = student.attendance_percentage
        if att_pct >= 95 and att_pct < 100:
            milestones_achieved.append({
                'icon': '📋',
                'title': 'Perfect Attendance',
                'desc': f'You have {att_pct}% attendance this year!',
            })

        # Check academic milestone
        results = student.result_set.all()
        if results.exists():
            total_obt = sum(r.marks_obtained for r in results)
            total_pos = sum(r.total_marks for r in results)
            avg_pct = (total_obt / total_pos * 100) if total_pos > 0 else 0
            if avg_pct >= 90:
                milestones_achieved.append({
                    'icon': '🌟',
                    'title': 'Outstanding Grades',
                    'desc': f'Your average is {avg_pct:.1f}% — excellent!',
                })
            elif avg_pct >= 75:
                milestones_achieved.append({
                    'icon': '👍',
                    'title': 'Good Progress',
                    'desc': f'Your average is {avg_pct:.1f}%.',
                })

        # Check for unread assignment (deadline approach)
        from core.models import Assignment, AssignmentSubmission
        upcoming = Assignment.objects.filter(
            is_published=True,
            target_class=student.student_class,
            due_date__gte=timezone.now(),
        ).exclude(
            id__in=AssignmentSubmission.objects.filter(student=student).values_list('assignment_id', flat=True)
        )
        if upcoming.exists():
            next_due = upcoming.order_by('due_date').first()
            milestones_achieved.append({
                'icon': '📚',
                'title': 'Pending Assignment',
                'desc': f'"{next_due.title}" due {next_due.due_date.strftime("%b %d")}.',
            })

    return JsonResponse({'milestones': milestones_achieved})


# ═════════════════════════════════════════════════════════════════════════════
# T2 — LESSON PLAN & SYLLABUS TRACKER
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def lesson_plans(request):
    """Teacher: list, filter, and create lesson plans"""
    teacher = request.teacher
    is_admin = teacher.is_admin

    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'create':
            subject_id = request.POST.get('subject', '').strip()
            title = request.POST.get('title', '').strip()
            class_section = request.POST.get('class_section', '').strip()
            chapter_topic = request.POST.get('chapter_topic', '').strip()
            planned_date = request.POST.get('planned_date', '').strip()
            teaching_method = request.POST.get('teaching_method', 'lecture')
            duration = request.POST.get('duration_minutes', '45')
            description = request.POST.get('description', '').strip()
            learning_outcomes = request.POST.get('learning_outcomes', '').strip()
            resources_used = request.POST.get('resources_used', '').strip()
            status = request.POST.get('status', 'planned')
            remarks = request.POST.get('remarks', '').strip()

            if not (subject_id and title and class_section and chapter_topic and planned_date):
                messages.error(request, 'Subject, title, class/section, chapter/topic, and planned date are required.')
            else:
                try:
                    from datetime import date
                    subj = Subject.objects.get(id=subject_id)
                    if not is_admin and not teacher.can_access_subject(subj):
                        messages.error(request, 'You can only create lesson plans for your assigned subjects.')
                    else:
                        LessonPlan.objects.create(
                            teacher=teacher,
                            subject=subj,
                            title=title,
                            description=description,
                            class_section=class_section,
                            chapter_topic=chapter_topic,
                            planned_date=date.fromisoformat(planned_date),
                            duration_minutes=int(duration) if duration else 45,
                            teaching_method=teaching_method,
                            learning_outcomes=learning_outcomes,
                            resources_used=resources_used,
                            status=status,
                            remarks=remarks,
                        )
                        messages.success(request, f'Lesson plan "{title}" created successfully.')
                except Subject.DoesNotExist:
                    messages.error(request, 'Invalid subject selected.')
            return redirect('core:lesson_plans')

    # GET — filter by subject/class/status
    selected_subject = request.GET.get('subject', '')
    selected_class = request.GET.get('class_section', '')
    selected_status = request.GET.get('status', '')

    plans = LessonPlan.objects.select_related('teacher', 'subject')
    if not is_admin:
        plans = plans.filter(teacher=teacher)
    if selected_subject:
        plans = plans.filter(subject_id=selected_subject)
    if selected_class:
        plans = plans.filter(class_section=selected_class)
    if selected_status:
        plans = plans.filter(status=selected_status)

    teacher_subjects = teacher.subjects.all() if not is_admin else Subject.objects.all()

    return render(request, 'core/dashboard/lesson_plans.html', {
        'plans': plans,
        'teacher_subjects': teacher_subjects,
        'selected_subject': selected_subject,
        'selected_class': selected_class,
        'selected_status': selected_status,
        'TEACHING_METHOD_CHOICES': LessonPlan.TEACHING_METHOD_CHOICES,
        'LESSON_STATUS_CHOICES': LessonPlan.LESSON_STATUS_CHOICES,
    })


@teacher_required
def lesson_plan_detail(request, plan_id):
    """View / edit a single lesson plan"""
    teacher = request.teacher
    plan = get_object_or_404(LessonPlan, id=plan_id)

    if not plan.teacher == teacher and not teacher.is_admin:
        messages.error(request, 'Access denied. You do not own this lesson plan.')
        return redirect('core:lesson_plans')

    if request.method == 'POST':
        plan.title = request.POST.get('title', plan.title)
        plan.description = request.POST.get('description', plan.description)
        plan.class_section = request.POST.get('class_section', plan.class_section)
        plan.chapter_topic = request.POST.get('chapter_topic', plan.chapter_topic)
        plan.teaching_method = request.POST.get('teaching_method', plan.teaching_method)
        plan.learning_outcomes = request.POST.get('learning_outcomes', plan.learning_outcomes)
        plan.resources_used = request.POST.get('resources_used', plan.resources_used)
        plan.status = request.POST.get('status', plan.status)
        plan.remarks = request.POST.get('remarks', plan.remarks)

        planned_date_str = request.POST.get('planned_date', '')
        if planned_date_str:
            try:
                plan.planned_date = date.fromisoformat(planned_date_str)
            except ValueError:
                pass

        actual_date_str = request.POST.get('actual_date', '')
        if actual_date_str:
            try:
                plan.actual_date = date.fromisoformat(actual_date_str)
            except ValueError:
                pass

        duration = request.POST.get('duration_minutes', '')
        if duration:
            plan.duration_minutes = int(duration)

        subject_id = request.POST.get('subject', '')
        if subject_id:
            try:
                plan.subject = Subject.objects.get(id=subject_id)
            except Subject.DoesNotExist:
                pass

        plan.save()
        messages.success(request, 'Lesson plan updated successfully.')
        return redirect('core:lesson_plan_detail', plan_id=plan.id)

    return render(request, 'core/dashboard/lesson_plan_detail.html', {
        'plan': plan,
        'TEACHING_METHOD_CHOICES': LessonPlan.TEACHING_METHOD_CHOICES,
        'LESSON_STATUS_CHOICES': LessonPlan.LESSON_STATUS_CHOICES,
        'teacher_subjects': teacher.subjects.all(),
    })


@teacher_required
def lesson_plan_delete(request, plan_id):
    """Delete a lesson plan (POST only)"""
    teacher = request.teacher
    plan = get_object_or_404(LessonPlan, id=plan_id)

    if plan.teacher != teacher and not teacher.is_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)

    if request.method == 'POST':
        plan.delete()
        return JsonResponse({'success': True, 'message': 'Lesson plan deleted.'})
    return JsonResponse({'success': False, 'message': 'POST required'}, status=405)


@teacher_required
def syllabus_tracker_view(request):
    """Per-subject syllabus coverage progress with chapter breakdown"""
    teacher = request.teacher
    is_admin = teacher.is_admin

    subject_filter = request.GET.get('subject', '')
    section_filter = request.GET.get('class_section', '')

    # Get subjects teacher can access
    teacher_subjects = teacher.subjects.all()

    # Build progress data per subject per class_section
    coverage_data = {}
    subjects_qs = teacher_subjects if not is_admin else Subject.objects.all()

    for subj in subjects_qs:
        syllabus_items = SyllabusCoverage.objects.filter(subject=subj)
        if subject_filter:
            syllabus_items = syllabus_items.filter(subject_id=subject_filter)
        if section_filter:
            syllabus_items = syllabus_items.filter(class_section=section_filter)

        class_sections = syllabus_items.values_list('class_section', flat=True).distinct()
        for cs in class_sections:
            items = syllabus_items.filter(class_section=cs).select_related('subject')
            total = items.count()
            completed = items.filter(is_completed=True).count()
            pct = round((completed / total * 100)) if total > 0 else 0
            key = f"{subj.id}_{cs}"
            coverage_data[key] = {
                'subject': subj,
                'class_section': cs,
                'total': total,
                'completed': completed,
                'pending': total - completed,
                'pct': pct,
                'items': list(items),
            }

    # Group data for display
    grouped = {}
    for key, data in coverage_data.items():
        subj_name = data['subject'].name
        if subj_name not in grouped:
            grouped[subj_name] = []
        grouped[subj_name].append(data)

    available_subjects = list(subjects_qs.order_by('name'))
    available_sections = sorted({v['class_section'] for v in coverage_data.values()})

    return render(request, 'core/dashboard/syllabus_tracker.html', {
        'grouped': grouped,
        'coverage_data': coverage_data,
        'selected_subject': subject_filter,
        'selected_section': section_filter,
        'available_subjects': available_subjects,
        'available_sections': available_sections,
    })


@teacher_required
@require_http_methods(["POST"])
def syllabus_mark_completed(request):
    """Mark a syllabus chapter as completed"""
    teacher = request.teacher
    item_id = request.POST.get('item_id', '').strip()
    is_completed = request.POST.get('is_completed', '0') == '1'

    try:
        item = SyllabusCoverage.objects.select_related('subject').get(id=item_id)
        if not teacher.is_admin and not teacher.can_access_subject(item.subject):
            return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)

        item.is_completed = is_completed
        if is_completed:
            item.actual_hours += 1
        item.update_completion_pct()
        return JsonResponse({
            'success': True,
            'pct': item.completion_pct,
            'is_completed': item.is_completed,
        })
    except SyllabusCoverage.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Item not found'}, status=404)


# ═════════════════════════════════════════════════════════════════════════════
# T3 — TWO-WAY GRADING RUBRICS
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def rubric_list(request):
    """List teacher's rubrics with filter/subject"""
    teacher = request.teacher
    rubrics = GradingRubric.objects.filter(teacher=teacher).select_related('subject').prefetch_related('criteria')

    subject_filter = request.GET.get('subject', '')
    if subject_filter:
        rubrics = rubrics.filter(subject_id=subject_filter)

    teacher_subjects = teacher.subjects.all()

    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'create':
            name = request.POST.get('name', '').strip()
            subject_id = request.POST.get('subject', '').strip()
            description = request.POST.get('description', '').strip()
            is_active = request.POST.get('is_active') == 'on'

            if name:
                from django.db import transaction
                try:
                    subj = Subject.objects.get(id=subject_id) if subject_id else None
                    rubric = GradingRubric.objects.create(
                        name=name, teacher=teacher, subject=subj,
                        description=description, is_active=is_active,
                    )
                    messages.success(request, f'Rubric "{name}" created. Add criteria from the rubric detail page.')
                    return redirect('core:rubric_detail', rubric_id=rubric.id)
                except Subject.DoesNotExist:
                    messages.error(request, 'Invalid subject selected.')
            else:
                messages.error(request, 'Rubric name is required.')
        return redirect('core:rubric_list')

    return render(request, 'core/dashboard/rubric_list.html', {
        'rubrics': rubrics,
        'teacher_subjects': teacher_subjects,
        'selected_subject': subject_filter,
    })


@teacher_required
def rubric_create(request):
    """Create a new grading rubric (POST with criteria JSON from client)"""
    teacher = request.teacher
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)

    name = request.POST.get('name', '').strip()
    subject_id = request.POST.get('subject', '').strip()
    description = request.POST.get('description', '').strip()
    is_active = request.POST.get('is_active') == 'on'

    if not name:
        return JsonResponse({'success': False, 'message': 'Rubric name required'}, status=400)

    rubric = GradingRubric(
        name=name, teacher=teacher,
        description=description, is_active=is_active,
    )
    if subject_id:
        try:
            rubric.subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid subject'}, status=400)

    rubric.save()

    # Add criteria from JSON
    criteria_json = request.POST.get('criteria_json', '[]')
    try:
        criteria_list = json.loads(criteria_json)
    except (json.JSONDecodeError, ValueError):
        criteria_list = []

    for idx, c in enumerate(criteria_list, start=1):
        RubricCriterion.objects.create(
            rubric=rubric,
            name=c.get('name', f'Criterion {idx}'),
            max_score=float(c.get('max_score', 10)),
            weight_percent=float(c.get('weight_percent', 33.3)),
            order=idx,
        )

    return JsonResponse({'success': True, 'rubric_id': rubric.id, 'message': f'Rubric "{name}" created.'})


@teacher_required
def rubric_detail(request, rubric_id):
    """View / edit a rubric's criteria"""
    teacher = request.teacher
    rubric = get_object_or_404(GradingRubric, id=rubric_id)

    if rubric.teacher != teacher and not teacher.is_admin:
        messages.error(request, 'Access denied.')
        return redirect('core:rubric_list')

    if request.method == 'POST':
        # Update rubric fields
        rubric.name = request.POST.get('name', rubric.name)
        rubric.description = request.POST.get('description', rubric.description)
        rubric.is_active = request.POST.get('is_active') == 'on'

        subject_id = request.POST.get('subject', '')
        if subject_id:
            try:
                rubric.subject = Subject.objects.get(id=subject_id)
            except Subject.DoesNotExist:
                pass
        rubric.save()

        # Update criteria
        criteria_ids = request.POST.getlist('criteria_id')
        for cid in criteria_ids:
            try:
                crit = rubric.criteria.get(id=int(cid))
                crit.name = request.POST.get(f'name_{cid}', crit.name)
                crit.max_score = float(request.POST.get(f'max_score_{cid}', crit.max_score))
                crit.weight_percent = float(request.POST.get(f'weight_{cid}', crit.weight_percent))
                crit.order = int(request.POST.get(f'order_{cid}', crit.order))
                crit.save()
            except (RubricCriterion.DoesNotExist, ValueError):
                continue

        messages.success(request, 'Rubric updated successfully.')
        return redirect('core:rubric_detail', rubric_id=rubric.id)

    context = {
        'rubric': rubric,
        'criteria': rubric.criteria.all(),
        'teacher_subjects': teacher.subjects.all(),
    }
    return render(request, 'core/dashboard/rubric_detail.html', context)


@teacher_required
@require_http_methods(["POST"])
def rubric_assign_to_submission(request, rubric_id):
    """Attach a rubric to an AssignmentSubmission"""
    teacher = request.teacher
    rubric = get_object_or_404(GradingRubric, id=rubric_id)

    if rubric.teacher != teacher and not teacher.is_admin:
        return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)

    sub_id = request.POST.get('submission_id', '').strip()
    if not sub_id:
        return JsonResponse({'success': False, 'message': 'submission_id required'}, status=400)

    try:
        submission = AssignmentSubmission.objects.select_related('assignment', 'student').get(id=int(sub_id))
        # Ensure teacher owns the rubric AND assignment
        if not teacher.is_admin and submission.assignment.teacher != teacher:
            return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)

        # Copy criteria from rubric into RubricScoreEntry (default 0)
        created = 0
        for crit in rubric.criteria.all():
            entry, _ = RubricScoreEntry.objects.get_or_create(
                submission=submission, criterion=crit,
                defaults={'score_obtained': 0},
            )
            created += 1

        return JsonResponse({'success': True, 'message': f'Rubric assigned to submission with {created} criteria.'})
    except AssignmentSubmission.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Submission not found'}, status=404)


@teacher_required
@require_http_methods(["POST"])
def rubric_save_scores(request, submission_id):
    """Save per-criterion scores for a submission; auto-compute total"""
    teacher = request.teacher
    try:
        submission = AssignmentSubmission.objects.select_related('assignment').get(id=int(submission_id))
        if not teacher.is_admin and submission.assignment.teacher != teacher:
            return JsonResponse({'success': False, 'message': 'Access denied'}, status=403)
    except AssignmentSubmission.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Submission not found'}, status=404)

    total_score = 0.0
    for key, value in request.POST.items():
        if key.startswith('score_'):
            entry_id = key.replace('score_', '').strip()
            try:
                entry = RubricScoreEntry.objects.get(id=int(entry_id), submission=submission)
                entry.score_obtained = float(value)
                entry.save()
                total_score += float(value)
            except (RubricScoreEntry.DoesNotExist, ValueError):
                continue

    # Update submission total marks from rubric total
    if total_score > 0:
        submission.marks_obtained = total_score
        submission.status = 'graded'
        from django.utils import timezone
        submission.graded_at = timezone.now()
        submission.graded_by = teacher
        submission.save()

    return JsonResponse({'success': True, 'total_score': total_score})


@teacher_required
def rubric_templates(request):
    """Browse global rubric templates and import one"""
    teacher = request.teacher
    templates = RubricTemplate.objects.filter(is_global=True)
    my_templates = RubricTemplate.objects.filter(is_global=False)

    return render(request, 'core/dashboard/rubric_templates.html', {
        'global_templates': templates,
        'my_templates': my_templates,
    })


# ═════════════════════════════════════════════════════════════════════════════
# T4 — ONLINE EXAM / QUIZ MODULE
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def online_exam_list(request):
    """Teacher: list their online exams"""
    teacher = request.teacher
    now = timezone.now()
    exams = OnlineExam.objects.filter(teacher=teacher).select_related('subject').order_by('-start_at')

    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        exams = [e for e in exams if e.is_published and not e.has_ended]
    elif status_filter == 'ended':
        exams = [e for e in exams if e.has_ended]
    elif status_filter == 'draft':
        exams = [e for e in exams if not e.is_published]

    for exam in exams:
        student_count = Student.objects.filter(student_class=exam.target_class)
        if exam.target_section:
            student_count = student_count.filter(section=exam.target_section)
        exam.student_count = student_count.count()
        exam.attempt_count = exam.attempts.count()

    return render(request, 'core/dashboard/online_exam_list.html', {
        'exams': exams,
        'selected_status': status_filter,
        'teacher_subjects': teacher.subjects.all(),
    })


@teacher_required
def online_exam_create(request):
    """Create a new online exam with config and one question builder"""
    teacher = request.teacher
    teacher_subjects = teacher.subjects.all()

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        subject_id = request.POST.get('subject', '').strip()
        target_class = request.POST.get('target_class', '').strip()
        target_section = request.POST.get('target_section', '').strip()
        total_marks = request.POST.get('total_marks', '100').strip()
        passing_marks = request.POST.get('passing_marks', '40').strip()
        duration = request.POST.get('duration_minutes', '60').strip()
        start_at_str = request.POST.get('start_at', '').strip()
        end_at_str = request.POST.get('end_at', '').strip()
        is_published = request.POST.get('is_published') == 'on'
        show_result = request.POST.get('show_result_immediately') == 'on'
        shuffle = request.POST.get('shuffle_questions') == 'on'
        max_attempts = request.POST.get('max_attempts', '1').strip()

        if not (title and subject_id and target_class and start_at_str and end_at_str):
            messages.error(request, 'Title, subject, class, start and end date/time are required.')
            return redirect('core:online_exam_create')

        try:
            subj = Subject.objects.get(id=subject_id)
            if not teacher.is_admin and not teacher.can_access_subject(subj):
                messages.error(request, 'You can only create exams for your assigned subjects.')
                return redirect('core:online_exam_create')

            from datetime import datetime
            start_at = datetime.fromisoformat(start_at_str)
            end_at = datetime.fromisoformat(end_at_str)

            exam = OnlineExam.objects.create(
                title=title, subject=subj, teacher=teacher,
                target_class=target_class, target_section=target_section,
                total_marks=int(total_marks) if total_marks else 100,
                passing_marks=int(passing_marks) if passing_marks else 40,
                duration_minutes=int(duration) if duration else 60,
                start_at=start_at, end_at=end_at,
                is_published=is_published,
                show_result_immediately=show_result,
                shuffle_questions=shuffle,
                max_attempts=int(max_attempts) if max_attempts else 1,
            )

            # Add initial questions if provided
            q_texts = request.POST.getlist('question_text[]')
            q_types = request.POST.getlist('question_type[]')
            q_options = request.POST.getlist('question_option[]')
            q_corrects = request.POST.getlist('question_correct[]')
            q_marks = request.POST.getlist('question_marks[]')
            q_order_list = request.POST.getlist('question_order[]')

            for i in range(len(q_texts)):
                q_text = (q_texts[i] or '').strip()
                q_type = q_types[i] if i < len(q_types) else 'MCQ'
                q_mark = int(q_marks[i]) if i < len(q_marks) and q_marks[i] else 1
                q_order = int(q_order_list[i]) if i < len(q_order_list) and q_order_list[i] else (i + 1)
                correct = q_corrects[i] if i < len(q_corrects) else ''
                raw_opts = q_options[i] if i < len(q_options) else ''

                opts = []
                if q_type == 'MCQ' and raw_opts:
                    option_labels = ['A', 'B', 'C', 'D']
                    raw_split = [o.strip() for o in raw_opts.split(',')]
                    for j, opt in enumerate(raw_split):
                        if j < 4:
                            opts.append({'label': option_labels[j], 'text': opt})

                Question.objects.create(
                    online_exam=exam, question_text=q_text, question_type=q_type,
                    options=opts, correct_answer=correct, marks=q_mark, order=q_order,
                )

            messages.success(request, f'Exam "{title}" created with {len(q_texts)} questions.')
            return redirect('core:online_exam_detail', exam_id=exam.id)
        except ValueError as e:
            messages.error(request, f'Invalid date format. Use YYYY-MM-DDTHH:MM. Error: {str(e)}')
        except Subject.DoesNotExist:
            messages.error(request, 'Invalid subject selected.')

    available_classes = []
    school_students = Student.objects.filter(level='school')
    college_students = Student.objects.filter(level='college')
    bachelor_students = Student.objects.filter(level='bachelor')
    if school_students.exists():
        available_classes += sorted(school_students.values_list('student_class', flat=True).distinct())
    if college_students.exists():
        available_classes += sorted(college_students.values_list('student_class', flat=True).distinct())
    if bachelor_students.exists():
        available_classes += sorted(bachelor_students.values_list('student_class', flat=True).distinct())

    return render(request, 'core/dashboard/online_exam_create.html', {
        'teacher_subjects': teacher_subjects,
        'available_classes': available_classes,
        'QUESTION_TYPE_CHOICES': Question.QUESTION_TYPE_CHOICES,
    })


@teacher_required
def online_exam_detail(request, exam_id):
    """View exam + questions, allow editing questions"""
    teacher = request.teacher
    exam = get_object_or_404(OnlineExam.objects.select_related('subject', 'teacher'), id=exam_id)

    if exam.teacher != teacher and not teacher.is_admin:
        messages.error(request, 'Access denied.')
        return redirect('core:online_exam_list')

    questions = exam.questions.prefetch_related(None).select_related(None).all().order_by('order')

    if request.method == 'POST':
        action = request.POST.get('action', 'edit')
        if action == 'add_question':
            q_text = request.POST.get('question_text', '').strip()
            q_type = request.POST.get('question_type', 'MCQ')
            q_mark = int(request.POST.get('question_marks', 1) or 1)
            correct = request.POST.get('question_correct', '')

            opts = []
            if q_type == 'MCQ':
                for label in ['A', 'B', 'C', 'D']:
                    opt_val = request.POST.get(f'opt_{label}', '')
                    if opt_val:
                        opts.append({'label': label, 'text': opt_val})

            next_order = (questions.aggregate(max_o=Count('order'))['max_o'] or 0) + 1
            Question.objects.create(
                online_exam=exam, question_text=q_text, question_type=q_type,
                options=opts, correct_answer=correct, marks=q_mark, order=next_order,
            )
            exam.total_marks = sum(q.marks for q in exam.questions.all()) or exam.total_marks
            exam.save(update_fields=['total_marks'])
            messages.success(request, 'Question added.')
        elif action == 'edit_exam':
            exam.title = request.POST.get('title', exam.title)
            exam.total_marks = int(request.POST.get('total_marks', exam.total_marks) or exam.total_marks)
            exam.passing_marks = int(request.POST.get('passing_marks', exam.passing_marks) or exam.passing_marks)
            exam.duration_minutes = int(request.POST.get('duration_minutes', exam.duration_minutes) or exam.duration_minutes)
            exam.shuffle_questions = request.POST.get('shuffle_questions') == 'on'
            exam.show_result_immediately = request.POST.get('show_result_immediately') == 'on'
            exam.max_attempts = int(request.POST.get('max_attempts', exam.max_attempts) or 1)
            exam.is_published = request.POST.get('is_published') == 'on'
            start_str = request.POST.get('start_at', '')
            end_str = request.POST.get('end_at', '')
            if start_str:
                try:
                    exam.start_at = datetime.fromisoformat(start_str)
                except ValueError:
                    pass
            if end_str:
                try:
                    exam.end_at = datetime.fromisoformat(end_str)
                except ValueError:
                    pass
            exam.save()
            messages.success(request, 'Exam settings updated.')
        elif action == 'delete_question':
            q_id = request.POST.get('question_id', '').strip()
            try:
                q = Question.objects.get(id=int(q_id), online_exam=exam)
                q.delete()
                exam.total_marks = sum(qs.marks for qs in exam.questions.all()) or exam.total_marks
                exam.save(update_fields=['total_marks'])
                messages.success(request, 'Question deleted.')
            except Question.DoesNotExist:
                messages.error(request, 'Question not found.')

        return redirect('core:online_exam_detail', exam_id=exam.id)

    questions = exam.questions.order_by('order')
    return render(request, 'core/dashboard/online_exam_detail.html', {
        'exam': exam,
        'questions': questions,
        'QUESTION_TYPE_CHOICES': Question.QUESTION_TYPE_CHOICES,
    })


@teacher_required
def online_exam_submissions(request, exam_id):
    """View who attempted, their scores, manually grade short-answer"""
    teacher = request.teacher
    exam = get_object_or_404(OnlineExam.objects.select_related('subject', 'teacher'), id=exam_id)

    if exam.teacher != teacher and not teacher.is_admin:
        messages.error(request, 'Access denied.')
        return redirect('core:online_exam_list')

    attempts = exam.attempts.select_related('student').prefetch_related('answers').order_by('-started_at')
    students = Student.objects.filter(student_class=exam.target_class).order_by('name')

    for attempt in attempts:
        attempt.manual_answers = attempt.answers.filter(
            question__question_type__in=['SHORT_ANSWER', 'LONG_ANSWER'],
        ).select_related('question')

    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'grade':
            ans_id = request.POST.get('answer_id', '').strip()
            try:
                answer = ExamAnswer.objects.select_related('attempt', 'question').get(id=int(ans_id))
                if not teacher.is_admin and answer.attempt.online_exam.teacher != teacher:
                    messages.error(request, 'Access denied.')
                else:
                    marks = float(request.POST.get('marks_awarded', 0) or 0)
                    answer.marks_awarded = marks
                    answer.save()
                    # Recompute attempt score
                    attempt = answer.attempt
                    total = sum(a.marks_awarded or 0 for a in attempt.answers.all())
                    attempt.score_obtained = total
                    attempt.status = 'SUBMITTED'
                    attempt.save(update_fields=['score_obtained', 'status'])
                    messages.success(request, f'Marks ({marks}) awarded for answer.')
            except (ExamAnswer.DoesNotExist, ValueError):
                messages.error(request, 'Invalid request.')
        return redirect('core:online_exam_submissions', exam_id=exam.id)

    return render(request, 'core/dashboard/online_exam_submissions.html', {
        'exam': exam,
        'attempts': attempts,
        'students': students,
    })


@login_required
def student_exam_list(request):
    """Student: see their available, active, and completed exams"""
    teacher = get_teacher_profile(request.user)
    is_student = not teacher and not _is_admin_user(request.user)
    student = _get_current_student(request)
    if not student:
        return render(request, 'core/dashboard/student_exam_list.html', {
            'exams': [], 'upcoming': [], 'active': [], 'completed': [], 'is_student': False,
        })

    now = timezone.now()
    exams_qs = OnlineExam.objects.select_related('subject').filter(
        is_published=True,
        target_class=student.student_class,
    ).filter(
        Q(target_section='') | Q(target_section=student.section)
    ).order_by('-start_at')

    upcoming = [e for e in exams_qs if e.start_at > now]
    active = []
    completed = []
    for exam in exams_qs:
        attempt = exam.attempts.filter(student=student).first()
        if not attempt:
            if not (e.start_at > now):
                continue
        elif attempt.status == 'IN_PROGRESS':
            active.append(exam)
        else:
            completed.append({
                'exam': exam,
                'attempt': attempt,
            })

    all_active_exams = list(active) + [c['exam'] for c in completed]

    return render(request, 'core/dashboard/student_exam_list.html', {
        'exams': all_active_exams, 'upcoming': upcoming,
        'active_exams': active, 'completed_exams': completed,
        'student': student, 'is_student': True,
    })


@login_required
def student_exam_start(request, exam_id):
    """Student: start attempt — shuffle questions, render timer"""
    student = _get_current_student(request)
    exam = get_object_or_404(OnlineExam.objects.select_related('subject', 'teacher'), id=exam_id)

    if not student:
        messages.error(request, 'Student profile not found.')
        return redirect('core:student_exam_list')

    # Check student is targeted for this exam
    if exam.target_class != student.student_class or (
        exam.target_section and exam.target_section != student.section
    ):
        messages.error(request, 'This exam is not for your class/section.')
        return redirect('core:student_exam_list')

    # Check exam window
    if not exam.is_published:
        messages.error(request, 'This exam is not yet published.')
        return redirect('core:student_exam_list')
    if not exam.is_available:
        messages.error(request, 'Exam has not started yet or has ended.')
        return redirect('core:student_exam_list')

    # Check attempts
    existing = exam.attempts.filter(student=student)
    if existing.count() >= exam.max_attempts:
        messages.error(request, 'Maximum attempts reached.')
        return redirect('core:student_exam_list')

    # Get / create attempt
    attempt = existing.first()
    if not attempt:
        attempt = ExamAttempt.objects.create(student=student, online_exam=exam, status='IN_PROGRESS')
    elif attempt.status in ['SUBMITTED', 'TIMED_OUT']:
        if existing.count() >= exam.max_attempts:
            messages.error(request, 'You have used all attempts.')
            return redirect('core:student_exam_list')
        attempt = ExamAttempt.objects.create(student=student, online_exam=exam, attempt_number=existing.count() + 1, status='IN_PROGRESS')

    # Fetch questions
    questions = list(exam.questions.order_by('order'))
    if exam.shuffle_questions:
        import random
        random.shuffle(questions)

    # Get student's answers so far
    answers = {a.question_id: a for a in attempt.answers.select_related('question').all()}

    time_left = None
    start_ts = attempt.started_at.timestamp() if attempt.started_at else timezone.now().timestamp()
    deadline = start_ts + exam.duration_minutes * 60
    time_left = max(0, int(deadline - timezone.now().timestamp()))

    return render(request, 'core/dashboard/student_exam_start.html', {
        'exam': exam, 'attempt': attempt, 'questions': questions,
        'answers': answers, 'time_left': time_left,
    })


@teacher_required
def online_exam_submissions_redirect(request, exam_id):
    """Alias for online_exam_submissions — used as route for /online-exams/<id>/submissions/"""
    return online_exam_submissions(request, exam_id=exam_id)


@login_required
def student_exam_submit(request, exam_id):
    """Student: submit answers — auto-grade MCQ/TF, store text answers"""
    student = _get_current_student(request)
    exam = get_object_or_404(OnlineExam, id=exam_id)

    if not student:
        return JsonResponse({'success': False, 'message': 'Not found'}, status=404)

    attempt = exam.attempts.filter(student=student, status='IN_PROGRESS').first()
    if not attempt:
        return JsonResponse({'success': False, 'message': 'No active attempt found'}, status=403)

    questions = exam.questions.order_by('order')
    total_score = 0.0

    for q in questions:
        if q.question_type in ['MCQ', 'TRUE_FALSE']:
            selected = request.POST.get(f'q_{q.id}', '')
            ans, _ = ExamAnswer.objects.get_or_create(
                attempt=attempt, question=q,
                defaults={'selected_option': selected},
            )
            if not ans.selected_option:
                ans.selected_option = selected
                ans.save()

            # Auto-grade
            correct = (q.correct_answer or '').strip().upper()
            is_correct = (selected or '').strip().upper() == correct
            awarded = q.marks if is_correct else 0
            ans.marks_awarded = awarded
            ans.save()
            total_score += awarded
        else:
            text = request.POST.get(f'q_{q.id}', '') or ''
            ans, _ = ExamAnswer.objects.get_or_create(
                attempt=attempt, question=q,
                defaults={'answer_text': text},
            )
            if not ans.answer_text:
                ans.answer_text = text
                ans.save()
            # Manual mark default = 0; teacher will grade later

    attempt.score_obtained = total_score
    attempt.status = 'SUBMITTED'
    attempt.submitted_at = timezone.now()
    attempt.save(update_fields=['score_obtained', 'status', 'submitted_at'])

    if exam.show_result_immediately:
        return JsonResponse({'success': True, 'redirect': reverse('core:student_exam_result', args=[exam_id, attempt.id])})
    return JsonResponse({'success': True, 'redirect': reverse('core:student_exam_list')})


@login_required
def student_exam_result(request, exam_id):
    """Student: view score and per-question review"""
    student = _get_current_student(request)
    exam = get_object_or_404(OnlineExam.objects.select_related('subject', 'teacher'), id=exam_id)

    if not student:
        messages.error(request, 'Student profile not found.')
        return redirect('core:student_exam_list')

    attempt = exam.attempts.filter(student=student).order_by('-started_at').first()
    if not attempt:
        messages.error(request, 'No attempt found.')
        return redirect('core:student_exam_list')

    answers = attempt.answers.select_related('question').order_by('question__order')

    result_data = []
    total_marks_attempted = 0
    total_marks_awarded = 0
    for ans in answers:
        q = ans.question
        correct = (q.correct_answer or '').strip().upper()
        selected = (ans.selected_option or '').strip().upper()
        is_correct = selected == correct
        if q.question_type in ['MCQ', 'TRUE_FALSE']:
            total_marks_attempted += q.marks
            total_marks_awarded += (ans.marks_awarded or 0)

        result_data.append({
            'question': q,
            'answer': ans,
            'is_correct': is_correct,
            'correct_answer': correct,
        })

    percentage = round((total_marks_awarded / total_marks_attempted * 100), 1) if total_marks_attempted > 0 else 0
    is_pass = percentage >= (exam.passing_marks / exam.total_marks * 100)

    return render(request, 'core/dashboard/student_exam_result.html', {
        'exam': exam,
        'attempt': attempt,
        'result_data': result_data,
        'total_marks_awarded': total_marks_awarded,
        'total_marks_attempted': total_marks_attempted,
        'percentage': percentage,
        'is_pass': is_pass,
    })


# ═════════════════════════════════════════════════════════════════════════════
# T5 — SEATING PLAN & HALL TICKET GENERATOR
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def seating_plan_list(request):
    """List all seating plans created by the teacher"""
    teacher = request.teacher
    plans = ExamSeatingPlan.objects.filter(teacher=teacher).select_related('exam', 'teacher').prefetch_related('allocations').order_by('-created_at')

    for plan in plans:
        plan.student_count = plan.allocations.count()

    return render(request, 'core/dashboard/seating_plan_list.html', {
        'plans': plans,
    })


@teacher_required
def seating_plan_create(request):
    """Create a seating plan — select exam + add rooms; auto-generate allocations"""
    teacher = request.teacher

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)

    exam_id = request.POST.get('exam_id', '').strip()
    room_name = request.POST.get('room_name', '').strip()
    room_capacity = int(request.POST.get('room_capacity', 30) or 30)
    arrangement = request.POST.get('arrangement_type', 'row_wise')

    if not exam_id or not room_name:
        return JsonResponse({'success': False, 'message': 'Exam and room name are required'}, status=400)

    try:
        exam = Exam.objects.get(id=exam_id)
    except Exam.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Exam not found'}, status=404)

    plan, created = ExamSeatingPlan.objects.get_or_create(
        exam=exam, teacher=teacher, room_name=room_name,
        defaults={'room_capacity': room_capacity, 'arrangement_type': arrangement},
    )
    if not created:
        plan.room_capacity = room_capacity
        plan.arrangement_type = arrangement
        plan.save()

    # Auto-generate seat allocations if none exist
    if plan.allocations.count() == 0:
        students = Student.objects.filter(student_class=exam.target_class).order_by('roll_number', 'name')
        if exam.target_section:
            students = students.filter(section=exam.target_section)

        student_list = list(students[:room_capacity])

        if arrangement == 'row_wise':
            cols = min(5, max(1, int(room_capacity ** 0.5)))
        elif arrangement == 'column_wise':
            cols = max(1, int(room_capacity ** 0.5))
        else:  # checkerboard
            cols = min(5, max(1, int(room_capacity ** 0.5)))

        row_num = 1
        col_num = 1
        letter = 'A'
        for idx, stu in enumerate(student_list):
            seat_label = f"{letter}-{col_num:02d}"
            SeatAllocation.objects.create(
                seating_plan=plan,
                student=stu,
                seat_number=seat_label,
                row=row_num,
                column=col_num,
            )
            col_num += 1
            if col_num > (cols if arrangement != 'column_wise' else 5):
                col_num = 1
                letter = chr(ord(letter) + 1)
                row_num += 1

    return JsonResponse({'success': True, 'plan_id': plan.id, 'message': f'Seating plan "{room_name}" created with {plan.allocations.count()} allocations.'})


@teacher_required
def seating_plan_view(request, plan_id):
    """View and edit the seat grid for a seating plan"""
    teacher = request.teacher
    plan = get_object_or_404(ExamSeatingPlan.objects.select_related('exam', 'teacher'), id=plan_id)

    if plan.teacher != teacher:
        messages.error(request, 'Access denied.')
        return redirect('core:seating_plan_list')

    if request.method == 'POST':
        # Allow editing seat_number/row/column
        for alloc in plan.allocations.all():
            f_seat = f'seat_{alloc.id}'
            f_row = f'row_{alloc.id}'
            f_col = f'col_{alloc.id}'
            if f_seat in request.POST:
                alloc.seat_number = request.POST.get(f_seat, alloc.seat_number)
            if f_row in request.POST:
                try:
                    alloc.row = int(request.POST.get(f_row))
                except ValueError:
                    pass
            if f_col in request.POST:
                try:
                    alloc.column = int(request.POST.get(f_col))
                except ValueError:
                    pass
            alloc.save()
        messages.success(request, 'Seat allocations updated.')
        return redirect('core:seating_plan_view', plan_id=plan.id)

    allocations = plan.allocations.select_related('student').all().order_by('row', 'column')

    return render(request, 'core/dashboard/seating_plan_view.html', {
        'plan': plan,
        'allocations': allocations,
    })


@teacher_required
def seating_plan_hall_ticket(request, plan_id, student_id):
    """Generate individual hall ticket with QR for a student"""
    teacher = request.teacher
    plan = get_object_or_404(ExamSeatingPlan.objects.select_related('exam'), id=plan_id)
    student = get_object_or_404(Student, id=student_id)

    if not teacher.is_admin and plan.teacher != teacher:
        messages.error(request, 'Access denied.')
        return redirect('core:seating_plan_list')

    seat = SeatAllocation.objects.filter(seating_plan=plan, student=student).first()
    if not seat:
        seat = None

    # QR code
    qr_data = f"{student.name}|{student.roll_number or ''}|{plan.exam.title}|{seat.seat_number if seat else ''}"
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=8, border=3)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    return render(request, 'core/dashboard/seating_plan_hall_ticket.html', {
        'plan': plan, 'student': student, 'seat': seat, 'qr_code': qr_b64,
    })


@teacher_required
def seating_plan_bulk_hall_tickets(request, plan_id):
    """Generate A4 print layout with all students for a seating plan"""
    teacher = request.teacher
    plan = get_object_or_404(ExamSeatingPlan.objects.select_related('exam'), id=plan_id)

    if not teacher.is_admin and plan.teacher != teacher:
        messages.error(request, 'Access denied.')
        return redirect('core:seating_plan_list')

    allocations = plan.allocations.select_related('student').all().order_by('row', 'column')
    return render(request, 'core/dashboard/seating_plan_bulk_hall_tickets.html', {
        'plan': plan,
        'allocations': allocations,
    })


# ═════════════════════════════════════════════════════════════════════════════
# T6 — TEACHER LEAVE MANAGEMENT
# ═════════════════════════════════════════════════════════════════════════════

@teacher_required
def teacher_leave_list(request):
    """Teacher: view their own leave history"""
    teacher = request.teacher
    leaves = TeacherLeave.objects.filter(teacher=teacher).order_by('-start_date')
    status_filter = request.GET.get('status', '')
    if status_filter:
        leaves = leaves.filter(status=status_filter)
    return render(request, 'core/dashboard/teacher_leave_list.html', {
        'leaves': leaves,
        'selected_status': status_filter,
        'LEAVE_TYPE_CHOICES': TeacherLeave.LEAVE_TYPE_CHOICES,
        'LEAVE_STATUS_CHOICES': TeacherLeave.LEAVE_STATUS_CHOICES,
    })


@teacher_required
def teacher_leave_create(request):
    """Teacher: create a new leave request"""
    teacher = request.teacher
    other_teachers = Teacher.objects.filter(is_active=True).exclude(id=teacher.id).order_by('first_name')

    if request.method == 'POST':
        leave_type = request.POST.get('leave_type', 'casual').strip()
        start_date = request.POST.get('start_date', '').strip()
        end_date = request.POST.get('end_date', '').strip()
        reason = request.POST.get('reason', '').strip()
        substitute_id = request.POST.get('substitute_teacher', '').strip()

        if not (start_date and end_date and reason):
            messages.error(request, 'All fields except substitute teacher are required.')
            return redirect('core:teacher_leave_create')

        try:
            from datetime import date
            sd = date.fromisoformat(start_date)
            ed = date.fromisoformat(end_date)
            if ed < sd:
                messages.error(request, 'End date cannot be before start date.')
                return redirect('core:teacher_leave_create')

            sub = None
            if substitute_id:
                sub = Teacher.objects.get(id=substitute_id)

            leave = TeacherLeave.objects.create(
                teacher=teacher, leave_type=leave_type,
                start_date=sd, end_date=ed, reason=reason,
                substitute_teacher=sub, status='PENDING',
            )
            return JsonResponse({'success': True, 'message': 'Leave request submitted.', 'leave_id': leave.id})
        except (ValueError, Teacher.DoesNotExist) as e:
            messages.error(request, f'Invalid input: {str(e)}')

        return redirect('core:teacher_leave_create')

    return render(request, 'core/dashboard/teacher_leave_create.html', {
        'other_teachers': other_teachers,
        'LEAVE_TYPE_CHOICES': TeacherLeave.LEAVE_TYPE_CHOICES,
    })


@teacher_required
def teacher_leave_detail(request, leave_id):
    """Teacher: view a leave request detail; cancel if pending"""
    teacher = request.teacher
    leave = get_object_or_404(TeacherLeave, id=leave_id)

    if leave.teacher != teacher:
        messages.error(request, 'Access denied.')
        return redirect('core:teacher_leave_list')

    if request.method == 'POST' and leave.status == 'PENDING':
        action = request.POST.get('action', '')
        if action == 'cancel':
            leave.status = 'CANCELLED'
            leave.save(update_fields=['status'])
            messages.success(request, 'Leave request cancelled.')
            return redirect('core:teacher_leave_detail', leave_id=leave.id)

    return render(request, 'core/dashboard/teacher_leave_detail.html', {
        'leave': leave,
        'LEAVE_TYPE_CHOICES': TeacherLeave.LEAVE_TYPE_CHOICES,
        'LEAVE_STATUS_CHOICES': TeacherLeave.LEAVE_STATUS_CHOICES,
    })


@_admin_required
def teacher_leave_approve(request, leave_id):
    """Admin: approve / reject a leave request"""
    leave = get_object_or_404(TeacherLeave, id=leave_id)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)

    action = request.POST.get('action', '').strip()
    approval_reason = request.POST.get('approval_reason', '').strip()
    substitute_id = request.POST.get('substitute_teacher', '').strip()

    if action == 'approve':
        leave.status = 'APPROVED'
        leave.approval_reason = approval_reason
        leave.approved_by = request.user
        leave.approved_at = timezone.now()

        sub = None
        if substitute_id:
            try:
                sub = Teacher.objects.get(id=substitute_id)
            except Teacher.DoesNotExist:
                sub = None
        leave.substitute_teacher = sub
        leave.save()

        NotificationService.log_activity(
            request.user, 'other',
            f'Approved leave request {leave.id} for {leave.teacher.get_full_name()}'
        )
        return JsonResponse({'success': True, 'message': 'Leave request approved.'})

    elif action == 'reject':
        leave.status = 'REJECTED'
        leave.approval_reason = approval_reason
        leave.approved_by = request.user
        leave.approved_at = timezone.now()
        leave.save()
        NotificationService.log_activity(
            request.user, 'other',
            f'Rejected leave request {leave.id} for {leave.teacher.get_full_name()}'
        )
        return JsonResponse({'success': True, 'message': 'Leave request rejected.'})

    return JsonResponse({'success': False, 'message': 'Invalid action'}, status=400)


@_admin_required
def admin_leave_dashboard(request):
    """Admin: all pending leave requests; quick approve/reject actions"""
    leaves = TeacherLeave.objects.all().order_by('-created_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        leaves = leaves.filter(status=status_filter)

    all_teachers = Teacher.objects.filter(is_active=True).order_by('first_name')

    if request.method == 'POST':
        leave_id = request.POST.get('leave_id', '').strip()
        action = request.POST.get('action', '')
        reason = request.POST.get('reason', '').strip()
        sub_id = request.POST.get('substitute_id', '').strip()

        try:
            leave = TeacherLeave.objects.get(id=int(leave_id))

            if action == 'approve':
                leave.status = 'APPROVED'
                leave.approval_reason = reason
                leave.approved_by = request.user
                leave.approved_at = timezone.now()
                if sub_id:
                    try:
                        leave.substitute_teacher = Teacher.objects.get(id=sub_id)
                    except Teacher.DoesNotExist:
                        pass
                leave.save()
                messages.success(request, f'Leave for {leave.teacher.get_full_name()} approved.')
            elif action == 'reject':
                leave.status = 'REJECTED'
                leave.approval_reason = reason
                leave.approved_by = request.user
                leave.approved_at = timezone.now()
                leave.save()
                messages.success(request, f'Leave for {leave.teacher.get_full_name()} rejected.')
        except TeacherLeave.DoesNotExist:
            messages.error(request, 'Leave request not found.')
        return redirect('core:admin_leave_dashboard')

    return render(request, 'core/dashboard/admin_leave_dashboard.html', {
        'leaves': leaves,
        'all_teachers': all_teachers,
        'selected_status': status_filter,
        'LEAVE_STATUS_CHOICES': TeacherLeave.LEAVE_STATUS_CHOICES,
    })


@login_required
def leave_calendar(request):
    """Calendar view: show which teachers are on leave each day in the month"""
    from datetime import timedelta
    from django.db.models.functions import TruncMonth

    now = timezone.now()
    year = request.GET.get('year', str(now.year))
    month = request.GET.get('month', str(now.month))
    try:
        year = int(year)
        month = int(month)
    except ValueError:
        year = now.year
        month = now.month

    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)

    approved_leaves = TeacherLeave.objects.filter(
        status='APPROVED',
        start_date__lte=last_day,
        end_date__gte=first_day,
    ).select_related('teacher', 'substitute_teacher').order_by('start_date')

    # Build day → list of leaves map
    day_leaves = {}
    d = first_day
    while d <= last_day:
        day_leaves[d.isoformat()] = []
        d += timedelta(days=1)

    for leave in approved_leaves:
        d = leave.start_date
        while d <= leave.end_date and d <= last_day:
            key = d.isoformat()
            if key in day_leaves:
                day_leaves[key].append(leave)
            d += timedelta(days=1)

    # Build calendar grid (weeks → days)
    import calendar as cal_module
    cal = cal_module.Calendar(firstweekday=6)  # Sunday = 0
    weeks = []
    for week in cal.monthdatescalendar(year, month):
        row = []
        for day_date in week:
            is_current_month = (day_date.month == month)
            row.append({
                'date': day_date,
                'is_current_month': is_current_month,
                'leaves': day_leaves.get(day_date.isoformat(), []),
                'is_today': day_date == now.date(),
            })
        weeks.append(row)

    teachers = sorted(set(l.teacher for l in approved_leaves), key=lambda t: t.first_name)
    prev_month = (first_day - timedelta(days=1))
    next_month = (last_day + timedelta(days=1))

    return render(request, 'core/dashboard/leave_calendar.html', {
        'weeks': weeks,
        'month_name': first_day.strftime('%B %Y'),
        'year': year, 'month': month,
        'teachers': teachers,
        'prev_month': prev_month.month, 'prev_year': prev_month.year,
        'next_month': next_month.month, 'next_year': next_month.year,
        'LEAVE_TYPE_CHOICES': TeacherLeave.LEAVE_TYPE_CHOICES,
    })


# ═════════════════════════════════════════════════════════════════════════════
# T7 — OFFLINE DATA ENTRY (PWA)
# ═════════════════════════════════════════════════════════════════════════════

@login_required
@require_http_methods(["POST"])
def offline_sync(request):
    """Accept queued offline attendance + marks data; batch-create records"""
    teacher = get_teacher_profile(request.user)
    is_admin = not teacher

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

    attendance_data = data.get('attendance', [])
    marks_data = data.get('marks', [])

    summary = {'attendance_created': 0, 'attendance_updated': 0, 'attendance_errors': 0,
               'marks_created': 0, 'marks_updated': 0, 'marks_errors': 0, 'errors': []}

    for att_entry in attendance_data:
        student_id = att_entry.get('student_id')
        status = att_entry.get('status', 'present')
        date_str = att_entry.get('date', '')
        if not student_id or not date_str:
            summary['attendance_errors'] += 1
            summary['errors'].append(f'Attendance: missing student_id or date')
            continue
        try:
            student = Student.objects.get(id=student_id)
            att_date = date.fromisoformat(date_str)
            subj_id = att_entry.get('subject_id')
            subj = Subject.objects.get(id=subj_id) if subj_id else None
            att, created = Attendance.objects.update_or_create(
                student=student, subject=subj, date=att_date,
                defaults={
                    'status': status,
                    'recorded_by': teacher if teacher else None,
                },
            )
            if created:
                summary['attendance_created'] += 1
            else:
                summary['attendance_updated'] += 1
        except (Student.DoesNotExist, Subject.DoesNotExist, ValueError) as e:
            summary['attendance_errors'] += 1
            summary['errors'].append(f'Attendance student {student_id}: {str(e)}')

    for mark_entry in marks_data:
        student_id = mark_entry.get('student_id')
        subject_id = mark_entry.get('subject_id')
        terminal = mark_entry.get('terminal', '1st')
        marks_obtained = mark_entry.get('marks_obtained')
        total_marks = mark_entry.get('total_marks', 100)

        if not student_id or not subject_id or marks_obtained is None:
            summary['marks_errors'] += 1
            summary['errors'].append(f'Marks: missing required fields')
            continue

        try:
            student = Student.objects.get(id=student_id)
            subject = Subject.objects.get(id=subject_id)
            result, created = Result.objects.update_or_create(
                student=student, subject=subject, terminal=terminal,
                defaults={'marks_obtained': float(marks_obtained), 'total_marks': float(total_marks)},
            )
            if created:
                summary['marks_created'] += 1
            else:
                summary['marks_updated'] += 1
        except (Student.DoesNotExist, Subject.DoesNotExist, ValueError) as e:
            summary['marks_errors'] += 1
            summary['errors'].append(f'Marks student {student_id}: {str(e)}')

    total_errors = summary['attendance_errors'] + summary['marks_errors']
    if total_errors == 0:
        return JsonResponse({'success': True, 'message': 'Sync completed successfully.', 'summary': summary})
    return JsonResponse({'success': True, 'message': f'Synced with {total_errors} errors.', 'summary': summary})


# %% T8 SMART QUESTION BANK %%
@teacher_required
def question_bank(request):
    from core.models import QuestionBank, QUESTION_TYPE_CHOICES, DIFFICULTY_CHOICES
    qs = QuestionBank.objects.filter(teacher=request.teacher).select_related("subject")
    subject_id   = request.GET.get("subject", "").strip()
    chapter      = request.GET.get("chapter", "").strip()
    difficulty   = request.GET.get("difficulty", "").strip()
    qtype        = request.GET.get("type", "").strip()
    if subject_id:   qs = qs.filter(subject_id=subject_id)
    if chapter:      qs = qs.filter(chapter__icontains=chapter)
    if difficulty:   qs = qs.filter(difficulty=difficulty)
    if qtype:        qs = qs.filter(question_type=qtype)

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=question_bank.csv"
        writer = csv.writer(response)
        writer.writerow(["Subject", "Type", "Question", "Options", "Correct", "Difficulty", "Chapter", "Marks"])
        for q in qs:
            opts = "; ".join("{key}: {text}".format(key=o.get("key",""), text=o.get("text","")) for o in (q.options or []))
            writer.writerow([q.subject.code, q.get_question_type_display(), q.question_text,
                             opts, q.correct_answer, q.get_difficulty_display(), q.chapter, q.marks])
        return response

    subjects = request.teacher.subjects.all()
    chapters = sorted(set(QuestionBank.objects.filter(teacher=request.teacher)
                         .values_list("chapter", flat=True).exclude(chapter="")))

    context = {
        "questions": qs, "subjects": subjects, "chapters": chapters,
        "difficulties": DIFFICULTY_CHOICES, "qtypes": QUESTION_TYPE_CHOICES,
        "filter_subject": subject_id, "filter_chapter": chapter,
        "filter_difficulty": difficulty, "filter_type": qtype,
    }
    return render(request, "core/dashboard/question_bank.html", context)


def _qb_form(request, qid=None):
    """Shared form handler for question_bank_add and question_bank_edit"""
    from core.models import QuestionBank, QUESTION_TYPE_CHOICES, DIFFICULTY_CHOICES
    question = None
    if qid:
        question = get_object_or_404(QuestionBank, id=qid, teacher=request.teacher)

    if request.method == "POST":
        subject_id     = request.POST.get("subject")
        question_text  = request.POST.get("question_text", "").strip()
        question_type  = request.POST.get("question_type", "")
        options_raw    = request.POST.get("options_json", "[]")
        correct_answer = request.POST.get("correct_answer", "").strip()
        difficulty     = request.POST.get("difficulty", "medium")
        chapter        = request.POST.get("chapter", "").strip()
        marks          = int(request.POST.get("marks") or 1)
        tags           = request.POST.get("tags", "").strip()
        try:
            options = json.loads(options_raw)
        except (json.JSONDecodeError, ValueError):
            options = question.options if question else []

        if qid:
            question.subject_id = subject_id
            question.question_text = question_text
            question.question_type = question_type
            question.options = options
            question.correct_answer = correct_answer
            question.difficulty = difficulty
            question.chapter = chapter
            question.marks = marks
            question.tags = tags
            question.save()
            messages.success(request, "Question updated successfully.")
        else:
            subject = get_object_or_404(Subject, id=subject_id)
            QuestionBank.objects.create(
                teacher=request.teacher, subject=subject,
                question_text=question_text, question_type=question_type,
                options=options, correct_answer=correct_answer,
                difficulty=difficulty, chapter=chapter, marks=marks, tags=tags,
            )
            messages.success(request, "Question added successfully.")
        return redirect("core:question_bank")

    subjects = request.teacher.subjects.all()
    ctx = {
        "question": question, "subjects": subjects,
        "qtypes": QUESTION_TYPE_CHOICES, "difficulties": DIFFICULTY_CHOICES,
        "options_json": json.dumps(question.options or []) if question else "[]",
    }
    return render(request, "core/dashboard/question_bank_add.html", ctx)


@teacher_required
def question_bank_add(request):
    return _qb_form(request, qid=None)


@teacher_required
def question_bank_edit(request, qid):
    return _qb_form(request, qid=qid)


@teacher_required
def question_bank_delete(request, qid):
    question = get_object_or_404(QuestionBank, id=qid, teacher=request.teacher)
    question.delete()
    messages.success(request, "Question deleted successfully.")
    return redirect("core:question_bank")


@teacher_required
def question_bank_import(request):
    from core.models import QuestionBank
    if request.method == "POST":
        uploaded = request.FILES.get("file")
        if not uploaded:
            messages.error(request, "Please select a file.")
            return redirect("core:question_bank_import")

        try:
            data = json.loads(uploaded.read().decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            try:
                data = json.loads(uploaded.read().decode("latin-1", errors="replace"))
            except json.JSONDecodeError:
                messages.error(request, "Invalid file format. Upload a valid JSON file.")
                return redirect("core:question_bank_import")

        if not isinstance(data, list):
            messages.error(request, "JSON root must be an array of question objects.")
            return redirect("core:question_bank_import")

        created = 0
        for item in data:
            subj_code = item.get("subject") or item.get("subject_code")
            subj = Subject.objects.filter(code=subj_code).first() if subj_code else None
            if not subj or not request.teacher.subjects.filter(id=subj.id).exists():
                continue
            QuestionBank.objects.update_or_create(
                subject=subj,
                question_text=item.get("question_text", "").strip(),
                defaults={
                    "teacher": request.teacher,
                    "question_type": item.get("question_type", "short_answer"),
                    "options": item.get("options", []),
                    "correct_answer": item.get("correct_answer", ""),
                    "difficulty": item.get("difficulty", "medium"),
                    "chapter": item.get("chapter", ""),
                    "marks": int(item.get("marks", 1)),
                    "tags": item.get("tags", ""),
                },
            )
            created += 1
        messages.success(request, f"{created} questions imported successfully.")
        return redirect("core:question_bank")

    return render(request, "core/dashboard/question_bank_import.html")


# ── T8. PAPER TEMPLATE ──────────────────────────────────────────────────────────

@teacher_required
def paper_template_list(request):
    from core.models import QuestionPaperTemplate
    templates = QuestionPaperTemplate.objects.filter(teacher=request.teacher).select_related("subject")
    return render(request, "core/dashboard/paper_template_list.html", {"templates": templates})


@teacher_required
def paper_template_create(request):
    from core.models import QuestionPaperTemplate as QPT
    if request.method == "POST":
        name             = request.POST.get("name", "").strip()
        subject_id       = request.POST.get("subject")
        total_marks      = int(request.POST.get("total_marks") or 100)
        duration_minutes = int(request.POST.get("duration_minutes") or 90)
        try:
            distribution = json.loads(request.POST.get("distribution_json", "{}"))
        except (ValueError, json.JSONDecodeError):
            distribution = {}
        try:
            chapters = json.loads(request.POST.get("chapters_json", "[]"))
        except (ValueError, json.JSONDecodeError):
            chapters = []
        subject = get_object_or_404(Subject, id=subject_id)
        QPT.objects.create(
            name=name, subject=subject, total_marks=total_marks,
            duration_minutes=duration_minutes, distribution=distribution,
            chapters=chapters, teacher=request.teacher,
        )
        messages.success(request, "Paper template created.")
        return redirect("core:paper_template_list")

    subjects = request.teacher.subjects.all()
    return render(request, "core/dashboard/paper_template_create.html", {"subjects": subjects})


@teacher_required
def paper_template_generate(request, tid):
    from core.models import QuestionPaperTemplate as QPT
    template = get_object_or_404(QPT, id=tid, teacher=request.teacher)
    dist = template.distribution or {}
    qs = QuestionBank.objects.filter(
        teacher=request.teacher, subject=template.subject, difficulty__in=["easy", "medium", "hard"]
    )
    if template.chapters:
        qs = qs.filter(chapter__in=template.chapters)
    selected = []
    for diff, cnt in dist.items():
        selected.extend(list(qs.filter(difficulty=diff)[:int(cnt)]))
    total_marks = sum(q.marks for q in selected)
    return render(request, "core/dashboard/paper_template_generate.html", {
        "template": template, "questions": selected, "total_marks": total_marks,
    })


# %% T9 REMINDER SCHEDULER %%
@teacher_required
def reminder_list(request):
    from core.models import Reminder
    reminders = Reminder.objects.filter(teacher=request.teacher).select_related("student", "subject")
    return render(request, "core/dashboard/reminder_list.html", {"reminders": reminders})


@teacher_required
def reminder_create(request):
    from core.models import Reminder, REMINDER_TYPE_CHOICES, RECURRENCE_CHOICES
    if request.method == "POST":
        reminder_type    = request.POST.get("reminder_type", "")
        student_id       = request.POST.get("student", "").strip()
        subject_id       = request.POST.get("subject", "").strip()
        title            = request.POST.get("title", "").strip()
        message          = request.POST.get("message", "").strip()
        scheduled_for    = request.POST.get("scheduled_for", "").strip()
        recurrence       = request.POST.get("recurrence", "none")
        recurrence_until = request.POST.get("recurrence_until", "").strip()
        student = Student.objects.filter(id=student_id).first() if student_id else None
        subject = Subject.objects.filter(id=subject_id).first() if subject_id else None
        ru_date = None
        if recurrence_until:
            try:
                ru_date = datetime.strptime(recurrence_until, "%Y-%m-%d").date()
            except ValueError:
                pass
        Reminder.objects.create(
            teacher=request.teacher, student=student, subject=subject,
            reminder_type=reminder_type, title=title, message=message,
            scheduled_for=scheduled_for, recurrence=recurrence,
            recurrence_until=ru_date,
        )
        messages.success(request, "Reminder created.")
        return redirect("core:reminder_list")

    subjects = request.teacher.subjects.all()
    students = request.teacher.students.all().order_by("name")
    return render(request, "core/dashboard/reminder_create.html", {
        "subjects": subjects, "students": students,
        "rtypes": REMINDER_TYPE_CHOICES, "recurrences": RECURRENCE_CHOICES,
    })


@teacher_required
@require_http_methods(["POST"])
def reminder_delete(request, reminder_id):
    reminder = get_object_or_404(Reminder, id=reminder_id, teacher=request.teacher)
    reminder.delete()
    messages.success(request, "Reminder deleted.")
    return redirect("core:reminder_list")


@_admin_required
def scheduled_tasks_list(request):
    from core.models import ScheduledTask, SCHEDULED_TASK_TYPE_CHOICES
    tasks = ScheduledTask.objects.all().order_by("-scheduled_for")[:200]
    return render(request, "core/dashboard/scheduled_tasks_list.html", {
        "tasks": tasks, "task_types": SCHEDULED_TASK_TYPE_CHOICES,
    })


# %% T10 SMART SUBJECT ANALYSIS %%
@teacher_required
def subject_analytics(request):
    from core.models import QuestionBank, QuestionPaperTemplate
    teacher         = request.teacher
    subject_id      = request.GET.get("subject", "").strip()
    terminal        = request.GET.get("terminal", "all")

    teacher_subjects = teacher.subjects.all()
    teacher_students = teacher.students.all()
    all_subjects = teacher_subjects.order_by("name")

    base_qs = Result.objects.filter(
        student__in=teacher_students, subject__in=all_subjects
    ).select_related("student", "subject")
    if subject_id:
        base_qs = base_qs.filter(subject_id=subject_id)

    bins_labels = ["0-39", "40-49", "50-59", "60-69", "70-79", "80-89", "90-100"]
    bins = [0] * 7
    pct_list = []
    for r in base_qs:
        pct = (r.marks_obtained / r.total_marks * 100) if r.total_marks > 0 else 0
        pct_list.append(pct)
        if pct < 40:    bins[0] += 1
        elif pct < 50:  bins[1] += 1
        elif pct < 60:  bins[2] += 1
        elif pct < 70:  bins[3] += 1
        elif pct < 80:  bins[4] += 1
        elif pct < 90:  bins[5] += 1
        else:           bins[6] += 1

    total       = len(pct_list)
    avg_pct     = round(sum(pct_list) / total, 1) if total else 0
    highest_pct = round(max(pct_list), 1) if pct_list else 0
    lowest_pct  = round(min(pct_list), 1) if pct_list else 0
    pass_count  = sum(1 for p in pct_list if p >= 40)
    pass_rate   = round((pass_count / total) * 100, 1) if total else 0
    below_40    = sum(1 for p in pct_list if p < 40)

    insight_lines = []
    if below_40 > 0:
        s = "s" if below_40 > 1 else ""
        insight_lines.append(
            f"[!] Warning: {below_40} student{s} scored below 40% - consider arranging a remedial class."
        )
    if avg_pct >= 80:
        insight_lines.append(f"Excellent! Class average is {avg_pct}%.")
    elif avg_pct >= 70:
        insight_lines.append(f"Good progress - class average is {avg_pct}%.")
    elif avg_pct >= 50:
        insight_lines.append(f"Class average is {avg_pct}%. Focus on weaker students.")
    else:
        insight_lines.append(f"Class average is {avg_pct}%. Needs urgent attention.")

    terminals = ["1st", "2nd", "3rd", "Final"]
    run_terminals = terminals if terminal == "all" else [terminal]

    trender = {}
    student_result_qs = base_qs.order_by("student__name", "terminal")
    for r in student_result_qs:
        sid = r.student.id
        if sid not in trender:
            trender[sid] = {"student": r.student, "marks_by_term": {}}
        trender[sid]["marks_by_term"][r.terminal] = r.marks_obtained

    trend_rows = []
    for sid in sorted(trender, key=lambda s: trender[s]["student"].name):
        info = trender[sid]
        student  = info["student"]
        mbt      = info["marks_by_term"]
        prev_vals = [mbt.get(t) for t in terminals[:-1]]
        if len(prev_vals) >= 2 and prev_vals[-1] is not None and prev_vals[-2] is not None:
            diff   = prev_vals[-1] - prev_vals[-2]
            arrow  = "up" if diff > 0.5 else ("down" if diff < -0.5 else "same")
        else:
            arrow = "same"
        # Build an aligned marks list so the template can iterate without dict keys
        marks_list = [mbt.get(t) for t in running_terminals]
        mark_tuples = list(zip(running_terminals, marks_list))
        trend_rows.append({"student": student, "marks_list": marks_list,
                           "mark_tuples": mark_tuples, "arrow": arrow})

    available_classes = sorted(set(s.student_class for s in teacher_students))

    return render(request, "core/dashboard/subject_analytics.html", {
        "subject_obj": Subject.objects.filter(id=subject_id).first(),
        "subject_id": subject_id, "terminal": terminal,
        "running_terminals": run_terminals,
        "total": total, "avg_pct": avg_pct, "highest_pct": highest_pct,
        "lowest_pct": lowest_pct, "pass_rate": pass_rate,
        "pass_count": pass_count, "below_40": below_40,
        "bins_labels": bins_labels, "bins": bins,
        "trend_rows": trend_rows, "insight_lines": insight_lines,
        "available_subjects": all_subjects, "available_classes": available_classes,
        "terminals": terminals, "is_teacher": True,
    })


# %% A12 AUTOMATED DIGESTS (admin-triggered) %%
@_admin_required
def trigger_weekly_digest(request):
    from core.models import ScheduledTask
    from datetime import timedelta
    now  = timezone.now()
    count = 0
    for student in Student.objects.all():
        ScheduledTask.objects.create(
            task_type="sms_digest",
            scheduled_for=now + timedelta(minutes=count),
            payload={"student_id": student.id, "type": "weekly_attendance"},
        )
        count += 1
    messages.success(request, f"Queued {count} weekly attendance digest tasks.")
    return redirect("core:scheduled_tasks_list")


@_admin_required
def trigger_monthly_report(request):
    from core.models import ScheduledTask
    from datetime import timedelta
    now  = timezone.now()
    count = 0
    for student in Student.objects.all():
        ScheduledTask.objects.create(
            task_type="sms_digest",
            scheduled_for=now + timedelta(minutes=count),
            payload={"student_id": student.id, "type": "monthly_report"},
        )
        count += 1
    messages.success(request, f"Queued {count} monthly report tasks.")
    return redirect("core:scheduled_tasks_list")


# ═════════════════════════════════════════════════════════════════════════════
# A3 — INVOICE BILLING / SUBSCRIPTION MODULE
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def subscription_management(request):
    """List subscriptions and allow creating/upgrading."""
    subscriptions = Subscription.objects.select_related('created_by').all()
    selected_sub = None
    if request.GET.get('id'):
        selected_sub = Subscription.objects.filter(id=request.GET['id']).first()
    if request.method == 'POST':
        inst = request.POST.get('institution', '').strip()
        plan = request.POST.get('plan', 'starter')
        price = request.POST.get('monthly_price', '0')
        expires = request.POST.get('expires_at', '')
        renew = request.POST.get('auto_renew') == 'on'
        active = request.POST.get('is_active') == 'on'
        sub = Subscription.objects.create(
            institution=inst, plan=plan, monthly_price=price or 0,
            is_active=active, auto_renew=renew,
            created_by=request.user,
            expires_at=timezone.make_aware(datetime.fromisoformat(expires)) if expires else None,
        )
        messages.success(request, f'Subscription created for {inst}.')
        return redirect('core:subscription_management')
    return render(request, 'core/dashboard/subscription_management.html', {
        'subscriptions': subscriptions, 'selected_sub': selected_sub,
        'plan_choices': Subscription.PLAN_CHOICES,
    })


@_admin_required
def invoice_list(request):
    """List all invoices with status filter."""
    status = request.GET.get('status', '')
    invoices = Invoice.objects.select_related('subscription', 'created_by').all()
    if status:
        invoices = invoices.filter(status=status)
    return render(request, 'core/dashboard/invoice_list.html', {
        'invoices': invoices,
        'status_filter': status,
        'status_choices': Invoice.INVOICE_STATUS_CHOICES,
    })


@_admin_required
def invoice_create(request):
    """Create a manual invoice with line items."""
    if request.method == 'POST':
        form = InvoiceQuickCreateForm(request.POST)
        if form.is_valid():
            try:
                line_items = json.loads(form.cleaned_data['line_items'] or '[]')
            except json.JSONDecodeError:
                line_items = []

            sub_id = form.cleaned_data.get('subscription')
            subscription = None
            if sub_id:
                try:
                    subscription = Subscription.objects.get(id=sub_id)
                except Subscription.DoesNotExist:
                    pass

            inv = Invoice.objects.create(
                subscription=subscription,
                issued_to=form.cleaned_data['issued_to'],
                amount=form.cleaned_data['amount'],
                tax_amount=form.cleaned_data.get('tax_amount') or 0,
                issue_date=form.cleaned_data['issue_date'],
                due_date=form.cleaned_data['due_date'],
                status='DRAFT',
                line_items=line_items,
                created_by=request.user,
            )
            messages.success(request, f'Invoice {inv.invoice_number} created.')
            return redirect('core:invoice_detail', invoice_number=inv.invoice_number)
    else:
        form = InvoiceQuickCreateForm()

    subs = Subscription.objects.filter(is_active=True)
    form.fields['subscription'].choices = [('', '— No subscription —')] + [
        (str(s.id), f'{s.institution} ({s.get_plan_display()})') for s in subs
    ]
    return render(request, 'core/dashboard/invoice_create.html', {
        'form': form, 'status_choices': Invoice.INVOICE_STATUS_CHOICES,
    })


@_admin_required
def invoice_detail(request, invoice_number):
    """Invoice preview/print page."""
    inv = get_object_or_404(Invoice.objects.select_related('subscription', 'created_by'), invoice_number=invoice_number)
    return render(request, 'core/dashboard/invoice_detail.html', {
        'invoice': inv,
    })


@_admin_required
def invoice_pdf(request, invoice_number):
    """PDF download using ReportLab."""
    from django.http import FileResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    inv = get_object_or_404(Invoice.objects.select_related('subscription', 'created_by'), invoice_number=invoice_number)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    accent = colors.HexColor('#4F46E5')

    title_style = ParagraphStyle('_title', parent=styles['Title'], textColor=accent, fontSize=22, spaceAfter=4)
    body_style  = ParagraphStyle('_body',  parent=styles['Normal'],  fontSize=10, spaceAfter=2)
    hdr_style   = ParagraphStyle('_hdr',   parent=styles['Normal'],  fontSize=9,  textColor=colors.grey, spaceAfter=2)

    elems = []
    elems.append(Paragraph('ACADEMIC MANAGEMENT SYSTEM', title_style))
    elems.append(Paragraph('Invoice', ParagraphStyle('_inv', parent=styles['Normal'], fontSize=16, textColor=accent, spaceAfter=2)))
    elems.append(Spacer(1, 0.5*cm))
    elems.append(HRFlowable(width='100%', thickness=2, color=accent, spaceAfter=0.5*cm))

    # Info table
    info = [
        ['Invoice #', inv.invoice_number],
        ['Issued To', inv.issued_to],
        ['Issue Date', inv.issue_date.strftime('%d %B %Y') if inv.issue_date else '—'],
        ['Due Date',   inv.due_date.strftime('%d %B %Y') if inv.due_date else '—'],
        ['Status',     inv.get_status_display()],
    ]
    info_rows = [[Paragraph(k, body_style), Paragraph(str(v), body_style)] for k, v in info]
    info_tbl = Table(info_rows, colWidths=[5*cm, 11*cm])
    info_tbl.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#EEF2FF')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
    ]))
    elems.append(info_tbl)
    elems.append(Spacer(1, 0.5*cm))

    # Line items
    items = inv.line_items or []
    if items:
        elems.append(Paragraph('Line Items', body_style))
        item_hdr = Table(
            [[Paragraph(h, body_style) for h in ['Description','Qty','Unit Price','Total']]],
            colWidths=[8*cm, 2*cm, 3*cm, 3*cm]
        )
        item_hdr.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), accent),
            ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        elems.append(item_hdr)
        for li in items:
            desc   = li.get('description', '')
            qty    = str(li.get('qty', 1))
            uprice = str(li.get('unit_price', 0))
            total  = str(li.get('total', 0))
            row_tbl = Table(
                [[Paragraph(desc, body_style), Paragraph(qty, body_style),
                  Paragraph(uprice, body_style), Paragraph(total, body_style)]],
                colWidths=[8*cm, 2*cm, 3*cm, 3*cm]
            )
            row_tbl.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            elems.append(row_tbl)
        elems.append(Spacer(1, 0.3*cm))

    # Totals
    tot_data = [
        ['Subtotal:', f'Rs. {inv.amount:,.2f}'],
        ['Tax:',       f'Rs. {inv.tax_amount:,.2f}'],
        ['TOTAL:',     f'Rs. {inv.total_amount:,.2f}'],
    ]
    tot_rows = [[Paragraph(k, body_style), Paragraph(str(v), body_style)] for k, v in tot_data]
    tot_table = Table(tot_rows, colWidths=[8*cm, 8*cm], hAlign='RIGHT')
    tot_table.setStyle(TableStyle([
        ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#EEF2FF')),
        ('FONTNAME',   (0,2), (-1,2), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elems.append(tot_table)
    elems.append(Spacer(1, 1*cm))
    elems.append(HRFlowable(width='100%', thickness=1, color=colors.grey, spaceAfter=0.3*cm))
    elems.append(Paragraph(f'Generated on  {timezone.now().strftime("%d %B %Y %H:%M")}', hdr_style))

    doc.build(elems)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{inv.invoice_number}.pdf')


@_admin_required
def invoice_mark_paid(request, invoice_id):
    """POST-only: mark invoice as paid."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)
    inv = get_object_or_404(Invoice, id=invoice_id)
    inv.status = 'PAID'
    inv.paid_date = timezone.now().date()
    inv.save()
    messages.success(request, f'Invoice {inv.invoice_number} marked as paid.')
    return redirect('core:invoice_list')


# ═════════════════════════════════════════════════════════════════════════════
# A8 — ANALYTICS EXPORT BUNDLE
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def export_smart_dashboard_pdf(request):
    """Generate a PDF export of the smart analytics dashboard."""
    if not SystemConfig.get_bool('enable_export_bundle', True):
        messages.warning(request, 'Export bundle is disabled by system configuration.')
        return redirect('core:smart_dashboard')
    # Enforce same filters as smart_dashboard
    req_args = {'path': request.path, 'GET': request.GET.copy()}
    return export_utils.export_smart_dashboard_pdf(request)


@_admin_required
def export_smart_dashboard_excel(request):
    """Generate an Excel workbook export of the smart analytics dashboard."""
    if not SystemConfig.get_bool('enable_export_bundle', True):
        messages.warning(request, 'Export bundle is disabled by system configuration.')
        return redirect('core:smart_dashboard')
    # Build context by calling the same idea as smart_dashboard but manually
    from django.db.models import Avg, Sum as _Sum
    period = request.GET.get('period', 'all')
    selected_class = request.GET.get('student_class', '')
    all_results = Result.objects.select_related('student', 'subject')
    if selected_class:
        all_results = all_results.filter(student__student_class=selected_class)

    gs = GradeScale.objects.filter(is_active=True).first()
    pass_mark_pct = gs.pass_mark_percent if gs else 40.0

    student_scores = []
    for stu in Student.objects.prefetch_related('result_set'):
        res = stu.result_set.all()
        if not res.exists():
            continue
        obt = sum(r.marks_obtained for r in res)
        pos = sum(r.total_marks for r in res)
        p = (obt / pos * 100) if pos > 0 else 0
        student_scores.append({'student': stu, 'percentage': round(p, 1), 'obtained': obt, 'total': pos})
    student_scores.sort(key=lambda x: x['percentage'], reverse=True)
    top_students    = student_scores[:10]
    weak_students   = [s for s in student_scores if s['percentage'] < pass_mark_pct][:10]

    subj_stats = []
    for subj in Subject.objects.all():
        s_res = all_results.filter(subject=subj)
        if s_res.exists():
            avg = s_res.aggregate(av=Avg('marks_obtained'))['av'] or 0
            pos_sum = s_res.aggregate(s=_Sum('total_marks'))['s'] or 0
            obt_sum = s_res.aggregate(s=_Sum('marks_obtained'))['s'] or 0
            pct = (obt_sum / pos_sum * 100) if pos_sum > 0 else 0
            subj_stats.append({'subject': subj.name, 'avg': round(avg,1), 'percentage': round(pct,1)})
    subj_stats.sort(key=lambda x: x['percentage'])

    class_stats = []
    for cls in Student.objects.values_list('student_class', flat=True).distinct().order_by('student_class'):
        cls_res = all_results.filter(student__student_class=cls)
        if cls_res.exists():
            pos = cls_res.aggregate(s=_Sum('total_marks'))['s'] or 0
            obt = cls_res.aggregate(s=_Sum('marks_obtained'))['s'] or 0
            pct = (obt / pos * 100) if pos > 0 else 0
            class_stats.append({'class': cls, 'students': Student.objects.filter(student_class=cls).count(), 'percentage': round(pct,1)})

    term_stats = []
    for term in ['1st','2nd','3rd','Final']:
        t_res = all_results.filter(terminal=term)
        if t_res.exists():
            pos = t_res.aggregate(s=_Sum('total_marks'))['s'] or 0
            obt = t_res.aggregate(s=_Sum('marks_obtained'))['s'] or 0
            pct = (obt / pos * 100) if pos > 0 else 0
            term_stats.append({'terminal': term, 'percentage': round(pct,1)})

    att_present = Attendance.objects.filter(status='present').count()
    att_total   = Attendance.objects.count()
    att_pct     = round((att_present / att_total * 100), 1) if att_total > 0 else 0

    context = {
        'total_students':   Student.objects.count(),
        'total_subjects':   Subject.objects.count(),
        'total_results':    all_results.count(),
        'total_teachers':   Teacher.objects.filter(is_active=True).count(),
        'top_students':     top_students,
        'weak_students':    weak_students,
        'subj_stats':       subj_stats,
        'class_stats':      class_stats,
        'term_stats':       term_stats,
        'att_present':      att_present,
        'att_total':        att_total,
        'att_pct':          att_pct,
        'selected_class':   selected_class,
        'pass_mark_pct':    pass_mark_pct,
    }
    return export_utils.export_smart_dashboard_excel(request, context)


# ═════════════════════════════════════════════════════════════════════════════
# A9 — CERTIFICATE GENERATOR
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def certificate_templates(request):
    """List / edit HTML certificate templates."""
    templates = CertificateTemplate.objects.all()
    if request.method == 'POST':
        tpl_id = request.POST.get('tpl_id')
        html = request.POST.get('html_content', '')
        if tpl_id:
            tpl = get_object_or_404(CertificateTemplate, id=tpl_id)
            tpl.html_content = html
            tpl.is_active = request.POST.get('is_active') == 'on'
            tpl.save()
            messages.success(request, f'Template "{tpl.name}" updated.')
        return redirect('core:certificate_templates')
    return render(request, 'core/dashboard/certificate_templates.html', {
        'templates': templates,
    })


@_admin_required
def generate_certificate(request):
    """Generate a new certificate PDF-ready record."""
    if request.method == 'POST':
        form = CertificateGenerateForm(request.POST)
        if form.is_valid():
            student = form.cleaned_data['student']
            cert_type = form.cleaned_data['certificate_type']
            template = CertificateTemplate.objects.filter(
                name__icontains=cert_type.lower(), is_active=True
            ).first()
            if not template:
                template = CertificateTemplate.objects.filter(is_active=True).first()
            if not template:
                messages.error(request, 'No active certificate template found. Create one first.')
                return redirect('core:certificate_templates')

            # Build QR verification string
            qr_str = f"ACADSTAT:CERT:{student.roll_number}:{student.name}:{cert_type}"

            cert = Certificate.objects.create(
                student=student,
                template=template,
                certificate_type=cert_type,
                issued_by=request.user,
                qr_data_string=qr_str,
                status='ISSUED',
            )
            try:
                NotificationService.log_activity(
                    request.user, 'other',
                    f'Certificate generated: {cert.certificate_number} for {student.name}',
                )
            except Exception:
                pass
            messages.success(request, f'Certificate {cert.certificate_number} issued for {student.name}.')
            return redirect('core:certificate_list')
    else:
        form = CertificateGenerateForm()
    return render(request, 'core/dashboard/generate_certificate.html', {
        'form': form,
    })


@_admin_required
def certificate_list(request):
    """List all issued/draft certificates."""
    status = request.GET.get('status', '')
    certs = Certificate.objects.select_related('student', 'template', 'issued_by')
    if status:
        certs = certs.filter(status=status)
    return render(request, 'core/dashboard/certificate_list.html', {
        'certificates': certs.order_by('-created_at'),
        'status_filter': status,
        'status_choices': Certificate.CERT_STATUS_CHOICES,
    })


@_admin_required
def certificate_view(request, cert_number):
    """Rendered certificate page."""
    cert = get_object_or_404(Certificate.objects.select_related('student', 'template', 'issued_by'),
                              certificate_number=cert_number)
    return render(request, 'core/dashboard/certificate_view.html', {
        'cert': cert,
    })


@_admin_required
def certificate_pdf(request, cert_number):
    """Generate PDF blob for a certificate using ReportLab."""
    from django.http import FileResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import qrcode as qr_gen
    from io import BytesIO

    cert = get_object_or_404(Certificate.objects.select_related('student', 'template', 'issued_by'),
                              certificate_number=cert_number)
    student = cert.student

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
        leftMargin=2*cm, rightMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    gold   = colors.HexColor('#D4A017')
    navy   = colors.HexColor('#1E3A5F')
    white  = colors.white
    black  = colors.HexColor('#1a1a1a')

    T = ParagraphStyle('_T', fontName='Helvetica-Bold',  fontSize=24, textColor=gold,
                       alignment=TA_CENTER, spaceAfter=6)
    S = ParagraphStyle('_S', fontName='Helvetica-Bold',  fontSize=14, textColor=navy,
                       alignment=TA_CENTER, spaceAfter=4)
    SN = ParagraphStyle('_SN', fontName='Helvetica', fontSize=12, textColor=black,
                        alignment=TA_CENTER, spaceAfter=2)
    FN = ParagraphStyle('_FN', fontName='Helvetica', fontSize=10, textColor=colors.grey,
                        alignment=TA_CENTER)
    SFN = ParagraphStyle('_SFN', fontName='Helvetica-Bold', fontSize=10, textColor=black,
                         alignment=TA_CENTER)

    # QR code
    qr_buf = BytesIO()
    qr_img = qr_gen.make(cert.qr_data_string or cert.certificate_number)
    qr_img.save(qr_buf, format='PNG')
    qr_buf.seek(0)

    elems = []
    elems.append(Spacer(1, 0.5*cm))
    elems.append(Paragraph('ACADEMIC MANAGEMENT SYSTEM', ParagraphStyle(
        '_inst', fontName='Helvetica-Bold', fontSize=12, textColor=navy, alignment=TA_CENTER, spaceAfter=4)))
    elems.append(Paragraph(cert.get_certificate_type_display(), T))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(HRFlowable(width='100%', thickness=2, color=gold, spaceAfter=0.3*cm))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(Paragraph('This is to certify that', SN))
    elems.append(Paragraph(student.name, S))
    elems.append(Spacer(1, 0.2*cm))
    elems.append(Paragraph(f'Roll No: {student.roll_number or "N/A"}', FN))
    elems.append(Paragraph(f'Class   : {student.student_class} {student.section}', FN))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(HRFlowable(width='100%', thickness=0.5, color=gold, spaceAfter=0.3*cm))
    elems.append(Spacer(1, 0.3*cm))
    elems.append(Paragraph(f'Certificate No: {cert.certificate_number}', FN))
    elems.append(Paragraph(f'Issued Date   : {cert.issued_date.strftime("%d %B %Y") if cert.issued_date else "—"}', FN))
    issued_name = cert.issued_by.get_full_name() if cert.issued_by else 'Administrator'
    elems.append(Paragraph(f'Issued By     : {issued_name}', FN))
    elems.append(Spacer(1, 1*cm))

    # QR + signature
    qr_tbl = Table(
        [[Paragraph(f'Scan to verify:<br/>{cert.certificate_number}', FN), None]],
        colWidths=[5*cm, 5*cm]
    )
    qr_data = qr_buf.read()
    from reportlab.platypus import Image as RLImage
    qr_img_rl = RLImage(io.BytesIO(qr_data), width=2.5*cm, height=2.5*cm)
    qr_tbl = Table([[qr_img_rl, Paragraph('Authorised Signature', FN)]], colWidths=[5*cm, 9*cm])
    qr_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elems.append(qr_tbl)
    elems.append(Spacer(1, 0.5*cm))
    elems.append(HRFlowable(width='100%', thickness=1, color=gold))
    elems.append(Paragraph(f'Verified at AcadStat — {timezone.now().strftime("%d %B %Y")}', FN))

    doc.build(elems)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'{cert.certificate_number}.pdf')


def verify_certificate(request):
    """Public (or login): search by certificate number and show certificate or 'not found'."""
    cert_number = request.GET.get('certificate_number', '').strip() or request.POST.get('certificate_number', '').strip()
    cert = None
    if cert_number:
        cert = Certificate.objects.select_related('student', 'template').filter(
            certificate_number=cert_number
        ).first()
    return render(request, 'core/dashboard/verify_certificate.html', {
        'cert_number': cert_number, 'cert': cert,
    })


# ═════════════════════════════════════════════════════════════════════════════
# A5 — SUPPORT TICKET SYSTEM
# ═════════════════════════════════════════════════════════════════════════════

@login_required
def ticket_list(request):
    """Students see their own tickets, teachers see all, admins see all."""
    is_admin = _is_admin_user(request.user)
    teacher = get_teacher_profile(request.user)
    qs = SupportTicket.objects.select_related('reported_by', 'assigned_to', 'resolved_by')
    if not is_admin and teacher:
        qs = qs.filter(Q(reported_by=request.user) | Q(assigned_to=teacher))
    elif not is_admin:
        qs = qs.filter(reported_by=request.user)
    tickets = qs.order_by('-created_at')
    return render(request, 'core/dashboard/ticket_list.html', {
        'tickets': tickets, 'is_admin': is_admin, 'is_teacher': bool(teacher),
        'status_choices': SupportTicket.TICKET_STATUS_CHOICES,
        'priority_choices': SupportTicket.TICKET_PRIORITY_CHOICES,
    })


@login_required
def ticket_create(request):
    """Create a new support ticket."""
    if request.method == 'POST':
        form = SupportTicketForm(request.POST, request.FILES)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.reported_by = request.user
            ticket.status = 'Open'
            ticket.save()
            messages.success(request, f'Ticket {ticket.ticket_id} created successfully.')
            return redirect('core:ticket_detail', ticket_id=ticket.ticket_id)
    else:
        form = SupportTicketForm()
    return render(request, 'core/dashboard/ticket_create.html', {
        'form': form,
    })


@login_required
def ticket_detail(request, ticket_id):
    """Full ticket thread with comments, status change."""
    ticket = get_object_or_404(SupportTicket.objects.select_related(
        'reported_by', 'assigned_to', 'resolved_by'), ticket_id=ticket_id)
    if request.method == 'POST':
        # Comment
        if 'comment' in request.POST:
            body = request.POST.get('body', '').strip()
            if body:
                TicketComment.objects.create(
                    ticket=ticket, author=request.user, body=body,
                    is_internal=request.POST.get('is_internal') == 'on',
                )
                messages.success(request, 'Comment added.')
            return redirect('core:ticket_detail', ticket_id=ticket_id)

        # Status update
        if 'update_status' in request.POST:
            new_status = request.POST.get('status', '').strip()
            if new_status in dict(SupportTicket.TICKET_STATUS_CHOICES):
                ticket.status = new_status
                ticket.priority = request.POST.get('priority', ticket.priority)
                if new_status == 'Resolved':
                    ticket.resolved_at = timezone.now()
                    ticket.resolved_by = request.user
                    if not ticket.resolution_notes:
                        ticket.resolution_notes = request.POST.get('resolution_notes', '')
                if request.POST.get('assigned_to'):
                    try:
                        ticket.assigned_to = Teacher.objects.get(
                            id=int(request.POST['assigned_to']))
                    except (Teacher.DoesNotExist, ValueError):
                        pass
                ticket.save()
                if new_status == 'Resolved':
                    ticket_notify(ticket)
                NotificationService.log_activity(
                    request.user, 'other',
                    f'Status updated: Ticket {ticket.ticket_id} → {Ticket.objects.filter(id=ticket.id).get()}',
                )
                messages.success(request, f'Ticket status updated to {new_status}.')
            return redirect('core:ticket_detail', ticket_id=ticket_id)

    comments = ticket.comments.select_related('author').all()
    teachers = Teacher.objects.filter(is_active=True).order_by('first_name')
    return render(request, 'core/dashboard/ticket_detail.html', {
        'ticket': ticket, 'comments': comments, 'teachers': teachers,
    })


@login_required
def ticket_update_status(request, ticket_id):
    """POST-only: update ticket status/priority/assignment."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)
    ticket = get_object_or_404(SupportTicket, ticket_id=ticket_id)
    new_status = request.POST.get('status', '').strip()
    if new_status in dict(SupportTicket.TICKET_STATUS_CHOICES):
        ticket.status = new_status
        ticket.priority = request.POST.get('priority', ticket.priority)
        if new_status == 'Resolved':
            ticket.resolved_at = timezone.now()
            ticket.resolved_by = request.user
        if request.POST.get('assigned_to'):
            try:
                ticket.assigned_to = Teacher.objects.get(id=int(request.POST['assigned_to']))
            except (Teacher.DoesNotExist, ValueError):
                pass
        ticket.save()
        if new_status == 'Resolved':
            ticket_notify(ticket)
        return JsonResponse({'success': True, 'status': ticket.status})
    return JsonResponse({'success': False, 'message': 'Invalid status'}, status=400)


@login_required
def ticket_comment(request, ticket_id):
    """POST-only: add a comment to a ticket."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)
    ticket = get_object_or_404(SupportTicket, ticket_id=ticket_id)
    body = request.POST.get('body', '').strip()
    if not body:
        return JsonResponse({'success': False, 'message': 'Body required'}, status=400)
    TicketComment.objects.create(
        ticket=ticket, author=request.user, body=body,
        is_internal=request.POST.get('is_internal') == 'on',
    )
    return JsonResponse({'success': True})


def ticket_notify(ticket):
    """When a ticket is resolved, email the reporter."""
    try:
        reporter = ticket.reported_by
        ticket_title = ticket.title
        ticket_id = ticket.ticket_id
        if reporter and reporter.email:
            notify_student(
                Student.objects.filter(email=reporter.email).first() or Student.objects.get(name=reporter.username) if Student.objects.filter(email=reporter.email).exists() else None,
                title=f'Ticket Resolved: {ticket_title}',
                message=f'Your ticket ({ticket_id}) has been resolved. {ticket.resolution_notes or "Thank you for your patience."}',
            ) if False else _email_ticket_notify_helper(ticket)
    except Exception as exc:
        logger = __import__('logging').getLogger(__name__)
        logger.warning(f'ticket_notify: {exc}')


def _email_ticket_notify_helper(ticket):
    from core.email_service import send_mail
    reporter = ticket.reported_by
    if reporter and reporter.email:
        send_mail(
            subject=f'Ticket Resolved: {ticket.title}',
            body=f'Your ticket {ticket.ticket_id} has been resolved.\n\n{ticket.resolution_notes}',
            recipient_list=[reporter.email],
        )


# ═════════════════════════════════════════════════════════════════════════════
# A14 — API KEY & WEBHOOK SYSTEM (DEVELOPER PORTAL)
# ═════════════════════════════════════════════════════════════════════════════

@_admin_required
def developer_portal(request):
    """Overview: API usage summary, active keys, active webhooks, failures."""
    keys   = APIKey.objects.all()
    active_keys = keys.filter(is_active=True).count()
    total_keys  = keys.count()
    webhooks    = WebhookEndpoint.objects.all()
    active_webhooks = webhooks.filter(is_active=True).count()
    failures    = list(WebhookDeliveryLog.objects.filter(success=False).order_by('-delivered_at')[:20])
    daily_logs  = WebhookEndpoint.objects.select_related('created_by').prefetch_related(
        'deliveries').all().order_by('-last_triggered_at')[:10]

    return render(request, 'core/dashboard/developer_portal.html', {
        'total_keys': total_keys, 'active_keys': active_keys,
        'active_webhooks': active_webhooks, 'total_webhooks': webhooks.count(),
        'recent_failures': list(failures),
        'daily_logs': daily_logs,
    })


@_admin_required
def manage_api_keys(request):
    """CRUD on API keys."""
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'create':
            prefix = request.POST.get('prefix', 'acad_').strip()
            name = request.POST.get('name', '').strip()
            scopes_raw = request.POST.get('scopes', '').strip()
            scopes = [s.strip() for s in scopes_raw.split(',') if s.strip()]
            ip_whitelist_raw = request.POST.get('allowed_ips', '').strip()
            allowed_ips = [ip.strip() for ip in ip_whitelist_raw.split(',') if ip.strip()]
            expires_raw = request.POST.get('expires_at', '')
            expires_at = timezone.make_aware(datetime.fromisoformat(expires_raw)) if expires_raw else None

            raw_key = uuid.uuid4().hex + uuid.uuid4().hex
            full_key = f"{prefix}{raw_key[:32]}"
            APIKey.objects.create(
                key=full_key, name=name, prefix=prefix,
                scopes=scopes, allowed_ips=allowed_ips,
                expires_at=expires_at,
                created_by=request.user,
            )
            activity_log_flat(request.user, 'other', f'Created API key: {name}')
            messages.success(request, f'API key created: {full_key[:16]}...')
            return redirect('core:manage_api_keys')

        if action == 'revoke':
            key_id = request.POST.get('key_id')
            k = APIKey.objects.filter(id=key_id).first()
            if k:
                k.is_active = False
                k.save()
                messages.success(request, f'API key "{k.name}" revoked.')
            return redirect('core:manage_api_keys')

    keys = APIKey.objects.select_related('created_by').order_by('-created_at')
    return render(request, 'core/dashboard/manage_api_keys.html', {
        'keys': keys,
    })


@_admin_required
def manage_webhooks(request):
    """CRUD on webhook endpoints."""
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'create':
            url = request.POST.get('url', '').strip()
            events_raw = request.POST.get('events', '').strip()
            events = [e.strip() for e in events_raw.split(',') if e.strip()]
            if not url:
                messages.error(request, 'URL is required.')
            else:
                secret = uuid.uuid4().hex
                WebhookEndpoint.objects.create(
                    url=url, secret=secret, events=events, created_by=request.user,
                )
                messages.success(request, f'Webhook endpoint created: {url}')
            return redirect('core:manage_webhooks')
        if action == 'toggle':
            wh_id = request.POST.get('webhook_id')
            wh = WebhookEndpoint.objects.filter(id=wh_id).first()
            if wh:
                wh.is_active = not wh.is_active
                wh.save()
            return redirect('core:manage_webhooks')

    webhooks = WebhookEndpoint.objects.select_related('created_by').order_by('-created_at')
    return render(request, 'core/dashboard/manage_webhooks.html', {
        'webhooks': webhooks,
    })


@_admin_required
def api_docs(request):
    """Render interactive API documentation page."""
    docs = [
        {
            'method': 'GET', 'path': '/api/students/',
            'desc': 'List all enrolled students. Optionally filter by student_id.',
            'params': 'student_id (int, optional)',
        },
        {
            'method': 'GET', 'path': '/api/results/',
            'desc': 'List all stored results. Optionally filter by student, subject, or terminal.',
            'params': 'student_id (int, optional), subject_id (int, optional), terminal (str, optional)',
        },
        {
            'method': 'GET', 'path': '/api/results/<id>/',
            'desc': 'Get a single result record with full details.',
            'params': 'id (int)',
        },
        {
            'method': 'GET', 'path': '/api/student-notifications/',
            'desc': 'Fetch notifications for the current logged-in student.',
            'params': 'unread_only (bool, GET param)',
        },
        {
            'method': 'POST', 'path': '/api/mark-notification-read/',
            'desc': 'Mark a notification as read.',
            'params': 'notification_id (int), student_id (int)',
        },
        {
            'method': 'POST', 'path': '/api/search/',
            'desc': 'Global search across students, teachers, and subjects.',
            'params': 'q (str, min 2 chars)',
        },
        {
            'method': 'GET', 'path': '/developer/api-keys/',
            'desc': 'List all API keys. Admin only.',
            'params': 'none',
        },
        {
            'method': 'POST', 'path': '/developer/webhooks/',
            'desc': 'Register a new webhook endpoint.',
            'params': 'url, events (comma-separated)',
        },
    ]
    return render(request, 'core/dashboard/api_docs.html', {'docs': docs})


def api_key_verify(request):
    """Public GET: test if an API key is valid and return its scopes."""
    raw_key = request.GET.get('key', '')
    if not raw_key:
        return JsonResponse({'valid': False, 'message': 'No key provided'}, status=400)
    k = APIKey.objects.filter(key=raw_key, is_active=True).first()
    if not k:
        return JsonResponse({'valid': False, 'message': 'Key not found or inactive'})
    expired = k.expires_at and (k.expires_at < timezone.now())
    if expired:
        return JsonResponse({'valid': False, 'message': 'Key expired', 'echo': k.prefix + '...'})
    k.last_used_at = timezone.now()
    k.save(update_fields=['last_used_at'])
    return JsonResponse({'valid': True, 'name': k.name, 'scopes': k.scopes, 'prefix': k.prefix})


def webhook_deliver(request):
    """Public POST: dispatch an event to all matching active webhooks."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)
    event_type = request.POST.get('event_type', request.headers.get('X-Webhook-Event', ''))
    raw_payload = request.POST.get('payload', '')
    try:
        payload = json.loads(raw_payload) if raw_payload else {}
    except json.JSONDecodeError:
        payload = {'raw': raw_payload}
    hmac_sig  = request.headers.get('X-Hmac-Signature', '')
    webhooks  = WebhookEndpoint.objects.filter(is_active=True, events__contains=[event_type])
    logs = []
    for wh in webhooks:
        try:
            import requests
            body = json.dumps(payload)
            expected_sig = hmac.new(
                wh.secret.encode(), body.encode(), hashlib.sha256
            ).hexdigest()
            sig_valid = hmac.compare_digest(hmac_sig, expected_sig) if hmac_sig else True
            resp = requests.post(wh.url, data=body,
                                 headers={'Content-Type': 'application/json',
                                          'X-Hmac-Signature': expected_sig},
                                 timeout=5)
            WebhookDeliveryLog.objects.create(
                endpoint=wh, event_type=event_type, payload=payload,
                response_code=resp.status_code, response_body=resp.text[:1000],
                success=resp.status_code < 400,
            )
            wh.last_triggered_at = timezone.now()
            wh.save(update_fields=['last_triggered_at'])
            logs.append({'url': wh.url, 'success': True, 'code': resp.status_code})
        except Exception as exc:
            WebhookDeliveryLog.objects.create(
            endpoint=wh, event_type=event_type, payload=payload, success=False,
            response_body=str(exc)[:500],
            )
            logs.append({'url': wh.url, 'success': False, 'error': str(exc)[:200]})
    return JsonResponse({'dispatched': len(logs), 'results': logs})


# ═════════════════════════════════════════════════════════════════════════════
# A10 — GOOGLE / MICROSOFT / GITHUB SSO
# ═════════════════════════════════════════════════════════════════════════════

def sso_login(request, provider):
    """Initiate OAuth2 redirect to the given provider."""
    provider_obj = SSOProvider.objects.filter(name=provider, is_active=True).first()
    if not provider_obj or not provider_obj.client_id:
        messages.warning(request, f'SSO provider "{provider}" is not configured.')
        return redirect('core:login')
    # Stub: log and redirect
    activity_log_flat(request.user, 'login', f'SSO OAuth redirect initiated for {provider}')
    messages.info(request, f'SSO provider {provider} not configured. Set up OAuth credentials in the admin panel.')
    return redirect('core:login')


def sso_callback(request, provider):
    """Handle OAuth2 callback from provider."""
    provider_obj = SSOProvider.objects.filter(name=provider, is_active=True).first()
    if not provider_obj or not provider_obj.client_id:
        messages.warning(request, f'SSO provider "{provider}" not configured.')
        return redirect('core:login')
    activity_log_flat(request.user, 'login', f'SSO OAuth callback received for {provider}')
    messages.info(request, f'[{provider.upper()}] SSO not configured — add OAuth credentials first.')
    return redirect('core:login')


def sso_microsoft(request):
    provider = 'microsoft'
    return sso_login(request, provider)


def sso_microsoft_callback(request):
    provider = 'microsoft'
    return sso_callback(request, provider)


def sso_github(request):
    provider = 'github'
    return sso_login(request, provider)


def sso_github_callback(request):
    provider = 'github'
    return sso_callback(request, provider)


def sso_status(request):
    """JSON API: list which SSO providers are configured and active."""
    data = []
    for prov in SSOProvider.objects.all():
        configured = bool(prov.client_id and prov.client_secret)
        data.append({
            'name': prov.name,
            'type': prov.provider_type,
            'active': prov.is_active,
            'configured': configured,
        })
    return JsonResponse({'providers': data})


def activity_log_flat(user, action, description=''):
    """Fire-and-forget activity log helper used in views posterior to this definition."""
    try:
        from core.notification_service import NotificationService
        NotificationService.log_activity(user, action, description)
    except Exception:
        pass

