from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import Student, Subject, Result, Teacher
from .views import get_teacher_profile
import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Four terminal sheet names
TERMINAL_SHEETS = ['1st Terminal', '2nd Terminal', '3rd Terminal', 'Final Terminal']
TERMINAL_CODES = {
    '1st Terminal':   '1st',
    '2nd Terminal':   '2nd',
    '3rd Terminal':   '3rd',
    'Final Terminal': 'Final',
}

# Column headers used in every sheet
HEADERS = ['Roll/Student ID', 'Student Name', 'Subject(s)', 'Marks Obtained', 'Total Marks']


# ── Workbook helpers ──────────────────────────────────────────────────────────

def _apply_header_style(ws):
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    hdr_align    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    col_widths = [22, 26, 46, 22, 20]
    for col_idx, header in enumerate(HEADERS, start=1):
        cell       = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = hdr_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.row_dimensions[1].height = 28


def _apply_instruction_row(ws):
    inst_font = Font(italic=True, size=9, color="555555")
    instructions = [
        "Roll number or full name (used to find student)",
        "Student full name — optional, for human reference",
        'Comma-separated subjects, e.g.  Mathematics,English,Science',
        'Comma-separated marks in same order, e.g.  85,78,90',
        'Per-subject total, e.g.  100,100,100  —  one value applies to all',
    ]
    for col_idx, inst in enumerate(instructions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=inst)
        cell.font  = inst_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 56


def _add_sample_rows(ws, teacher):
    """Paint 2 illustrative filled-in rows so the teacher sees the expected pattern."""
    sample_font   = Font(italic=True, size=10, color="555555")
    sample_fill   = PatternFill(start_color="F0F0FF", end_color="F0F0FF", fill_type="solid")
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    subjs   = list(teacher.subjects.values_list('name', flat=True)[:5]) if teacher else []
    students = list(teacher.students.all().order_by('roll_number')[:2]) if teacher else []

    def _paint(row, roll, name, su, ma, to):
        vals = [roll, name, su, ma, to]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font   = sample_font
            cell.fill   = sample_fill
            cell.border = thin_border

    if subjs and students:
        _paint(3, students[0].roll_number or 'R001', students[0].name,
               ','.join(subjs[:3]), '85,78,92', ','.join(['100'] * min(3, len(subjs[:3]))))
        _paint(4, students[1].roll_number or 'R002', students[1].name,
               subjs[0], '88', '100')
    else:
        _paint(3, 'R001', 'Sample Student', 'Mathematics,English', '85,78', '100,100')
        _paint(4, 'R002', 'Another Student', 'Science', '92', '100')


def _build_sheet(wb, sheet_name, teacher):
    ws = wb.create_sheet(title=sheet_name)
    _apply_header_style(ws)
    _apply_instruction_row(ws)
    _add_sample_rows(ws, teacher)
    return ws


# ── Public views ─────────────────────────────────────────────────────────────

@login_required
@require_http_methods(["GET"])
def download_excel_template(request):
    """
    Generate and stream an .xlsx workbook with 4 sheets (one per terminal).
    Each sheet: Roll/Student ID | Student Name | Subject(s) | Marks Obtained | Total Marks
    """
    teacher = get_teacher_profile(request.user)

    wb = openpyxl.Workbook()
    # Remove the default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name in TERMINAL_SHEETS:
        _build_sheet(wb, sheet_name, teacher)

    buf = BytesIO()
    try:
        wb.save(buf)
    except Exception as exc:
        import traceback as tb
        traceback.print_exc()
        return HttpResponse(
            f'<h1>Excel Error</h1><p>{type(exc).__name__}: {exc}</p>'
            f'<pre>{tb.format_exc()}</pre>',
            content_type='text/html', status=500,
        )

    buf.seek(0)
    data = buf.read()

    response = HttpResponse(
        data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bulk_marks_upload_template.xlsx"'
    response['Content-Length'] = str(len(data))
    return response


@login_required
@require_http_methods(["GET", "POST"])
def upload_marks_excel(request):
    """
    GET  → renders upload page with teacher's class/subject info.
    POST → reads the uploaded Excel file, iterates all 4 terminal sheets,
           expands comma-separated subject/mark rows into individual Result records,
           and reports success/error counts back via Django messages.
    """
    teacher = get_teacher_profile(request.user)

    if request.method == 'GET':
        teacher_classes   = []
        teacher_subjects  = []

        if teacher:
            teacher_classes = sorted(
                set(teacher.students.values_list('student_class', flat=True)
                    .exclude(student_class__isnull=True)
                    .exclude(student_class__exact='')
                    .distinct())
            )
            teacher_subjects = sorted(
                set(teacher.subjects.values_list('name', flat=True))
            )

        return render(request, 'core/dashboard/upload_marks_excel.html', {
            'teacher_classes':  teacher_classes,
            'teacher_subjects': teacher_subjects,
        })

    # ── POST ────────────────────────────────────────────────────────────────
    terminal_from_form = request.POST.get('terminal', '').strip()
    class_filter       = request.POST.get('student_class', '').strip()

    if 'excel_file' not in request.FILES:
        messages.error(request, 'No file uploaded. Please select an Excel file.')
        return redirect('core:upload_marks_excel')

    excel_file = request.FILES['excel_file']
    if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
        return redirect('core:upload_marks_excel')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception:
        messages.error(request, 'Failed to read Excel file. Ensure it is a valid .xlsx file and not corrupted.')
        return redirect('core:upload_marks_excel')

    # ── Teacher permission sets ──────────────────────────────────────────────
    teacher_subject_ids = set(teacher.subjects.values_list('id', flat=True)) if teacher else set()
    if teacher:
        qs = teacher.students.all()
        if class_filter:
            qs = qs.filter(student_class=class_filter)
        teacher_student_ids = set(qs.values_list('id', flat=True))
    else:
        teacher_student_ids = set()

    # ── Row-level resolvers ──────────────────────────────────────────────────
    def resolve_student(identifier):
        if not identifier or str(identifier).strip().lower() in ('none', 'null', ''):
            return None
        ident = str(identifier).strip()
        try:
            return Student.objects.get(roll_number__iexact=ident)
        except Student.DoesNotExist:
            try:
                return Student.objects.get(name__iexact=ident)
            except Student.DoesNotExist:
                return Student.objects.filter(name__icontains=ident).first()

    def resolve_subject(name):
        key = name.lower().strip()
        try:
            return Subject.objects.get(name__iexact=key)
        except Subject.DoesNotExist:
            try:
                return Subject.objects.get(code__iexact=key)
            except Subject.DoesNotExist:
                return None

    # ── Process one sheet ─────────────────────────────────────────────────────
    def process_sheet(ws, terminal_code):
        """Return dict: success, updated, errors, skipped (-1 = wrong columns)."""
        res = {'success': 0, 'updated': 0, 'errors': [], 'skipped': 0}
        student_cache, subject_cache = {}, {}

        # ── Map headers (flexible, case-insensitive) ──
        col = {}
        for idx, cell in enumerate(ws[1]):
            if not cell.value:
                continue
            h = str(cell.value).strip().lower()
            if h in ('roll', 'roll number', 'roll_no', 'roll/student id'):
                col['roll']   = idx
            elif h in ('student name', 'name', 'student'):
                col['name']   = idx
            elif h in ('subject(s)', 'subject', 'subjects'):
                col['subject'] = idx
            elif h in ('marks obtained', 'marks', 'score', 'marks_obtained'):
                col['marks']  = idx
            elif h in ('total marks', 'total', 'total_marks'):
                col['total']  = idx

        if 'subject' not in col or 'marks' not in col:
            return {'success': 0, 'updated': 0, 'errors': [], 'skipped': -1}

        # ── Parse each row ───────────────────────────────────────────────────
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None or str(v).strip() == '' for v in row):
                continue

            raw_roll  = row[col.get('roll', -1)]     if 'roll'  in col else None
            raw_name  = row[col.get('name', -1)]     if 'name'  in col else None
            raw_subj  = row[col['subject']]
            raw_marks = row[col['marks']]
            raw_total = row[col.get('total', -1)]    if 'total' in col else None

            # ── Resolve student ───────────────────────────────────────────────
            student = None
            for ident in (raw_roll, raw_name):
                if ident is None:
                    continue
                ident_str = str(ident).strip()
                if ident_str.lower() in ('none', 'null', ''):
                    continue
                if ident_str not in student_cache:
                    student_cache[ident_str] = resolve_student(ident_str)
                if student_cache[ident_str]:
                    student = student_cache[ident_str]
                    break

            if not student:
                res['errors'].append(f'[{terminal_code}] Row {row_num}: Student "{raw_roll or raw_name}" not found')
                res['skipped'] += 1
                continue

            if teacher and student.id not in teacher_student_ids:
                res['errors'].append(f'[{terminal_code}] Row {row_num}: "{student.name}" is not assigned to you')
                res['skipped'] += 1
                continue

            if class_filter and student.student_class != class_filter:
                res['errors'].append(f'[{terminal_code}] Row {row_num}: "{student.name}" not in class {class_filter}')
                res['skipped'] += 1
                continue

            # ── Parse subject(s) (comma-separated) ───────────────────────────
            if raw_subj is None or str(raw_subj).strip().lower() in ('none', 'null', ''):
                res['errors'].append(f'[{terminal_code}] Row {row_num}: Subject is required')
                res['skipped'] += 1; continue
            subject_names = [s.strip() for s in str(raw_subj).split(',') if s.strip()]
            if not subject_names:
                res['errors'].append(f'[{terminal_code}] Row {row_num}: Invalid subject column')
                res['skipped'] += 1; continue

            # ── Parse marks (comma-separated) ────────────────────────────────
            if raw_marks is None or str(raw_marks).strip().lower() in ('none', 'null', ''):
                res['errors'].append(f'[{terminal_code}] Row {row_num}: Marks required')
                res['skipped'] += 1; continue
            marks_list = [m.strip() for m in str(raw_marks).split(',') if m.strip()]
            try:
                marks_vals = [float(m) for m in marks_list]
            except ValueError:
                res['errors'].append(f'[{terminal_code}] Row {row_num}: Non-numeric marks')
                res['skipped'] += 1; continue

            if len(marks_vals) != len(subject_names):
                res['errors'].append(
                    f'[{terminal_code}] Row {row_num}: {len(subject_names)} subject(s) but '
                    f'{len(marks_vals)} mark value(s) — counts must match'
                )
                res['skipped'] += 1; continue

            # ── Parse total marks (comma-separated or single) ─────────────────
            if raw_total is not None and str(raw_total).strip().lower() not in ('none', 'null', ''):
                parts = [t.strip() for t in str(raw_total).split(',') if t.strip()]
                if len(parts) == len(subject_names):
                    try:
                        total_vals = [float(t) for t in parts]
                    except ValueError:
                        total_vals = [100.0] * len(subject_names)
                elif len(parts) == 1:
                    try:
                        total_vals = [float(parts[0])] * len(subject_names)
                    except ValueError:
                        total_vals = [100.0] * len(subject_names)
                else:
                    total_vals = [100.0] * len(subject_names)
            else:
                total_vals = [100.0] * len(subject_names)

            # ── Create / update Result per subject ───────────────────────────
            row_errors = []
            for subj_name, mk, tot in zip(subject_names, marks_vals, total_vals):
                if mk < 0:
                    row_errors.append(f'Negative marks in {subj_name}')
                    continue
                if tot <= 0:
                    row_errors.append(f'Invalid total in {subj_name}')
                    continue

                key = subj_name.lower().strip()
                if key not in subject_cache:
                    subject_cache[key] = resolve_subject(subj_name)
                subject = subject_cache[key]
                if not subject:
                    row_errors.append(f'Subject "{subj_name}" not found')
                    continue
                if teacher and subject.id not in teacher_subject_ids:
                    row_errors.append(f'Not authorised for {subject.name}')
                    continue

                existing = Result.objects.filter(
                    student=student, subject=subject, terminal=terminal_code
                ).first()
                if existing:
                    existing.marks_obtained = mk
                    existing.total_marks   = tot
                    existing.save()
                    res['updated'] += 1
                else:
                    Result.objects.create(
                        student=student, subject=subject, terminal=terminal_code,
                        marks_obtained=mk, total_marks=tot,
                    )
                    res['success'] += 1

            if row_errors:
                res['errors'].append(f'[{terminal_code}] Row {row_num}: {"; ".join(row_errors)}')
                res['skipped'] += 1

        return res

    # ── Iterate all terminal sheets ──────────────────────────────────────────
    all_success = all_updated = 0
    all_errors  = []

    sheets_found = {s: TERMINAL_CODES[s] for s in TERMINAL_SHEETS if s in wb.sheetnames}
    if not sheets_found:
        messages.error(request, 'The file has no recognised terminal sheets.')
        return redirect('core:upload_marks_excel')

    for sheet_name, tcode in sheets_found.items():
        result = process_sheet(wb[sheet_name], tcode)
        if result['skipped'] == -1:
            continue  # sheet has no expected headers — silently skip
        all_success += result['success']
        all_updated += result['updated']
        all_errors.extend(result['errors'])

    # ── Feedback ──────────────────────────────────────────────────────────────
    total = all_success + all_updated
    if total > 0:
        parts = [f'{all_success} added' if all_success else '',
                 f'{all_updated} updated' if all_updated else '']
        messages.success(request, f'Bulk upload complete: {", ".join(p for p in parts if p)}.')
    else:
        messages.warning(request, 'No marks were saved. Check each sheet in the file.')

    if all_errors:
        shown   = all_errors[:12]
        details = '<br>• '.join(shown)
        suffix  = f'<br>• ... and {len(all_errors) - 12} more' if len(all_errors) > 12 else ''
        messages.warning(request, f'{len(all_errors)} issue(s):<br>• {details}{suffix}')

    return redirect('core:upload_marks_excel')
