"""
A8 — Analytics Export Bundle
Generates PDF and Excel analytics reports from smart_dashboard context.
"""
import io
import logging
from datetime import date
from django.db.models import Sum, Avg
from django.http import HttpResponse


logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def style_header(ws, headers):
    """Apply bold header styling to the first row of an openpyxl worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill("solid", fgColor="4F46E5")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = __import__('openpyxl').styles.Border(
        left=__import__('openpyxl').styles.Side(style='thin'),
        right=__import__('openpyxl').styles.Side(style='thin'),
        top=__import__('openpyxl').styles.Side(style='thin'),
        bottom=__import__('openpyxl').styles.Side(style='thin'),
    )
    ws.row_dimensions[1].height = 20
    for col, val in enumerate(headers, 1):
        cell = ws.cell(1, col, val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin


def auto_width(ws):
    """Auto-fit column widths in an openpyxl worksheet."""
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value or ''))
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 40)


def _get_styles():
    from openpyxl.styles import Font, PatternFill
    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill("solid", fgColor="4F46E5")
    return hf, hfill


def export_smart_dashboard_excel(request, context):
    """
    Generate an Excel workbook with 7 sheets matching the smart_dashboard sections:
    1. KPI Overview  2. Top Students  3. Weak Students
    4. Subject Stats  5. Class Stats  6. Terminal Stats  7. Attendance
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_students   = context.get('total_students', 0)
    total_subjects   = context.get('total_subjects', 0)
    total_results    = context.get('total_results', 0)
    total_teachers   = context.get('total_teachers', 0)
    pass_rate        = context.get('pass_rate', 0)
    pass_fail_ratio  = context.get('pass_fail_ratio', '0:0')
    top_students     = context.get('top_students', [])
    weak_students    = context.get('weak_students', [])
    subj_stats       = context.get('subj_stats', [])
    class_stats      = context.get('class_stats', [])
    term_stats       = context.get('term_stats', [])
    att_present      = context.get('att_present', 0)
    att_total        = context.get('att_total', 0)
    att_pct          = context.get('att_pct', 0)
    pass_mark_pct    = context.get('pass_mark_pct', 40.0)
    selected_class   = context.get('selected_class', '')
    overall_pct      = context.get('overall_percentage', 0)
    overall_avg      = context.get('overall_average', 0)
    pass_count       = context.get('pass_count', 0)
    fail_count       = context.get('fail_count', 0)

    hf, hfill = _get_styles()
    thin_border = __import__('openpyxl').styles.Border(
        left=__import__('openpyxl').styles.Side(style='thin'),
        right=__import__('openpyxl').styles.Side(style='thin'),
        top=__import__('openpyxl').styles.Side(style='thin'),
        bottom=__import__('openpyxl').styles.Side(style='thin'),
    )
    alt_fill  = PatternFill("solid", fgColor="F8FAFC")
    alt2_fill = PatternFill("solid", fgColor="F1F5F9")
    red2_fill = PatternFill("solid", fgColor="FEF2F2")

    def _write_sheet(wb, title, headers, rows, col_widths=None, alt=None):
        ws = wb.create_sheet(title)
        for col, val in enumerate(headers, 1):
            cell = ws.cell(1, col, val)
            cell.font  = hf
            cell.fill  = hfill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 20
        for row_i, row_data in enumerate(rows, 2):
            fill = alt[row_i % len(alt)] if alt else (alt_fill if row_i % 2 == 0 else None)
            for col_i, val in enumerate(row_data, 1):
                cell = ws.cell(row_i, col_i, val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if fill:
                    cell.fill = fill
        if col_widths:
            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[__import__('openpyxl').utils.get_column_letter(ci)].width = w

    # Sheet 1 – KPI
    _write_sheet(wb, 'KPI Overview',
        ['Metric', 'Value'],
        [
            ['Selected Class', selected_class or 'All Classes'],
            ['Total Students', total_students],
            ['Total Subjects', total_subjects],
            ['Total Teachers', total_teachers],
            ['Total Results',  total_results],
            ['Overall Average', overall_avg],
            ['Overall %', f'{overall_pct}%'],
            ['Pass Count', pass_count],
            ['Fail Count', fail_count],
            ['Pass Rate', f'{pass_rate}%'],
            ['Pass : Fail Ratio', pass_fail_ratio],
            ['Pass Mark Threshold', f'{pass_mark_pct}%'],
            ['Attendance %', f'{att_pct}%'],
            ['Attendance Present / Total', f'{att_present} / {att_total}'],
        ], col_widths=[28, 20], alt=[alt_fill])

    # Sheet 2 – Top Students
    _write_sheet(wb, 'Top Students',
        ['Rank', 'Name', 'Class', 'Section', 'Percentage %', 'Marks Obtained', 'Total Marks'],
        [[r, s['student'].name, s['student'].student_class, s['student'].section,
          s['percentage'], s.get('obtained', 0), s.get('total', 0)]
          for r, s in enumerate(top_students, 1)],
        col_widths=[8, 28, 14, 12, 16, 18, 18], alt=[alt_fill, alt2_fill])

    # Sheet 3 – Weak Students
    _write_sheet(wb, 'Weak Students',
        ['Rank', 'Name', 'Class', 'Section', 'Percentage %', 'Marks Obtained', 'Total Marks'],
        [[r, s['student'].name, s['student'].student_class, s['student'].section,
          s['percentage'], s.get('obtained', 0), s.get('total', 0)]
          for r, s in enumerate(weak_students, 1)],
        col_widths=[8, 28, 14, 12, 16, 18, 18], alt=[red2_fill, alt_fill])

    # Sheet 4 – Subject Stats
    _write_sheet(wb, 'Subject Stats',
        ['Subject', 'Average Marks', 'Percentage %', 'Status'],
        [[ss['subject'], ss['avg'], ss['percentage'],
          'PASS' if ss['percentage'] >= pass_mark_pct else 'BELOW']
          for ss in subj_stats],
        col_widths=[30, 18, 18, 12], alt=[alt_fill, alt2_fill])

    # Sheet 5 – Class Stats
    _write_sheet(wb, 'Class Stats',
        ['Class', 'Students', 'Percentage %', 'Status'],
        [[cs['class'], cs['students'], cs['percentage'],
          'PASS' if cs['percentage'] >= pass_mark_pct else 'BELOW']
          for cs in class_stats],
        col_widths=[14, 14, 18, 12], alt=[alt_fill, alt2_fill])

    # Sheet 6 – Terminal Stats
    _write_sheet(wb, 'Terminal Stats',
        ['Terminal', 'Percentage %', 'Status'],
        [[ts['terminal'], ts['percentage'],
          'PASS' if ts['percentage'] >= pass_mark_pct else 'BELOW']
          for ts in term_stats],
        col_widths=[14, 18, 12], alt=[alt_fill, alt2_fill])

    # Sheet 7 – Attendance Summary
    _write_sheet(wb, 'Attendance Summary',
        ['Metric', 'Value'],
        [['Attendance %', f'{att_pct}%'],
         ['Present Count',              att_present],
         ['Total Count',                att_total],
         ['Absent / Excused / Late',    att_total - att_present]],
        col_widths=[28, 20], alt=[alt_fill])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    selected_class_label = (selected_class.replace(' ', '_') if selected_class else 'all')
    resp['Content-Disposition'] = f'attachment; filename="analytics_{selected_class_label}_{date.today()}.xlsx"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_smart_dashboard_pdf(request, context=None):
    """
    Generate a PDF report using ReportLab based on smart_dashboard context.
    Falls back to xhtml2pdf HTML rendering if ReportLab fails.
    """
    from core.models import Student, Subject, Teacher, Result, Attendance, GradeScale

    if context is None:
        all_results = Result.objects.select_related('student', 'subject')
        subj_list   = list(Subject.objects.all())
        cls_list    = list(Student.objects.values_list('student_class', flat=True).distinct().order_by('student_class'))

        gs = GradeScale.objects.filter(is_active=True).first()
        pass_mark = gs.pass_mark_percent if gs else 40.0

        student_scores = []
        for stu in Student.objects.prefetch_related('result_set'):
            res = stu.result_set.all()
            if not res.exists():
                continue
            obt = sum(r.marks_obtained for r in res)
            pos = sum(r.total_marks for r in res)
            p = (obt / pos * 100) if pos > 0 else 0
            student_scores.append({'student': stu, 'percentage': round(p, 1)})
        student_scores.sort(key=lambda x: x['percentage'], reverse=True)
        top  = student_scores[:10]
        weak = [s for s in student_scores if s['percentage'] < pass_mark][:10]

        subj_stats = []
        for subj in subj_list:
            s_res = all_results.filter(subject=subj)
            if s_res.exists():
                avg = s_res.aggregate(av=Avg('marks_obtained'))['av'] or 0
                pos_sum = s_res.aggregate(s=Sum('total_marks'))['s'] or 0
                obt_sum = s_res.aggregate(s=Sum('marks_obtained'))['s'] or 0
                pct = (obt_sum / pos_sum * 100) if pos_sum > 0 else 0
                subj_stats.append({'subject': subj.name, 'avg': round(avg,1), 'percentage': round(pct,1)})
        subj_stats.sort(key=lambda x: x['percentage'])

        context = {
            'total_students':   Student.objects.count(),
            'total_subjects':   len(subj_list),
            'total_results':    all_results.count(),
            'total_teachers':   Teacher.objects.filter(is_active=True).count(),
            'top_students':     top,
            'weak_students':    weak,
            'subj_stats':       subj_stats,
            'att_present':      Attendance.objects.filter(status='present').count(),
            'att_total':        Attendance.objects.count(),
        }
        p_total  = context['att_total'] or 1
        context['att_pct']       = round((context['att_present'] / p_total) * 100, 1)
        context['pass_mark_pct'] = pass_mark

    try:
        return _pdf_reportlab(context, request)
    except Exception as exc:
        logger.warning("ReportLab PDF generation failed: %s — falling back to HTML", exc)
        return _pdf_html_fallback(context, request)


def _pdf_reportlab(context, request):
    """Build a PDF using ReportLab's canvas."""
    from django.http import FileResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        title='Smart Analytics Report',
    )
    styles   = getSampleStyleSheet()
    accent   = colors.HexColor('#4F46E5')
    grey     = colors.grey

    def P(text, style_name='Normal', size=10, color=None, bold=False, align=TA_CENTER):
        color = color or colors.HexColor('#1e293b')
        st = ParagraphStyle(
            '_tmp', parent=styles.get(style_name, styles['Normal']),
            fontSize=size, textColor=color, alignment=align, spaceAfter=4,
        )
        if bold:
            st.fontName = 'Helvetica-Bold'
        return Paragraph(text, st)

    elements = []
    elements.append(P('ACADEMIC SMART ANALYTICS REPORT', 'Title', size=18, color=accent, bold=True))
    elements.append(P('AcadStat \u2014 Academic Management System', 'Normal', size=9, color=grey))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(HRFlowable(width='100%', thickness=2, color=accent, spaceAfter=0.4*cm))

    # KPI table
    hdr_row = [P('Total Students','Normal',9,color=colors.white,bold=True),
               P('Total Subjects','Normal',9,color=colors.white,bold=True),
               P('Total Results', 'Normal',9,color=colors.white,bold=True),
               P('Total Teachers','Normal',9,color=colors.white,bold=True)]
    val_row = [P(str(context.get('total_students',0)),  'Normal',16,bold=True),
               P(str(context.get('total_subjects',0)),  'Normal',16,bold=True),
               P(str(context.get('total_results',0)),   'Normal',16,bold=True),
               P(str(context.get('total_teachers',0)),  'Normal',16,bold=True)]
    kpi_data = [hdr_row, val_row]

    kpi_table = Table(kpi_data, colWidths=[4*cm]*4)
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), accent),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#EEF2FF')),
        ('GRID',      (0,0), (-1,-1), 0.5, colors.white),
        ('ALIGN',     (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',    (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.5*cm))

    # Attendance
    att_pct   = context.get('att_pct', 0)
    att_total = context.get('att_total', 0)
    att_pres  = context.get('att_present', 0)
    elements.append(P(f'Attendance: {att_pct}%', 'Normal', 12, bold=True, align=TA_CENTER))
    att_data = [['Present', 'Total', 'Others']]
    for val in [att_pres, att_total, att_total - att_pres]:
        att_data.append([str(val)])
    att_table = Table(att_data, colWidths=[4*cm, 4*cm, 4*cm])
    att_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), accent),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('GRID',       (0,0), (-1,-1), 0.5, grey),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    elements.append(att_table)
    elements.append(Spacer(1, 0.5*cm))

    # Top Students
    top = context.get('top_students', [])
    if top:
        elements.append(P('TOP PERFORMERS', 'Normal', 12, bold=True, color=accent))
        top_data = [[P('Rank','Normal',9,color=colors.white,bold=True),
                     P('Name','Normal',9,color=colors.white,bold=True),
                     P('Class','Normal',9,color=colors.white,bold=True),
                     P('%','Normal',9,color=colors.white,bold=True)]]
        for r, s in enumerate(top, 1):
            top_data.append([P(str(r)), P(s['student'].name), P(s['student'].student_class), P(str(s['percentage']))])
        tt = Table(top_data, colWidths=[2*cm, 6*cm, 4*cm, 4*cm])
        tt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), accent),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F1F5F9')]),
            ('GRID',(0,0),(-1,-1),0.5,grey),('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        elements.append(tt)
        elements.append(Spacer(1, 0.5*cm))

    # Weak Students
    weak = context.get('weak_students', [])
    if weak:
        elements.append(P('WEAK STUDENTS (Below Pass Mark)', 'Normal', 12, bold=True, color=colors.HexColor('#ef4444')))
        wdata = [[P('Rank','Normal',9,color=colors.white,bold=True),
                  P('Name','Normal',9,color=colors.white,bold=True),
                  P('Class','Normal',9,color=colors.white,bold=True),
                  P('%','Normal',9,color=colors.white,bold=True)]]
        for r, s in enumerate(weak, 1):
            wdata.append([P(str(r)), P(s['student'].name), P(s['student'].student_class), P(str(s['percentage']))])
        wt = Table(wdata, colWidths=[2*cm, 6*cm, 4*cm, 4*cm])
        wt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#ef4444')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#FEF2F2')]),
            ('GRID',(0,0),(-1,-1),0.5,grey),('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        elements.append(wt)
        elements.append(Spacer(1, 0.5*cm))

    # Subject Stats
    ss = context.get('subj_stats', [])
    if ss:
        elements.append(P('SUBJECT-WISE PERFORMANCE', 'Normal', 12, bold=True, color=accent))
        sdata = [[P('Subject',9,color=colors.white,bold=True), P('Avg',9,color=colors.white,bold=True), P('%',9,color=colors.white,bold=True)]]
        for s in ss:
            sdata.append([P(s['subject'].__str__() if s['subject'] else str(s['subject'])), P(str(s['avg'])), P(str(s['percentage']))])
        st = Table(sdata, colWidths=[6*cm, 6*cm, 4*cm])
        st.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),accent),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F1F5F9')]),
            ('GRID',(0,0),(-1,-1),0.5,grey),('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        elements.append(st)
        elements.append(Spacer(1, 0.5*cm))

    # Class Stats
    cs = context.get('class_stats', [])
    if cs:
        elements.append(P('CLASS-WISE PERFORMANCE', 'Normal', 12, bold=True, color=accent))
        cdata = [[P('Class',9,color=colors.white,bold=True), P('Students',9,color=colors.white,bold=True), P('%',9,color=colors.white,bold=True)]]
        for c in cs:
            cdata.append([P(str(c['class'])), P(str(c['students'])), P(str(c['percentage']))])
        ct = Table(cdata, colWidths=[6*cm, 6*cm, 4*cm])
        ct.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),accent),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F1F5F9')]),
            ('GRID',(0,0),(-1,-1),0.5,grey),('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        elements.append(ct)
        elements.append(Spacer(1, 0.5*cm))

    # Terminal Stats
    ts = context.get('term_stats', [])
    if ts:
        elements.append(P('TERMINAL-WISE PERFORMANCE', 'Normal', 12, bold=True, color=accent))
        tdata = [[P('Terminal',9,color=colors.white,bold=True), P('%',9,color=colors.white,bold=True)]]
        for t in ts:
            tdata.append([P(t['terminal']), P(str(t['percentage']))])
        tt = Table(tdata, colWidths=[8*cm, 8*cm])
        tt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),accent),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F1F5F9')]),
            ('GRID',(0,0),(-1,-1),0.5,grey),('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        elements.append(tt)
        elements.append(Spacer(1, 0.5*cm))

    # Footer
    elements.append(HRFlowable(width='100%', thickness=1, color=grey, spaceBefore=0.5*cm))
    elements.append(P(f'Generated on {date.today().strftime("%B %d, %Y")} \u2014 AcadStat Academic System',
                      'Normal', 8, color=grey))

    doc.build(elements)
    buf.seek(0)
    return FileResponse(buf, as_attachment=True,
                        filename=f'smart_analytics_report_{date.today()}.pdf')


def _pdf_html_fallback(context, request):
    """Return a rich HTML page the user can print→Save as PDF when ReportLab isn't available."""
    from django.http import HttpResponse as _Resp
    from django.template.loader import render_to_string
    html = render_to_string('core/dashboard/smart_dashboard_pdf.html', context)
    resp = _Resp(html, content_type='text/html; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="analytics_report_{date.today()}.html"'
    return resp
